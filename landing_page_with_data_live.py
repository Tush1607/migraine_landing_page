import streamlit as st
import streamlit.components.v1 as components

st.set_page_config(
    page_title="Migraine Intelligence Hub",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# --- Data (cached from Snowflake via DSS) ---
import plotly.graph_objects as go
from datetime import datetime

# --- User Access Restriction (resolved after finance data loads) ---
def _get_current_user_email():
    """Get current VIEWER email - tries ALL methods and reports results."""
    _debug_info = []
    _candidates = {}

    # Method 1: Dataiku get_auth_info
    try:
        import dataiku
        client = dataiku.api_client()
        auth_info = client.get_auth_info()
        m1_user = auth_info.get('authIdentifier', '')
        m1_assoc = auth_info.get('associatedDSSUser', '')
        m1_imperson = auth_info.get('userForImpersonation', '')
        _debug_info.append(f"Method 1 (get_auth_info) - authIdentifier: '{m1_user}', associatedDSSUser: '{m1_assoc}', userForImpersonation: '{m1_imperson}'")
        _debug_info.append(f"  Full auth_info: {auth_info}")
        _candidates['get_auth_info'] = m1_user
    except Exception as ex:
        _debug_info.append(f"Method 1 (get_auth_info) FAILED: {ex}")

    # Method 2: Dataiku webapp get_webapp_config
    try:
        import dataiku
        from dataiku.webapp import get_webapp_config
        webapp_config = get_webapp_config()
        _debug_info.append(f"Method 2 (webapp_config) - config: {webapp_config}")
    except Exception as ex:
        _debug_info.append(f"Method 2 (webapp_config) FAILED: {ex}")

    # Method 3: Streamlit headers
    try:
        headers = st.context.headers
        all_headers = dict(headers)
        dku_user = headers.get('X-DKU-User', '') or headers.get('x-dku-user', '')
        remote_user = headers.get('X-Remote-User', '') or headers.get('x-remote-user', '')
        _debug_info.append(f"Method 3 (st.context.headers) - X-DKU-User: '{dku_user}', X-Remote-User: '{remote_user}'")
        _debug_info.append(f"  All headers: {all_headers}")
        if dku_user:
            _candidates['headers_dku'] = dku_user
        if remote_user:
            _candidates['headers_remote'] = remote_user
    except Exception as ex:
        _debug_info.append(f"Method 3 (st.context.headers) FAILED: {ex}")

    # Method 4: Streamlit experimental_user
    try:
        exp_email = st.experimental_user.email or ''
        _debug_info.append(f"Method 4 (st.experimental_user) - email: '{exp_email}'")
        if exp_email:
            _candidates['experimental_user'] = exp_email
    except Exception as ex:
        _debug_info.append(f"Method 4 (st.experimental_user) FAILED: {ex}")

    # Method 5: Env vars
    try:
        import os
        m5_dku = os.environ.get('DKU_CURRENT_USER', '')
        m5_dataiku = os.environ.get('DATAIKU_USER', '')
        _debug_info.append(f"Method 5 (env vars) - DKU_CURRENT_USER: '{m5_dku}', DATAIKU_USER: '{m5_dataiku}'")
        if m5_dku:
            _candidates['env_dku'] = m5_dku
        if m5_dataiku:
            _candidates['env_dataiku'] = m5_dataiku
    except Exception as ex:
        _debug_info.append(f"Method 5 (env vars) FAILED: {ex}")

    # Method 6: Dataiku get_own_user
    try:
        import dataiku
        client = dataiku.api_client()
        user = client.get_own_user()
        settings = user.get_settings().get_raw()
        m6_email = settings.get('email', '')
        m6_login = settings.get('login', '')
        _debug_info.append(f"Method 6 (get_own_user) - email: '{m6_email}', login: '{m6_login}'")
        _candidates['get_own_user_email'] = m6_email
        _candidates['get_own_user_login'] = m6_login
    except Exception as ex:
        _debug_info.append(f"Method 6 (get_own_user) FAILED: {ex}")

    _debug_info.append(f"ALL CANDIDATES: {_candidates}")

    # Pick best: prefer headers > experimental_user > env > auth_info > get_own_user
    chosen = (_candidates.get('headers_dku') or _candidates.get('headers_remote')
              or _candidates.get('experimental_user')
              or _candidates.get('env_dku') or _candidates.get('env_dataiku')
              or _candidates.get('get_auth_info')
              or _candidates.get('get_own_user_email') or _candidates.get('get_own_user_login') or '')
    _debug_info.append(f"CHOSEN: '{chosen}'")
    return chosen, _debug_info

_current_user_email, _user_debug_info = _get_current_user_email()

def _resolve_finance_restriction():
    """Dynamically build restricted email list from data + hardcoded additions."""
    restricted = ['Harshit.Gupta@pfizer.com', 'Dhruv.Bajpai@pfizer.com']
    try:
        fin_df = load_finance_stacked()
        if 'USER_EMAIL_RESTRICT' in fin_df.columns:
            raw = fin_df['USER_EMAIL_RESTRICT'].dropna().iloc[0] if len(fin_df) > 0 else ''
            if raw:
                restricted += [e.strip() for e in str(raw).split(';') if e.strip()]
    except:
        pass
    # Normalize: compare lowercase, and also match just the prefix before @ 
    user_email = _current_user_email.lower().strip()
    user_prefix = user_email.split('@')[0] if '@' in user_email else user_email
    for r in restricted:
        r_lower = r.lower().strip()
        r_prefix = r_lower.split('@')[0] if '@' in r_lower else r_lower
        if user_email == r_lower or user_prefix == r_prefix:
            return True
    return False

def week_to_date(wid):
    return datetime.strptime(str(int(wid)), '%Y%m%d')

import pandas as pd

from data_queries_live import (
    fetch_brand_data, fetch_channel_data, fetch_npa_stacked,
    fetch_xponent_trends, fetch_xponent_stacked,
    fetch_finance_trends, fetch_finance_stacked,
    fetch_data_refresh
)

@st.cache_data(ttl=3600)
def load_brand_data(segment):
    return fetch_brand_data(segment=segment, rx_classification="Overall", channel_type="Overall")

@st.cache_data(ttl=3600)
def load_channel_data(segment, brand):
    return fetch_channel_data(segment=segment, rx_classification="Overall", brand=brand)

@st.cache_data(ttl=3600)
def load_acute_prev_brand_data(segment, rx_class):
    return fetch_brand_data(segment=segment, rx_classification=rx_class, channel_type="Overall")

@st.cache_data(ttl=3600)
def load_acute_prev_channel_data(segment, rx_class, brand):
    return fetch_channel_data(segment=segment, rx_classification=rx_class, brand=brand)

@st.cache_data(ttl=3600)
def load_xponent_trends():
    return fetch_xponent_trends()

@st.cache_data(ttl=3600)
def load_xponent_stacked():
    return fetch_xponent_stacked()

@st.cache_data(ttl=3600)
def load_finance_trends():
    return fetch_finance_trends()

@st.cache_data(ttl=3600)
def load_finance_stacked():
    return fetch_finance_stacked()

@st.cache_data(ttl=3600)
def load_data_refresh():
    return fetch_data_refresh()

# --- NPA Overall Brand Data ---
npa_brand_df = load_brand_data("TRx")
weeks = sorted(npa_brand_df['WEEK_ID'].unique())
nurtec_data = npa_brand_df[npa_brand_df['BRAND'] == 'NURTEC'].sort_values('WEEK_ID')
ubrelvy_data = npa_brand_df[npa_brand_df['BRAND'] == 'UBRELVY'].sort_values('WEEK_ID')
qulipta_data = npa_brand_df[npa_brand_df['BRAND'] == 'QULIPTA'].sort_values('WEEK_ID')

nbrx_brand_df = load_brand_data("NBRx")
nbrx_nurtec = nbrx_brand_df[nbrx_brand_df['BRAND'] == 'NURTEC'].sort_values('WEEK_ID')
nbrx_ubrelvy = nbrx_brand_df[nbrx_brand_df['BRAND'] == 'UBRELVY'].sort_values('WEEK_ID')
nbrx_qulipta = nbrx_brand_df[nbrx_brand_df['BRAND'] == 'QULIPTA'].sort_values('WEEK_ID')

# --- NPA Overall Channel Data ---
def get_channel_dict(segment):
    result = {}
    for brand in ['NURTEC', 'UBRELVY', 'QULIPTA']:
        ch_df = load_channel_data(segment, brand)
        retail_df = ch_df[ch_df['CHANNEL_TYPE'] == 'Retail'].sort_values('WEEK_ID')
        mail_df = ch_df[ch_df['CHANNEL_TYPE'] == 'MAIL'].sort_values('WEEK_ID')
        ltc_df = ch_df[ch_df['CHANNEL_TYPE'] == 'Long term'].sort_values('WEEK_ID')
        result[brand] = {
            "Retail": {"actuals": list(retail_df['ACTUALS']), "stly": list(retail_df['STLY'])},
            "Mail": {"actuals": list(mail_df['ACTUALS']), "stly": list(mail_df['STLY'])},
            "LTC": {"actuals": list(ltc_df['ACTUALS']) if len(ltc_df) > 0 else [0]*len(retail_df), "stly": list(ltc_df['STLY']) if len(ltc_df) > 0 else [0]*len(retail_df)}
        }
    return result

_channel_trx_data = get_channel_dict("TRx")
_channel_nbrx_data = get_channel_dict("NBRx")

# --- Acute/Preventive Brand Data ---
_acute_trx_df = load_acute_prev_brand_data("TRx", "Acute")
_acute_nbrx_df = load_acute_prev_brand_data("NBRx", "Acute")
_prev_trx_df = load_acute_prev_brand_data("TRx", "Preventive")
_prev_nbrx_df = load_acute_prev_brand_data("NBRx", "Preventive")

# --- Xponent Trends Data (Live) ---
_xpt_df = load_xponent_trends()
_xpt_weeks = sorted(_xpt_df['WEEK_ID'].unique())
_xpt_dates = [week_to_date(w) for w in _xpt_weeks]

def _get_xpt_series(df, cut_type, cut_value, prescription):
    sub = df[(df['CUT_TYPE'] == cut_type) & (df['CUT_VALUE'] == cut_value) & (df['PRESCRIPTION'] == prescription)].sort_values('WEEK_ID')
    return list(sub['SHARE_PCT'])

_xpt_payer_trx = {k: _get_xpt_series(_xpt_df, 'Payer', k, 'TRx') for k in ['Commercial', 'Medicare', 'Medicaid', 'Other']}
_xpt_payer_nrx = {k: _get_xpt_series(_xpt_df, 'Payer', k, 'NRx') for k in ['Commercial', 'Medicare', 'Medicaid', 'Other']}
_xpt_ch_trx = {'Retail': _get_xpt_series(_xpt_df, 'Channel', 'RETAIL', 'TRx'), 'Mail-Order': _get_xpt_series(_xpt_df, 'Channel', 'MAIL_ORDER', 'TRx'), 'LTC': _get_xpt_series(_xpt_df, 'Channel', 'LTC', 'TRx')}
_xpt_ch_nrx = {'Retail': _get_xpt_series(_xpt_df, 'Channel', 'RETAIL', 'NRx'), 'Mail-Order': _get_xpt_series(_xpt_df, 'Channel', 'MAIL_ORDER', 'NRx'), 'LTC': _get_xpt_series(_xpt_df, 'Channel', 'LTC', 'NRx')}

# --- Xponent Trends Excel downloads (base64-encoded) ---
import io, base64
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def _build_xpt_excel(dates, data_dict, sheet_name):
    buf = io.BytesIO()
    df = pd.DataFrame({'Week': [d.strftime('%Y-%m-%d') for d in dates]})
    for k, v in data_dict.items():
        df[k] = v
    df = df.sort_values('Week', ascending=False).reset_index(drop=True)
    df.to_excel(buf, index=False, sheet_name=sheet_name, engine='openpyxl')
    buf.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(buf)
    ws = wb.active
    header_fill = PatternFill(start_color='0000C9', end_color='0000C9', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return base64.b64encode(out.read()).decode()

_xpt_dl_payer_trx_b64 = _build_xpt_excel(_xpt_dates, _xpt_payer_trx, 'TRx Share by Payer')
_xpt_dl_payer_nrx_b64 = _build_xpt_excel(_xpt_dates, _xpt_payer_nrx, 'NRx Share by Payer')
_xpt_dl_ch_trx_b64 = _build_xpt_excel(_xpt_dates, _xpt_ch_trx, 'TRx Share by Channel')
_xpt_dl_ch_nrx_b64 = _build_xpt_excel(_xpt_dates, _xpt_ch_nrx, 'NRx Share by Channel')

# --- NPA Market Insights Excel downloads ---
_npa_dates = [week_to_date(w) for w in weeks]

def _build_npa_brand_excel(brand_df, metric, sheet_name):
    dates_col = [week_to_date(w).strftime('%Y-%m-%d') for w in sorted(brand_df['WEEK_ID'].unique())]
    data = {'Week': dates_col}
    for brand in ['NURTEC', 'UBRELVY', 'QULIPTA']:
        bdf = brand_df[brand_df['BRAND'] == brand].sort_values('WEEK_ID')
        data[f'{brand} Actuals'] = list(bdf['ACTUALS'])
        data[f'{brand} STLY'] = list(bdf['STLY'])
    return _build_xpt_excel(_npa_dates, {k: v for k, v in data.items() if k != 'Week'}, sheet_name)

def _build_npa_channel_excel(channel_dict, brand, metric, sheet_name):
    ch = channel_dict[brand]
    data = {
        'Retail Actuals': ch['Retail']['actuals'], 'Retail STLY': ch['Retail']['stly'],
        'Mail-Order Actuals': ch['Mail']['actuals'], 'Mail-Order STLY': ch['Mail']['stly'],
        'LTC Actuals': ch['LTC']['actuals'], 'LTC STLY': ch['LTC']['stly'],
    }
    return _build_xpt_excel(_npa_dates, data, sheet_name)

def _build_acute_prev_brand_excel(trx_df, nbrx_df, rx_class, brands):
    wks = sorted(trx_df['WEEK_ID'].unique())
    data = {}
    for brand in brands:
        t = trx_df[trx_df['BRAND'] == brand].sort_values('WEEK_ID')
        n = nbrx_df[nbrx_df['BRAND'] == brand].sort_values('WEEK_ID')
        data[f'{brand} TRx Actuals'] = list(t['ACTUALS'])
        data[f'{brand} TRx STLY'] = list(t['STLY'])
        data[f'{brand} NBRx Actuals'] = list(n['ACTUALS'])
        data[f'{brand} NBRx STLY'] = list(n['STLY'])
    dates = [week_to_date(w) for w in wks]
    return _build_xpt_excel(dates, data, f'{rx_class} Brand')

def _build_ap_channel_excel(segment, rx_class, brand, dates_list, sheet_name):
    ch_df = load_acute_prev_channel_data(segment, rx_class, brand)
    retail_df = ch_df[ch_df['CHANNEL_TYPE'] == 'Retail'].sort_values('WEEK_ID')
    mail_df = ch_df[ch_df['CHANNEL_TYPE'] == 'MAIL'].sort_values('WEEK_ID')
    ltc_df = ch_df[ch_df['CHANNEL_TYPE'] == 'Long term'].sort_values('WEEK_ID')
    data = {
        'Retail Actuals': list(retail_df['ACTUALS']), 'Retail STLY': list(retail_df['STLY']),
        'Mail-Order Actuals': list(mail_df['ACTUALS']), 'Mail-Order STLY': list(mail_df['STLY']),
        'LTC Actuals': list(ltc_df['ACTUALS']) if len(ltc_df) > 0 else [0]*len(dates_list),
        'LTC STLY': list(ltc_df['STLY']) if len(ltc_df) > 0 else [0]*len(dates_list),
    }
    return _build_xpt_excel(dates_list, data, sheet_name)

# NPA Overall Brand (2)
_npa_brand_trx_b64 = _build_npa_brand_excel(npa_brand_df, 'TRx', 'NPA Overall Brand TRx')
_npa_brand_nbrx_b64 = _build_npa_brand_excel(nbrx_brand_df, 'NBRx', 'NPA Overall Brand NBRx')
# NPA Overall Channel (6)
_npa_ch_nurtec_trx_b64 = _build_npa_channel_excel(_channel_trx_data, 'NURTEC', 'TRx', 'Channel Nurtec TRx')
_npa_ch_nurtec_nbrx_b64 = _build_npa_channel_excel(_channel_nbrx_data, 'NURTEC', 'NBRx', 'Channel Nurtec NBRx')
_npa_ch_ubrelvy_trx_b64 = _build_npa_channel_excel(_channel_trx_data, 'UBRELVY', 'TRx', 'Channel Ubrelvy TRx')
_npa_ch_ubrelvy_nbrx_b64 = _build_npa_channel_excel(_channel_nbrx_data, 'UBRELVY', 'NBRx', 'Channel Ubrelvy NBRx')
_npa_ch_qulipta_trx_b64 = _build_npa_channel_excel(_channel_trx_data, 'QULIPTA', 'TRx', 'Channel Qulipta TRx')
_npa_ch_qulipta_nbrx_b64 = _build_npa_channel_excel(_channel_nbrx_data, 'QULIPTA', 'NBRx', 'Channel Qulipta NBRx')
# Acute/Prev Brand (2)
_ap_dates = [week_to_date(w) for w in sorted(_acute_trx_df['WEEK_ID'].unique())]
_acute_brand_b64 = _build_acute_prev_brand_excel(_acute_trx_df, _acute_nbrx_df, 'Acute', ['NURTEC', 'UBRELVY'])
_prev_brand_b64 = _build_acute_prev_brand_excel(_prev_trx_df, _prev_nbrx_df, 'Preventive', ['NURTEC', 'QULIPTA'])
# Acute Channel (2)
_acute_ch_nurtec_b64 = _build_ap_channel_excel("TRx", "Acute", "NURTEC", _ap_dates, 'Acute Channel Nurtec')
_acute_ch_ubrelvy_b64 = _build_ap_channel_excel("TRx", "Acute", "UBRELVY", _ap_dates, 'Acute Channel Ubrelvy')
# Prev Channel (2)
_prev_ch_nurtec_b64 = _build_ap_channel_excel("TRx", "Preventive", "NURTEC", _ap_dates, 'Prev Channel Nurtec')
_prev_ch_qulipta_b64 = _build_ap_channel_excel("TRx", "Preventive", "QULIPTA", _ap_dates, 'Prev Channel Qulipta')

# --- Finance Trends Data (Live) ---
_fin_df = load_finance_trends()
_fin_gross_df = _fin_df[_fin_df['SECTION_NAME'] == 'Gross'].sort_values('DATE_PARSED')
_fin_net_df = _fin_df[_fin_df['SECTION_NAME'] == 'Net'].sort_values('DATE_PARSED')

_fin_gross_labels = list(_fin_gross_df['PERIOD_LABEL'])
_fin_gross_actuals = [v/1e6 if v and v > 0 else None for v in _fin_gross_df['ACTUAL_VALUE']]
_fin_gross_stly = [v/1e6 if v and v > 0 else None for v in _fin_gross_df['PRIOR_YEAR_VALUE']]
_fin_gross_budget = [v/1e6 if v and v > 0 else None for v in _fin_gross_df['BUDGET_VALUE']]

_fin_net_labels = list(_fin_net_df['PERIOD_LABEL'])
_fin_net_actuals = [v/1e6 if v and v > 0 else None for v in _fin_net_df['ACTUAL_VALUE']]
_fin_net_stly = [v/1e6 if v and v > 0 else None for v in _fin_net_df['PRIOR_YEAR_VALUE']]
_fin_net_budget = [v/1e6 if v and v > 0 else None for v in _fin_net_df['BUDGET_VALUE']]

# --- Finance Excel downloads ---
def _build_fin_excel(labels, actuals, stly, budget, sheet_name):
    data = {'Actual ($M)': actuals, 'STLY ($M)': stly, 'Budget ($M)': budget}
    dates = [datetime.strptime('2026-01-01', '%Y-%m-%d')] * len(labels)  # placeholder
    buf = io.BytesIO()
    df = pd.DataFrame({'Period': labels, 'Actual ($M)': actuals, 'STLY ($M)': stly, 'Budget ($M)': budget})
    df.to_excel(buf, index=False, sheet_name=sheet_name, engine='openpyxl')
    buf.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(buf)
    ws = wb.active
    header_fill = PatternFill(start_color='0000C9', end_color='0000C9', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return base64.b64encode(out.read()).decode()

_fin_gross_b64 = _build_fin_excel(_fin_gross_labels, _fin_gross_actuals, _fin_gross_stly, _fin_gross_budget, 'Gross Sales')
_fin_net_b64 = _build_fin_excel(_fin_net_labels, _fin_net_actuals, _fin_net_stly, _fin_net_budget, 'Net Sales')

# --- Xponent KPI values (Live) ---
_xpt_kpi_df = load_xponent_stacked()
_xpt_national = _xpt_kpi_df[_xpt_kpi_df['CUT_TYPE'] == 'National'].iloc[0] if len(_xpt_kpi_df[_xpt_kpi_df['CUT_TYPE'] == 'National']) > 0 else None

# --- Finance KPI values (Live) ---
_fin_kpi_df = load_finance_stacked()

# --- NPA Stacked Metrics (for tables & KPIs) ---
@st.cache_data(ttl=3600)
def load_npa_stacked():
    return fetch_npa_stacked(stack_key='NPA_STACKED_MATRICES')

_npa_stacked_df = load_npa_stacked()

def _fmt_val(v):
    if v is None or (isinstance(v, float) and v != v):
        return '—'
    v = float(v)
    if v >= 1000000:
        return f'{v/1000000:.1f}M' if v % 1000000 != 0 else f'{int(v/1000000)}M'
    return f'{int(v):,}'

def _fmt_pct(v, prefix=True):
    if v is None or (isinstance(v, float) and v != v):
        return '—'
    v = float(v)
    sign = '+' if v >= 0 and prefix else ''
    cls = 'delta-pos' if v >= 0 else 'delta-neg'
    return f'<span class="{cls}">{sign}{v:.1f}%</span>'

def _fmt_ms(v):
    if v is None or (isinstance(v, float) and v != v):
        return '—'
    return f'{float(v):.1f}%'

def _fmt_att(v):
    if v is None or (isinstance(v, float) and v != v):
        return '—'
    return f'{float(v):.1f}%'

def _build_npa_table_rows(brand, prescription, rx_class, compact=False):
    sub = _npa_stacked_df[
        (_npa_stacked_df['BRAND'] == brand) &
        (_npa_stacked_df['PRESCRIPTION'] == prescription) &
        (_npa_stacked_df['RX_CLASSIFICATION'] == rx_class)
    ]
    rows = []
    for row_label in ["Actuals '26", "Actuals '25", "Latest Goal OP'26"]:
        rd = sub[sub['ROW_LABEL'] == row_label]
        if len(rd) == 0:
            continue
        wk_row = rd[rd['TIME_PERIOD'] == 'Latest Week']
        qtd_row = rd[rd['TIME_PERIOD'] == 'QTD']
        ytd_row = rd[rd['TIME_PERIOD'] == 'YTD']
        wk_val = wk_row.iloc[0]['CURR_VALUE'] if len(wk_row) > 0 else None
        qtd_val = qtd_row.iloc[0]['CURR_VALUE'] if len(qtd_row) > 0 else None
        ytd_val = ytd_row.iloc[0]['CURR_VALUE'] if len(ytd_row) > 0 else None
        wk_pct = wk_row.iloc[0]['GROWTH_PCT'] if len(wk_row) > 0 else None
        qtd_pct = qtd_row.iloc[0]['GROWTH_PCT'] if len(qtd_row) > 0 else None
        ytd_pct = ytd_row.iloc[0]['GROWTH_PCT'] if len(ytd_row) > 0 else None
        wk_ms = wk_row.iloc[0]['MARKET_SHARE_PCT'] if len(wk_row) > 0 else None
        qtd_ms = qtd_row.iloc[0]['MARKET_SHARE_PCT'] if len(qtd_row) > 0 else None
        ytd_ms = ytd_row.iloc[0]['MARKET_SHARE_PCT'] if len(ytd_row) > 0 else None
        row_html = f'<tr><td>{row_label}</td>'
        row_html += f'<td style="text-align:right">{_fmt_val(wk_val)}</td>'
        row_html += f'<td style="text-align:right">{_fmt_val(qtd_val)}</td>'
        row_html += f'<td style="text-align:right">{_fmt_val(ytd_val)}</td>'
        row_html += f'<td style="text-align:right">{_fmt_pct(wk_pct)}</td>'
        if not compact:
            row_html += f'<td style="text-align:right">{_fmt_pct(qtd_pct)}</td>'
            row_html += f'<td style="text-align:right">{_fmt_pct(ytd_pct)}</td>'
        row_html += f'<td style="text-align:right">{_fmt_ms(wk_ms)}</td>'
        if not compact:
            row_html += f'<td style="text-align:right">{_fmt_ms(qtd_ms)}</td>'
            row_html += f'<td style="text-align:right">{_fmt_ms(ytd_ms)}</td>'
        row_html += '</tr>'
        rows.append(row_html)
    # Goal Attainment row (only for OVERALL)
    if not compact:
        att_rd = sub[(sub['ROW_LABEL'] == "Actuals '26")]
        if len(att_rd) > 0:
            wk_att = att_rd[att_rd['TIME_PERIOD'] == 'Latest Week']
            qtd_att = att_rd[att_rd['TIME_PERIOD'] == 'QTD']
            ytd_att = att_rd[att_rd['TIME_PERIOD'] == 'YTD']
            wk_a = wk_att.iloc[0]['GOAL_ATTAINMENT_PCT'] if len(wk_att) > 0 else None
            qtd_a = qtd_att.iloc[0]['GOAL_ATTAINMENT_PCT'] if len(qtd_att) > 0 else None
            ytd_a = ytd_att.iloc[0]['GOAL_ATTAINMENT_PCT'] if len(ytd_att) > 0 else None
            # Only show row if at least one value is a real number (not None/NaN)
            def _is_valid(v):
                return v is not None and not (isinstance(v, float) and v != v)
            if _is_valid(wk_a) or _is_valid(qtd_a) or _is_valid(ytd_a):
                row_html = f"<tr><td>Goal Attainment '26</td>"
                row_html += f'<td style="text-align:right">{_fmt_att(wk_a)}</td>'
                row_html += f'<td style="text-align:right">{_fmt_att(qtd_a)}</td>'
                row_html += f'<td style="text-align:right">{_fmt_att(ytd_a)}</td>'
                row_html += '<td style="text-align:right">\u2014</td><td style="text-align:right">\u2014</td><td style="text-align:right">\u2014</td>'
                row_html += '<td style="text-align:right">\u2014</td><td style="text-align:right">\u2014</td><td style="text-align:right">\u2014</td></tr>'
                rows.append(row_html)
    return '\n'.join(rows)

# Build all NPA table HTML
_npa_tables = {}
for _brand in ['NURTEC', 'UBRELVY', 'QULIPTA']:
    for _rx in ['TRx', 'NBRx']:
        _npa_tables[f'{_brand}_{_rx}_OVERALL'] = _build_npa_table_rows(_brand, _rx, 'OVERALL', compact=False)
for _brand in ['NURTEC', 'UBRELVY']:
    for _rx in ['TRx', 'NBRx']:
        _npa_tables[f'{_brand}_{_rx}_ACUTE'] = _build_npa_table_rows(_brand, _rx, 'ACUTE', compact=True)
for _brand in ['NURTEC', 'QULIPTA']:
    for _rx in ['TRx', 'NBRx']:
        _npa_tables[f'{_brand}_{_rx}_PREVENTIVE'] = _build_npa_table_rows(_brand, _rx, 'PREVENTIVE', compact=True)

# Build Hero KPI values from NPA stacked
def _get_npa_kpi(brand, prescription, rx_class, time_period, field):
    sub = _npa_stacked_df[
        (_npa_stacked_df['BRAND'] == brand) &
        (_npa_stacked_df['PRESCRIPTION'] == prescription) &
        (_npa_stacked_df['RX_CLASSIFICATION'] == rx_class) &
        (_npa_stacked_df['ROW_LABEL'] == "Actuals '26") &
        (_npa_stacked_df['TIME_PERIOD'] == time_period)
    ]
    if len(sub) > 0:
        return sub.iloc[0][field]
    return None


# Resolve finance access restriction (must be after load_finance_stacked is available)
FINANCE_RESTRICTED = _resolve_finance_restriction()


# --- Acute/Preventive Brand Charts ---
def build_acute_prev_chart(df, brand1, brand2, brand1_name, brand2_name, brand1_color, brand2_color, metric_label, dates_list):
    b1 = df[df['BRAND'] == brand1].sort_values('WEEK_ID')
    b2 = df[df['BRAND'] == brand2].sort_values('WEEK_ID')
    fig_ap = go.Figure()
    fig_ap.add_trace(go.Scatter(x=dates_list, y=list(b1['ACTUALS']), mode='lines', name=f'{brand1_name} Actuals', line=dict(color=brand1_color, width=2.5), hovertemplate=f'{brand1_name} Actuals<br>Week: %{{x|%d %b %y}}<br>{metric_label}: %{{y:,.0f}}<extra></extra>'))
    fig_ap.add_trace(go.Scatter(x=dates_list, y=list(b1['STLY']), mode='lines', name=f'{brand1_name} STLY', line=dict(color=brand1_color, width=2, dash='dash'), hovertemplate=f'{brand1_name} STLY<br>Week: %{{x|%d %b}} 25<br>{metric_label}: %{{y:,.0f}}<extra></extra>'))
    fig_ap.add_trace(go.Scatter(x=dates_list, y=list(b2['ACTUALS']), mode='lines', name=f'{brand2_name} Actuals', line=dict(color=brand2_color, width=2.5), hovertemplate=f'{brand2_name} Actuals<br>Week: %{{x|%d %b %y}}<br>{metric_label}: %{{y:,.0f}}<extra></extra>'))
    fig_ap.add_trace(go.Scatter(x=dates_list, y=list(b2['STLY']), mode='lines', name=f'{brand2_name} STLY', line=dict(color=brand2_color, width=2, dash='dash'), hovertemplate=f'{brand2_name} STLY<br>Week: %{{x|%d %b}} 25<br>{metric_label}: %{{y:,.0f}}<extra></extra>'))
    fig_ap.update_layout(
        height=300, margin=dict(l=50, r=10, t=10, b=80),
        plot_bgcolor='white', paper_bgcolor='white',
        xaxis=dict(tickfont=dict(size=8, color='#374151', family='Inter, sans-serif'), tickformat='%d %b %y', tickangle=-90, dtick=14*24*60*60*1000, showgrid=False, hoverformat=''),
        yaxis=dict(tickfont=dict(size=8, color='#374151', family='Inter, sans-serif'), showgrid=False, tickformat=',.0f', rangemode='tozero'),
        hovermode='closest', showlegend=False,
        hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'),
    )
    ap_html = fig_ap.to_html(full_html=False, include_plotlyjs=False, config={'displayModeBar': False, 'responsive': True})
    ap_html = ap_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')
    return ap_html

ap_dates = [week_to_date(w) for w in weeks]
acute_trx_chart = build_acute_prev_chart(_acute_trx_df, 'NURTEC', 'UBRELVY', 'Nurtec Acute', 'Ubrelvy Acute', '#7C6CFC', '#4ADE80', 'TRx', ap_dates)
acute_nbrx_chart = build_acute_prev_chart(_acute_nbrx_df, 'NURTEC', 'UBRELVY', 'Nurtec Acute', 'Ubrelvy Acute', '#7C6CFC', '#4ADE80', 'NBRx', ap_dates)
prev_trx_chart = build_acute_prev_chart(_prev_trx_df, 'NURTEC', 'QULIPTA', 'Nurtec Prev', 'Qulipta Prev', '#7C6CFC', '#FB923C', 'TRx', ap_dates)
prev_nbrx_chart = build_acute_prev_chart(_prev_nbrx_df, 'NURTEC', 'QULIPTA', 'Nurtec Prev', 'Qulipta Prev', '#7C6CFC', '#FB923C', 'NBRx', ap_dates)

# --- Acute/Preventive Channel Charts ---
def build_ap_channel_chart_live(segment, rx_class, brand, metric_label, dates_list):
    ch_df = load_acute_prev_channel_data(segment, rx_class, brand)
    retail_df = ch_df[ch_df['CHANNEL_TYPE'] == 'Retail'].sort_values('WEEK_ID')
    mail_df = ch_df[ch_df['CHANNEL_TYPE'] == 'MAIL'].sort_values('WEEK_ID')
    ltc_df = ch_df[ch_df['CHANNEL_TYPE'] == 'Long term'].sort_values('WEEK_ID')
    fig_ch = go.Figure()
    fig_ch.add_trace(go.Scatter(x=dates_list, y=list(retail_df['ACTUALS']), mode='lines', name='Retail Actuals', line=dict(color='#0891b2', width=2.5), hovertemplate='Retail Actuals<br>Week: %{x|%d %b %y}<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.add_trace(go.Scatter(x=dates_list, y=list(retail_df['STLY']), mode='lines', name='Retail STLY', line=dict(color='#0891b2', width=2, dash='dash'), hovertemplate='Retail STLY<br>Week: %{x|%d %b %y}<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.add_trace(go.Scatter(x=dates_list, y=list(mail_df['ACTUALS']), mode='lines', name='Mail-Order Actuals', line=dict(color='#7c3aed', width=2.5), hovertemplate='Mail-Order Actuals<br>Week: %{x|%d %b %y}<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.add_trace(go.Scatter(x=dates_list, y=list(mail_df['STLY']), mode='lines', name='Mail-Order STLY', line=dict(color='#7c3aed', width=2, dash='dash'), hovertemplate='Mail-Order STLY<br>Week: %{x|%d %b %y}<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    ltc_act = list(ltc_df['ACTUALS']) if len(ltc_df) > 0 else [0]*len(dates_list)
    ltc_stl = list(ltc_df['STLY']) if len(ltc_df) > 0 else [0]*len(dates_list)
    fig_ch.add_trace(go.Scatter(x=dates_list, y=ltc_act, mode='lines', name='LTC Actuals', line=dict(color='#9ca3af', width=2.5), hovertemplate='LTC Actuals<br>Week: %{x|%d %b %y}<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.add_trace(go.Scatter(x=dates_list, y=ltc_stl, mode='lines', name='LTC STLY', line=dict(color='#9ca3af', width=2, dash='dash'), hovertemplate='LTC STLY<br>Week: %{x|%d %b %y}<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.update_layout(
        height=300, margin=dict(l=50, r=10, t=10, b=80),
        plot_bgcolor='white', paper_bgcolor='white',
        xaxis=dict(tickfont=dict(size=8, color='#374151', family='Inter, sans-serif'), tickformat='%d %b %y', tickangle=-90, dtick=14*24*60*60*1000, showgrid=False, hoverformat=''),
        yaxis=dict(tickfont=dict(size=8, color='#374151', family='Inter, sans-serif'), showgrid=False, tickformat=',.0f', rangemode='tozero'),
        hovermode='closest', showlegend=False,
        hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'),
    )
    ch_html = fig_ch.to_html(full_html=False, include_plotlyjs=False, config={'displayModeBar': False, 'responsive': True})
    ch_html = ch_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')
    return ch_html

ap_ch_nurtec_acute_trx = build_ap_channel_chart_live("TRx", "Acute", "NURTEC", "TRx", ap_dates)
ap_ch_nurtec_acute_nbrx = build_ap_channel_chart_live("NBRx", "Acute", "NURTEC", "NBRx", ap_dates)
ap_ch_ubrelvy_acute_trx = build_ap_channel_chart_live("TRx", "Acute", "UBRELVY", "TRx", ap_dates)
ap_ch_ubrelvy_acute_nbrx = build_ap_channel_chart_live("NBRx", "Acute", "UBRELVY", "NBRx", ap_dates)
ap_ch_nurtec_prev_trx = build_ap_channel_chart_live("TRx", "Preventive", "NURTEC", "TRx", ap_dates)
ap_ch_nurtec_prev_nbrx = build_ap_channel_chart_live("NBRx", "Preventive", "NURTEC", "NBRx", ap_dates)
ap_ch_qulipta_prev_trx = build_ap_channel_chart_live("TRx", "Preventive", "QULIPTA", "TRx", ap_dates)
ap_ch_qulipta_prev_nbrx = build_ap_channel_chart_live("NBRx", "Preventive", "QULIPTA", "NBRx", ap_dates)


# --- Generate NPA Overall Charts ---
nurtec_dates = [week_to_date(w) for w in nurtec_data['WEEK_ID']]
ubrelvy_dates = [week_to_date(w) for w in ubrelvy_data['WEEK_ID']]
qulipta_dates = [week_to_date(w) for w in qulipta_data['WEEK_ID']]

fig = go.Figure()

# Nurtec Actuals
fig.add_trace(go.Scatter(x=nurtec_dates, y=list(nurtec_data['ACTUALS']),
    mode='lines', name='Nurtec Actuals', line=dict(color='#7C6CFC', width=2.5),
    hovertemplate='Nurtec Actuals<br>Week: %{x|%d %b %y}<br>TRx: %{y:,.0f}<extra></extra>'))
# Nurtec STLY
fig.add_trace(go.Scatter(x=nurtec_dates, y=list(nurtec_data['STLY']),
    mode='lines', name='Nurtec STLY', line=dict(color='#7C6CFC', width=2, dash='dash'),
    hovertemplate='Nurtec STLY<br>STLY Week: %{x|%d %b} 25<br>TRx: %{y:,.0f}<extra></extra>'))
# Ubrelvy Actuals
fig.add_trace(go.Scatter(x=ubrelvy_dates, y=list(ubrelvy_data['ACTUALS']),
    mode='lines', name='Ubrelvy Actuals', line=dict(color='#4ADE80', width=2.5),
    hovertemplate='Ubrelvy Actuals<br>Week: %{x|%d %b %y}<br>TRx: %{y:,.0f}<extra></extra>'))
# Ubrelvy STLY
fig.add_trace(go.Scatter(x=ubrelvy_dates, y=list(ubrelvy_data['STLY']),
    mode='lines', name='Ubrelvy STLY', line=dict(color='#4ADE80', width=2, dash='dash'),
    hovertemplate='Ubrelvy STLY<br>STLY Week: %{x|%d %b} 25<br>TRx: %{y:,.0f}<extra></extra>'))
# Qulipta Actuals
fig.add_trace(go.Scatter(x=qulipta_dates, y=list(qulipta_data['ACTUALS']),
    mode='lines', name='Qulipta Actuals', line=dict(color='#FB923C', width=2.5),
    hovertemplate='Qulipta Actuals<br>Week: %{x|%d %b %y}<br>TRx: %{y:,.0f}<extra></extra>'))
# Qulipta STLY
fig.add_trace(go.Scatter(x=qulipta_dates, y=list(qulipta_data['STLY']),
    mode='lines', name='Qulipta STLY', line=dict(color='#FB923C', width=2, dash='dash'),
    hovertemplate='Qulipta STLY<br>STLY Week: %{x|%d %b} 25<br>TRx: %{y:,.0f}<extra></extra>'))
# Goal
goal_vals = [v if v is not None and v > 0 else None for v in list(nurtec_data['LATEST_GOAL'])]
fig.add_trace(go.Scatter(x=nurtec_dates, y=goal_vals,
    mode='lines', name='Goal', line=dict(color='#f472b6', width=2, dash='dashdot'),
    hovertemplate='Goal<br>Week: %{x|%d %b %y}<br>TRx: %{y:,.0f}<extra></extra>'))

fig.update_layout(
    height=340,
    margin=dict(l=60, r=20, t=35, b=100),
    plot_bgcolor='white',
    paper_bgcolor='white',
    xaxis=dict(
        tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'),
        tickformat='%d %b %y',
        tickangle=-90,
        dtick=7*24*60*60*1000,
        gridcolor='rgba(0,0,0,0.04)',
        showgrid=False,
        hoverformat='',
    ),
    yaxis=dict(
        title=dict(text='TRx Volume', font=dict(size=10, color='#4b5563')),
        tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'),
        gridcolor='rgba(0,0,0,0.06)',
        showgrid=False,
        tickformat=',',
        rangemode='tozero',
    ),
    legend=dict(
        orientation='h',
        yanchor='top',
        y=-0.35,
        xanchor='center',
        x=0.5,
        font=dict(size=9),
    ),
    hovermode='closest',
    hoverlabel=dict(
        bgcolor='white',
        font=dict(size=11, color='#1a2332', family='Inter, sans-serif'),
        bordercolor='rgba(0,0,0,0)',
    ),
)

# Export to HTML div (self-contained with plotly.js CDN)
chart_html = fig.to_html(full_html=False, include_plotlyjs='cdn', config={'displayModeBar': False, 'responsive': True})
# Make chart div fill container width
chart_html = chart_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')
brand_chart_svg = chart_html

# Generate NBRx Plotly chart
fig_nbrx = go.Figure()
nbrx_dates = [week_to_date(w) for w in nbrx_nurtec['WEEK_ID']]

fig_nbrx.add_trace(go.Scatter(x=nbrx_dates, y=list(nbrx_nurtec['ACTUALS']), mode='lines', name='Nurtec Actuals', line=dict(color='#7C6CFC', width=2.5), hovertemplate='Nurtec Actuals<br>Week: %{x|%d %b %y}<br>NBRx: %{y:,.0f}<extra></extra>'))
fig_nbrx.add_trace(go.Scatter(x=nbrx_dates, y=list(nbrx_nurtec['STLY']), mode='lines', name='Nurtec STLY', line=dict(color='#7C6CFC', width=2, dash='dash'), hovertemplate='Nurtec STLY<br>STLY Week: %{x|%d %b} 25<br>NBRx: %{y:,.0f}<extra></extra>'))
fig_nbrx.add_trace(go.Scatter(x=nbrx_dates, y=list(nbrx_ubrelvy['ACTUALS']), mode='lines', name='Ubrelvy Actuals', line=dict(color='#4ADE80', width=2.5), hovertemplate='Ubrelvy Actuals<br>Week: %{x|%d %b %y}<br>NBRx: %{y:,.0f}<extra></extra>'))
fig_nbrx.add_trace(go.Scatter(x=nbrx_dates, y=list(nbrx_ubrelvy['STLY']), mode='lines', name='Ubrelvy STLY', line=dict(color='#4ADE80', width=2, dash='dash'), hovertemplate='Ubrelvy STLY<br>STLY Week: %{x|%d %b} 25<br>NBRx: %{y:,.0f}<extra></extra>'))
fig_nbrx.add_trace(go.Scatter(x=nbrx_dates, y=list(nbrx_qulipta['ACTUALS']), mode='lines', name='Qulipta Actuals', line=dict(color='#FB923C', width=2.5), hovertemplate='Qulipta Actuals<br>Week: %{x|%d %b %y}<br>NBRx: %{y:,.0f}<extra></extra>'))
fig_nbrx.add_trace(go.Scatter(x=nbrx_dates, y=list(nbrx_qulipta['STLY']), mode='lines', name='Qulipta STLY', line=dict(color='#FB923C', width=2, dash='dash'), hovertemplate='Qulipta STLY<br>STLY Week: %{x|%d %b} 25<br>NBRx: %{y:,.0f}<extra></extra>'))

fig_nbrx.update_layout(
    height=340, margin=dict(l=60, r=20, t=35, b=100),
    plot_bgcolor='white', paper_bgcolor='white',
    xaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), tickformat='%d %b %y', tickangle=-90, dtick=7*24*60*60*1000, showgrid=False, hoverformat=''),
    yaxis=dict(title=dict(text='NBRx Volume', font=dict(size=10, color='#4b5563')), tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), showgrid=False, tickformat=',', rangemode='tozero'),
    legend=dict(orientation='h', yanchor='top', y=-0.35, xanchor='center', x=0.5, font=dict(size=9)),
    hovermode='closest',
    hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'),
)

nbrx_chart_html = fig_nbrx.to_html(full_html=False, include_plotlyjs=False, config={'displayModeBar': False, 'responsive': True})
nbrx_chart_html = nbrx_chart_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')
nbrx_chart_svg = nbrx_chart_html

# --- Channel Performance Charts (6 total: 3 brands x 2 metrics) ---
def build_channel_chart(brand_data, metric_label, dates):
    fig_ch = go.Figure()
    retail_act = brand_data["Retail"]["actuals"]
    retail_stl = brand_data["Retail"]["stly"]
    mail_act = brand_data["Mail"]["actuals"]
    mail_stl = brand_data["Mail"]["stly"]
    ltc_act = brand_data.get("LTC", {}).get("actuals", [0]*len(dates))
    ltc_stl = brand_data.get("LTC", {}).get("stly", [0]*len(dates))
    fig_ch.add_trace(go.Scatter(x=dates, y=retail_act, mode='lines', name='Retail Actuals', line=dict(color='#0891b2', width=2.5), hovertemplate='Retail Actuals<br>Week: %{x|%d %b %y}<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.add_trace(go.Scatter(x=dates, y=retail_stl, mode='lines', name='Retail STLY', line=dict(color='#0891b2', width=2, dash='dash'), hovertemplate='Retail STLY<br>STLY Week: %{x|%d %b} 25<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.add_trace(go.Scatter(x=dates, y=mail_act, mode='lines', name='Mail-Order Actuals', line=dict(color='#7c3aed', width=2.5), hovertemplate='Mail-Order Actuals<br>Week: %{x|%d %b %y}<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.add_trace(go.Scatter(x=dates, y=mail_stl, mode='lines', name='Mail-Order STLY', line=dict(color='#7c3aed', width=2, dash='dash'), hovertemplate='Mail-Order STLY<br>STLY Week: %{x|%d %b} 25<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.add_trace(go.Scatter(x=dates, y=ltc_act, mode='lines', name='LTC Actuals', line=dict(color='#9ca3af', width=2.5), hovertemplate='LTC Actuals<br>Week: %{x|%d %b %y}<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.add_trace(go.Scatter(x=dates, y=ltc_stl, mode='lines', name='LTC STLY', line=dict(color='#9ca3af', width=2, dash='dash'), hovertemplate='LTC STLY<br>STLY Week: %{x|%d %b} 25<br>' + metric_label + ': %{y:,.0f}<extra></extra>'))
    fig_ch.update_layout(
        height=340, margin=dict(l=60, r=20, t=35, b=100),
        plot_bgcolor='white', paper_bgcolor='white',
        xaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), tickformat='%d %b %y', tickangle=-90, dtick=7*24*60*60*1000, showgrid=False, hoverformat=''),
        yaxis=dict(title=dict(text=metric_label + ' Volume', font=dict(size=10, color='#4b5563')), tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), showgrid=False, tickformat=',', rangemode='tozero'),
        legend=dict(orientation='h', yanchor='top', y=-0.35, xanchor='center', x=0.5, font=dict(size=9)),
        hovermode='closest',
        hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'),
    )
    ch_html = fig_ch.to_html(full_html=False, include_plotlyjs=False, config={'displayModeBar': False, 'responsive': True})
    ch_html = ch_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')
    return ch_html

channel_dates = [week_to_date(w) for w in weeks]
ch_nurtec_trx = build_channel_chart(_channel_trx_data["NURTEC"], "TRx", channel_dates)
ch_ubrelvy_trx = build_channel_chart(_channel_trx_data["UBRELVY"], "TRx", channel_dates)
ch_qulipta_trx = build_channel_chart(_channel_trx_data["QULIPTA"], "TRx", channel_dates)
ch_nurtec_nbrx = build_channel_chart(_channel_nbrx_data["NURTEC"], "NBRx", channel_dates)
ch_ubrelvy_nbrx = build_channel_chart(_channel_nbrx_data["UBRELVY"], "NBRx", channel_dates)
ch_qulipta_nbrx = build_channel_chart(_channel_nbrx_data["QULIPTA"], "NBRx", channel_dates)


# Build XPT Payer TRx chart
fig_xpt_p_trx = go.Figure()
fig_xpt_p_trx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_payer_trx['Commercial'], mode='lines', name='Commercial', line=dict(color='#3b82f6', width=2.5), hovertemplate='Commercial<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_p_trx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_payer_trx['Medicare'], mode='lines', name='Medicare', line=dict(color='#16a34a', width=2.5), hovertemplate='Medicare<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_p_trx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_payer_trx['Medicaid'], mode='lines', name='Medicaid', line=dict(color='#f59e0b', width=2.5), hovertemplate='Medicaid<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_p_trx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_payer_trx['Other'], mode='lines', name='Other', line=dict(color='#9ca3af', width=2.5), hovertemplate='Other<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_p_trx.update_layout(height=300, margin=dict(l=50, r=20, t=10, b=100), plot_bgcolor='white', paper_bgcolor='white',
    xaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), tickformat='%d %b %y', tickangle=-90, dtick=14*24*60*60*1000, showgrid=False, hoverformat='', showline=True, linewidth=1, linecolor='#1a2332'),
    yaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), showgrid=False, ticksuffix='%'),
    showlegend=False, hovermode='closest',
    hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'))
xpt_payer_trx_html = fig_xpt_p_trx.to_html(full_html=False, include_plotlyjs='cdn', config={'displayModeBar': False, 'responsive': True})
xpt_payer_trx_html = xpt_payer_trx_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')

# Build XPT Payer NRx chart
fig_xpt_p_nrx = go.Figure()
fig_xpt_p_nrx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_payer_nrx['Commercial'], mode='lines', name='Commercial', line=dict(color='#3b82f6', width=2.5), hovertemplate='Commercial<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_p_nrx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_payer_nrx['Medicare'], mode='lines', name='Medicare', line=dict(color='#16a34a', width=2.5), hovertemplate='Medicare<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_p_nrx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_payer_nrx['Medicaid'], mode='lines', name='Medicaid', line=dict(color='#f59e0b', width=2.5), hovertemplate='Medicaid<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_p_nrx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_payer_nrx['Other'], mode='lines', name='Other', line=dict(color='#9ca3af', width=2.5), hovertemplate='Other<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_p_nrx.update_layout(height=300, margin=dict(l=50, r=20, t=10, b=100), plot_bgcolor='white', paper_bgcolor='white',
    xaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), tickformat='%d %b %y', tickangle=-90, dtick=14*24*60*60*1000, showgrid=False, hoverformat='', showline=True, linewidth=1, linecolor='#1a2332'),
    yaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), showgrid=False, ticksuffix='%'),
    showlegend=False, hovermode='closest',
    hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'))
xpt_payer_nrx_html = fig_xpt_p_nrx.to_html(full_html=False, include_plotlyjs=False, config={'displayModeBar': False, 'responsive': True})
xpt_payer_nrx_html = xpt_payer_nrx_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')

# Build XPT Channel TRx chart
fig_xpt_c_trx = go.Figure()
fig_xpt_c_trx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_ch_trx['Retail'], mode='lines', name='Retail', line=dict(color='#0891b2', width=2.5), hovertemplate='Retail<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_c_trx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_ch_trx['Mail-Order'], mode='lines', name='Mail-Order', line=dict(color='#7c3aed', width=2.5), hovertemplate='Mail-Order<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_c_trx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_ch_trx['LTC'], mode='lines', name='LTC', line=dict(color='#9ca3af', width=2.5), hovertemplate='LTC<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_c_trx.update_layout(height=300, margin=dict(l=50, r=20, t=10, b=100), plot_bgcolor='white', paper_bgcolor='white',
    xaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), tickformat='%d %b %y', tickangle=-90, dtick=14*24*60*60*1000, showgrid=False, hoverformat='', showline=True, linewidth=1, linecolor='#1a2332'),
    yaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), showgrid=False, ticksuffix='%'),
    showlegend=False, hovermode='closest',
    hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'))
xpt_ch_trx_html = fig_xpt_c_trx.to_html(full_html=False, include_plotlyjs=False, config={'displayModeBar': False, 'responsive': True})
xpt_ch_trx_html = xpt_ch_trx_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')

# Build XPT Channel NRx chart
fig_xpt_c_nrx = go.Figure()
fig_xpt_c_nrx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_ch_nrx['Retail'], mode='lines', name='Retail', line=dict(color='#0891b2', width=2.5), hovertemplate='Retail<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_c_nrx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_ch_nrx['Mail-Order'], mode='lines', name='Mail-Order', line=dict(color='#7c3aed', width=2.5), hovertemplate='Mail-Order<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_c_nrx.add_trace(go.Scatter(x=_xpt_dates, y=_xpt_ch_nrx['LTC'], mode='lines', name='LTC', line=dict(color='#9ca3af', width=2.5), hovertemplate='LTC<br>Week: %{x|%d %b %y}<br>Share: %{y:.1f}%<extra></extra>'))
fig_xpt_c_nrx.update_layout(height=300, margin=dict(l=50, r=20, t=10, b=100), plot_bgcolor='white', paper_bgcolor='white',
    xaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), tickformat='%d %b %y', tickangle=-90, dtick=14*24*60*60*1000, showgrid=False, hoverformat='', showline=True, linewidth=1, linecolor='#1a2332'),
    yaxis=dict(tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), showgrid=False, ticksuffix='%'),
    showlegend=False, hovermode='closest',
    hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'))
xpt_ch_nrx_html = fig_xpt_c_nrx.to_html(full_html=False, include_plotlyjs=False, config={'displayModeBar': False, 'responsive': True})
xpt_ch_nrx_html = xpt_ch_nrx_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')


# --- Finance Trend Charts (from live data) ---
fig_fin_gross = go.Figure()
fig_fin_gross.add_trace(go.Bar(x=_fin_gross_labels, y=_fin_gross_actuals, name='Actual', marker_color='#7C6CFC', hovertemplate='Actual : $%{y:.1f}M<extra></extra>'))
fig_fin_gross.add_trace(go.Scatter(x=_fin_gross_labels, y=_fin_gross_stly, mode='lines', name='Actual (STLY)', line=dict(color='#0000C9', width=2.5), hovertemplate='Actual (STLY) : $%{y:.1f}M<extra></extra>'))
fig_fin_gross.add_trace(go.Scatter(x=_fin_gross_labels, y=_fin_gross_budget, mode='lines', name='Budget', line=dict(color='#9ca3af', width=2, dash='dash'), hovertemplate='Budget : $%{y:.1f}M<extra></extra>'))
fig_fin_gross.update_layout(height=340, margin=dict(l=60, r=20, t=10, b=100), plot_bgcolor='white', paper_bgcolor='white',
    xaxis=dict(tickfont=dict(size=8, color='#374151', family='Inter, sans-serif'), tickangle=-90, showgrid=False, hoverformat='', showline=True, linewidth=1, linecolor='#1a2332'),
    yaxis=dict(title=dict(text='Gross Sales ($M)', font=dict(size=10, color='#4b5563')), tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), showgrid=False, tickprefix='$', ticksuffix='M'),
    showlegend=False, hovermode='x unified', barmode='group', bargap=0.3,
    hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'))
fin_gross_html = fig_fin_gross.to_html(full_html=False, include_plotlyjs=False, config={'displayModeBar': False, 'responsive': True})
fin_gross_html = fin_gross_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')

# Net Sales chart (monthly - uses live data from lines 163-166)
fig_fin_net = go.Figure()
fig_fin_net.add_trace(go.Bar(x=_fin_net_labels, y=_fin_net_actuals, name='Actual', marker_color='#7C6CFC', hovertemplate='Actual : $%{y:.1f}M<extra></extra>'))
fig_fin_net.add_trace(go.Scatter(x=_fin_net_labels, y=_fin_net_stly, mode='lines', name='Actual (STLY)', line=dict(color='#0000C9', width=2.5), hovertemplate='Actual (STLY) : $%{y:.1f}M<extra></extra>'))
fig_fin_net.add_trace(go.Scatter(x=_fin_net_labels, y=_fin_net_budget, mode='lines', name='Budget', line=dict(color='#9ca3af', width=2, dash='dash'), hovertemplate='Budget : $%{y:.1f}M<extra></extra>'))
fig_fin_net.update_layout(height=340, margin=dict(l=60, r=10, t=10, b=80), plot_bgcolor='white', paper_bgcolor='white', autosize=True,
    xaxis=dict(type='category', categoryorder='array', categoryarray=_fin_net_labels, tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), tickangle=0, showgrid=False, hoverformat='', showline=True, linewidth=1, linecolor='#1a2332', range=[-0.5, 11.5], constrain='domain'),
    yaxis=dict(title=dict(text='Net Sales ($M)', font=dict(size=10, color='#4b5563')), tickfont=dict(size=9, color='#374151', family='Inter, sans-serif'), showgrid=False, tickprefix='$', ticksuffix='M'),
    showlegend=False, hovermode='x unified', barmode='group', bargap=0.15,
    hoverlabel=dict(bgcolor='white', font=dict(size=11, color='#1a2332', family='Inter, sans-serif'), bordercolor='rgba(0,0,0,0)'))
fin_net_html = fig_fin_net.to_html(full_html=False, include_plotlyjs='cdn', config={'displayModeBar': False, 'responsive': True})
fin_net_html = fin_net_html.replace('class="plotly-graph-div" style="', 'class="plotly-graph-div" style="width:100%;')








st.markdown("""
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
[data-testid="stSidebar"] {display: none;}
.block-container {padding: 0 !important; max-width: 100% !important;}
[data-testid="stAppViewBlockContainer"] {padding: 0 !important;}
iframe {height: 100vh !important; min-height: 100vh !important;}
</style>
""", unsafe_allow_html=True)

html_content = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Migraine Intelligence Hub</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<style>
:root {
    --navy-900: #0000C9;
    --navy-800: #0000C9;
    --navy-700: #0000C9;
    --navy-600: #0000C9;
    --navy-500: #0000C9;
    --accent: #0EA5E9;
    
    --bg: #F0F4FA;
    --bg-2: #EFF6FF;
    --surface: #FFFFFF;
    --surface-2: #F8FAFD;
    --text: #1A1A1A;
    --text-muted: #64748B;
    --text-soft: #475569;
    --hairline: rgba(15,23,42,0.08);
    --hairline-2: rgba(15,23,42,0.05);
    --up: #10B981;
    --down: #EF4444;
    --flat: #94A3B8;
    --shadow-xs: 0 1px 2px rgba(15,23,42,0.04);
    --shadow-sm: 0 2px 8px rgba(15,23,42,0.05), 0 1px 2px rgba(15,23,42,0.04);
    --shadow-md: 0 6px 16px rgba(15,23,42,0.07), 0 2px 4px rgba(15,23,42,0.04);
    --shadow-lg: 0 18px 40px rgba(15,23,42,0.10), 0 6px 12px rgba(15,23,42,0.06);
    --shadow-panel: 0 8px 24px rgba(15,23,42,0.07), 0 2px 6px rgba(15,23,42,0.04);
    --ease: cubic-bezier(0.4, 0, 0.2, 1);
    --ease-out: cubic-bezier(0.16, 1, 0.3, 1);
    --sidebar-w: 232px;
    --shell-pad: 10px;
    --panel-radius: 18px;
}

* { margin: 0; padding: 0; box-sizing: border-box; }
html, body { height: 100%; }

body {
    font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif;
    background: linear-gradient(160deg, #EFF6FF 0%, #F0F4FA 50%, #EEF2FF 100%);
    color: var(--text);
    line-height: 1.5;
    font-size: 14px;
    -webkit-font-smoothing: antialiased;
    -moz-osx-font-smoothing: grayscale;
    overflow: hidden;
}

h1, h2, h3, h4 { font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif; font-weight: 800; letter-spacing: -0.015em; }
a { color: inherit; text-decoration: none; }

/* ───── APP SHELL ───── */
.app { height: 100vh; display: grid; grid-template-columns: var(--sidebar-w) 1fr; gap: var(--shell-pad); padding: var(--shell-pad); overflow: hidden; }

/* ───── SIDEBAR ───── */
.sidebar { position: relative; background: rgba(255,255,255,0.62); backdrop-filter: saturate(180%) blur(22px); -webkit-backdrop-filter: saturate(180%) blur(22px); border: 1px solid var(--hairline); border-radius: var(--panel-radius); box-shadow: var(--shadow-panel); display: flex; flex-direction: column; overflow: hidden; z-index: 10; }

.sidebar-brand { padding: 1.4rem 1.2rem 1.2rem; display: flex; flex-direction: column; gap: 0.7rem; }
.sidebar-brand img { height: 28px; align-self: flex-start; }
.sidebar-brand .title { font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif; font-weight: 800; font-size: 1.22rem; color: var(--navy-900); line-height: 1.18; letter-spacing: -0.025em; }
.sidebar-brand .subtitle { font-size: 0.72rem; color: var(--text-muted); font-weight: 500; }

.sidebar-divider { height: 1px; background: var(--hairline); margin: 0 0.85rem; }

.sidebar-section-label { font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif; font-size: 0.62rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.12em; color: var(--text-muted); padding: 0.95rem 1.15rem 0.4rem; }

.nav { padding: 0 0.55rem; }
.nav-item {
    position: relative; display: flex; align-items: center; gap: 0.7rem;
    padding: 0.55rem 0.7rem; margin: 0.08rem 0; border-radius: 8px;
    font-size: 0.84rem; font-weight: 500; color: var(--text-soft);
    cursor: pointer; transition: background 0.18s var(--ease), color 0.18s var(--ease);
    background: transparent; border: none; width: 100%; text-align: left; font-family: inherit;
}
.nav-item .nav-icon { width: 18px; height: 18px; display: flex; align-items: center; justify-content: center; color: var(--text-muted); transition: color 0.18s var(--ease); flex-shrink: 0; }
.nav-item .nav-icon svg { width: 16px; height: 16px; stroke-width: 1.8; fill: none; stroke: currentColor; }
.nav-item .nav-label { flex: 1; min-width: 0; }
.nav-item .nav-count { font-size: 0.66rem; font-weight: 600; color: var(--text-muted); background: rgba(15,23,42,0.06); padding: 0.12rem 0.42rem; border-radius: 5px; font-variant-numeric: tabular-nums; transition: background 0.18s var(--ease), color 0.18s var(--ease); flex-shrink: 0; line-height: 1.3; }
.nav-item:hover { background: rgba(15,23,42,0.04); color: var(--text); }
.nav-item:hover .nav-icon { color: var(--navy-700); }
.nav-item:hover .nav-count { background: rgba(15,23,42,0.09); color: var(--text-soft); }

.nav-item.active { background: linear-gradient(90deg, rgba(28,79,192,0.10) 0%, rgba(28,79,192,0.04) 100%); color: var(--navy-700); font-weight: 600; }
.nav-item.active .nav-icon { color: var(--navy-700); }
.nav-item.active .nav-count { background: rgba(28,79,192,0.14); color: var(--navy-700); }
.nav-item.active::before { content: ''; position: absolute; left: -0.55rem; top: 6px; bottom: 6px; width: 3px; border-radius: 0 3px 3px 0; background: linear-gradient(180deg, var(--navy-600), var(--accent)); box-shadow: 0 0 8px rgba(28,79,192,0.3); }

/* ───── SUB-NAV (indented children) ───── */
.nav-sub {
    display: none;
    flex-direction: column;
    padding: 0.15rem 0 0.15rem 1.8rem;
}
.nav-sub.is-open { display: flex; }
.nav-sub-item {
    position: relative;
    display: flex;
    align-items: center;
    gap: 0.5rem;
    padding: 0.38rem 0.65rem;
    margin: 0.04rem 0;
    border-radius: 6px;
    font-size: 0.76rem;
    font-weight: 500;
    color: var(--text-muted);
    cursor: pointer;
    transition: background 0.18s var(--ease), color 0.18s var(--ease);
    background: transparent;
    border: none;
    width: 100%;
    text-align: left;
    font-family: inherit;
}
.nav-sub-item:hover { background: rgba(15,23,42,0.04); color: var(--text); }
.nav-sub-item.active { background: rgba(28,79,192,0.08); color: var(--navy-700); font-weight: 600; }
.nav-sub-item.external { color: var(--text-muted); font-style: italic; }
.nav-sub-item.external svg { width: 10px; height: 10px; stroke: currentColor; fill: none; stroke-width: 2; flex-shrink: 0; }

.sidebar-spacer { flex: 1; }

.sidebar-meta { padding: 0.85rem 1.15rem 1rem; font-size: 0.7rem; color: var(--text-muted); line-height: 1.55; border-top: 1px solid var(--hairline); background: linear-gradient(180deg, transparent 0%, rgba(28,79,192,0.025) 100%); overflow: hidden; }
.sidebar-meta strong { color: var(--text-soft); font-weight: 600; }
.sidebar-meta a { color: var(--navy-700); white-space: nowrap; font-size: 0.54rem; }
.sidebar-meta a:hover { text-decoration: underline; }
.sidebar-meta .meta-row { margin-bottom: 0.2rem; }

/* ───── MAIN ───── */
.main { background: rgba(255,255,255,0.55); backdrop-filter: saturate(180%) blur(14px); -webkit-backdrop-filter: saturate(180%) blur(14px); border: 1px solid var(--hairline); border-radius: var(--panel-radius); box-shadow: var(--shadow-panel); display: flex; flex-direction: column; overflow: hidden; min-width: 0; }

/* ───── CONTENT ───── */
.content { flex: 1; min-height: 0; overflow-y: auto; padding: 1.4rem; }
.content::-webkit-scrollbar { width: 6px; }
.content::-webkit-scrollbar-track { background: transparent; }
.content::-webkit-scrollbar-thumb { background: rgba(15,23,42,0.14); border-radius: 3px; }
.content::-webkit-scrollbar-thumb:hover { background: rgba(15,23,42,0.24); }

/* ───── DROPDOWN ───── */
.dropdown-wrap { position: relative; }
.icon-btn { display: inline-flex; align-items: center; gap: 0.4rem; padding: 0.4rem 0.7rem; border-radius: 7px; background: rgba(255,255,255,0.7); border: 1px solid var(--hairline); color: var(--text-soft); font-size: 0.75rem; font-weight: 500; cursor: pointer; font-family: inherit; transition: all 0.18s var(--ease); }
.icon-btn:hover { background: #fff; color: var(--navy-700); border-color: rgba(28,79,192,0.25); box-shadow: var(--shadow-xs); }
.icon-btn svg { width: 13px; height: 13px; stroke-width: 1.8; fill: none; stroke: currentColor; }
.dropdown { position: absolute; top: calc(100% + 6px); right: 0; background: rgba(255,255,255,0.96); backdrop-filter: saturate(180%) blur(20px); border: 1px solid var(--hairline); border-radius: 12px; box-shadow: var(--shadow-lg); min-width: 240px; padding: 0.45rem 0; opacity: 0; visibility: hidden; transform: translateY(-6px) scale(0.98); transform-origin: top right; transition: opacity 0.2s var(--ease), transform 0.2s var(--ease), visibility 0.2s; z-index: 200; }
.dropdown.show { opacity: 1; visibility: visible; transform: translateY(0) scale(1); }
.dropdown-header { font-size: 0.62rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; color: var(--text-muted); padding: 0.5rem 0.95rem 0.35rem; }
.dropdown-item { display: flex; align-items: center; justify-content: space-between; padding: 0.45rem 0.95rem; font-size: 0.78rem; }
.dropdown-item:hover { background: rgba(15,23,42,0.04); }
.dropdown-item .src { font-weight: 500; }
.dropdown-item .date { font-size: 0.7rem; color: var(--text-muted); font-variant-numeric: tabular-nums; }

/* ───── SECTIONS ───── */
.section { display: none; opacity: 0; transform: translateY(4px); transition: opacity 0.22s var(--ease), transform 0.22s var(--ease-out); }
.section.is-active { display: block; }
.section.is-visible { opacity: 1; transform: translateY(0); }

.section-head { margin-bottom: 1rem; }
.section-head-row { display: flex; align-items: center; justify-content: space-between; gap: 1rem; margin-bottom: 0.2rem; }
.section-head h2 { font-size: 1.35rem; font-weight: 700; color: var(--navy-900); letter-spacing: -0.02em; margin-bottom: 0.2rem; }
.section-head p { font-size: 0.84rem; color: var(--text-muted); max-width: 680px; }

.grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 1rem; }

/* ───── CARDS ───── */
.card { position: relative; display: flex; flex-direction: column; background: var(--surface); border-radius: 14px; padding: 1.15rem 1.2rem; min-height: 148px; overflow: hidden; box-shadow: var(--shadow-sm); transition: transform 0.28s var(--ease-out), box-shadow 0.28s var(--ease); cursor: pointer; }
.card::after { content: ''; position: absolute; inset: 0; border-radius: inherit; background: linear-gradient(135deg, rgba(255,255,255,0) 55%, rgba(65,182,230,0.05) 80%, rgba(28,79,192,0.07) 100%); opacity: 0; transition: opacity 0.28s var(--ease); pointer-events: none; }
.card:hover { transform: translateY(-3px); box-shadow: var(--shadow-lg); }
.card:hover::after { opacity: 1; }

.card-top { display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.7rem; }
.icon-chip { width: 38px; height: 38px; border-radius: 10px; display: flex; align-items: center; justify-content: center; flex-shrink: 0; }
.icon-chip svg { width: 19px; height: 19px; stroke-width: 1.8; fill: none; }
.chip-s1 { background: linear-gradient(135deg, #DBEAFE, #BFDBFE); } .chip-s1 svg { stroke: #1D4ED8; }
.chip-s2 { background: linear-gradient(135deg, #E0E7FF, #C7D2FE); } .chip-s2 svg { stroke: #4338CA; }
.chip-s3 { background: linear-gradient(135deg, #EDE9FE, #DDD6FE); } .chip-s3 svg { stroke: #6D28D9; }
.chip-s4 { background: linear-gradient(135deg, #CFFAFE, #A5F3FC); } .chip-s4 svg { stroke: #0E7490; }
.chip-s5 { background: linear-gradient(135deg, #DCFCE7, #BBF7D0); } .chip-s5 svg { stroke: #047857; }

.card-title { font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif; font-size: 1.02rem; font-weight: 700; color: var(--navy-900); line-height: 1.25; margin-bottom: 0.3rem; }
.card-desc { font-size: 0.8rem; color: var(--text-muted); line-height: 1.5; flex: 1; margin-bottom: 0.85rem; }
.dest-pill { display: inline-flex; align-items: center; gap: 0.35rem; font-size: 0.7rem; font-weight: 600; color: var(--text-soft); padding: 0.22rem 0.55rem; border-radius: 6px; background: rgba(15,23,42,0.05); align-self: flex-start; }
.dest-tableau .swatch { color: #1F77B4; } .dest-ppt .swatch { color: #D24726; } .dest-xlsx .swatch { color: #107C41; } .dest-agent .swatch { color: #7C3AED; } .dest-doc .swatch { color: #475569; }
.badge { font-size: 0.6rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; padding: 0.2rem 0.5rem; border-radius: 5px; background: rgba(15,23,42,0.05); color: var(--text-muted); }
.badge.weekly { background: rgba(16,185,129,0.12); color: #047857; }
.badge.monthly { background: rgba(59,130,246,0.12); color: #1E40AF; }

/* ───── HERO BANNER ───── */
.hero { position: relative; background: radial-gradient(ellipse 90% 80% at 20% 20%, rgba(28,79,192,0.06) 0%, transparent 50%), radial-gradient(ellipse 60% 70% at 80% 80%, rgba(65,182,230,0.05) 0%, transparent 50%), linear-gradient(135deg, rgba(255,255,255,0.9) 0%, rgba(248,250,253,0.95) 100%); border-radius: 16px; padding: 2rem 2rem 1.6rem; border: 1px solid var(--hairline-2); box-shadow: var(--shadow-sm); overflow: hidden; }
.hero::before { content: ''; position: absolute; top: -1px; left: 0; right: 0; height: 3px; background: linear-gradient(90deg, var(--navy-600), var(--accent), var(--navy-500)); border-radius: 16px 16px 0 0; opacity: 0.7; }
.hero-header { display: flex; align-items: flex-start; justify-content: space-between; margin-bottom: 1.5rem; }
.hero-title { font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif; font-size: 1.65rem; font-weight: 800; color: var(--navy-900); letter-spacing: -0.025em; line-height: 1.15; margin-bottom: 0.35rem; }
.hero-subtitle { font-size: 0.82rem; font-weight: 500; color: var(--text-muted); display: flex; align-items: center; gap: 0.5rem; }
.hero-subtitle .dot { width: 4px; height: 4px; border-radius: 50%; background: var(--text-muted); opacity: 0.5; }
.hero-badge { display: inline-flex; align-items: center; gap: 0.4rem; font-size: 0.68rem; font-weight: 600; color: var(--navy-700); background: rgba(28,79,192,0.08); padding: 0.3rem 0.7rem; border-radius: 8px; flex-shrink: 0; margin-top: 0.2rem; }
.hero-badge svg { width: 12px; height: 12px; stroke: currentColor; fill: none; stroke-width: 2; }
.hero-kpis { display: grid; grid-template-columns: repeat(5, 1fr); gap: 0.75rem; }
.hero-kpi { background: #EEF5FB; border: 1px solid var(--hairline-2); border-radius: 12px; padding: 0.85rem 1rem 0.8rem; transition: transform 0.25s var(--ease-out), box-shadow 0.25s var(--ease); }
.hero-kpi:hover { transform: translateY(-2px); box-shadow: var(--shadow-md); }
.hero-kpi .kpi-label { font-size: 0.7rem; color: var(--text-muted); font-weight: 500; margin-bottom: 0.25rem; }
.hero-kpi .kpi-value { font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif; font-size: 1.5rem; font-weight: 700; color: var(--navy-900); line-height: 1.1; letter-spacing: -0.02em; font-variant-numeric: tabular-nums; margin-bottom: 0.3rem; }
.hero-kpi .kpi-delta { display: inline-flex; align-items: center; gap: 0.25rem; font-size: 0.7rem; font-weight: 600; font-variant-numeric: tabular-nums; }
.hero-kpi .kpi-delta.up { color: var(--up); } .hero-kpi .kpi-delta.down { color: var(--down); } .hero-kpi .kpi-delta.flat { color: var(--flat); }
.hero-kpi .kpi-delta .tri { font-size: 0.65rem; line-height: 1; }
.hero-kpi .kpi-delta .vs { color: var(--text-muted); font-weight: 500; }

.workspace-divider { height: 1px; background: var(--hairline); margin: 1.4rem 0; }

/* ───── SUB-PANEL (full content area) ───── */
.sub-panel { display: none; }
.sub-panel.is-active { display: block; }
.sub-panel-title { font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif; font-size: 1.25rem; font-weight: 700; color: var(--navy-900); margin-bottom: 0.3rem; letter-spacing: -0.02em; }
.sub-panel-desc { font-size: 0.84rem; color: var(--text-muted); margin-bottom: 1.4rem; max-width: 640px; }
.placeholder-chart { border: 2px dashed var(--hairline); border-radius: 12px; padding: 3rem 2rem; display: flex; align-items: center; justify-content: center; color: var(--text-muted); font-size: 0.85rem; font-weight: 500; background: var(--surface-2); margin-bottom: 1.2rem; min-height: 220px; }
.placeholder-table { border: 1px solid var(--hairline); border-radius: 10px; overflow: hidden; }
.placeholder-table-header { display: grid; grid-template-columns: 2fr 1fr 1fr 1fr; padding: 0.6rem 1rem; background: var(--surface-2); font-size: 0.7rem; font-weight: 600; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.06em; border-bottom: 1px solid var(--hairline); }
.placeholder-table-row { display: grid; grid-template-columns: 2fr 1fr 1fr 1fr; padding: 0.55rem 1rem; font-size: 0.8rem; color: var(--text-soft); border-bottom: 1px solid var(--hairline-2); }
.placeholder-table-row:last-child { border-bottom: none; }

/* ───── RESPONSIVE ───── */
@media (max-width: 1180px) { :root { --sidebar-w: 208px; } }
@media (max-width: 980px) { .grid { grid-template-columns: repeat(2, 1fr); } .hero-kpis { grid-template-columns: repeat(3, 1fr); } }
@media (max-width: 760px) { :root { --sidebar-w: 60px; } .sidebar-brand .title, .sidebar-brand .subtitle, .sidebar-section-label, .sidebar-meta, .nav-item .nav-label, .nav-item .nav-count, .nav-sub { display: none; } .nav-item { justify-content: center; padding: 0.6rem 0; } .nav { padding: 0 0.3rem; } }
@media (max-width: 560px) { .grid { grid-template-columns: 1fr; } body { overflow: auto; } .app { height: auto; grid-template-columns: 1fr; padding: 0; gap: 0; } .sidebar { display: none; } .main { border-radius: 0; border: none; } .main, .content { overflow: visible; } .hero-kpis { grid-template-columns: repeat(2, 1fr); } .hero-title { font-size: 1.35rem; } .hero { padding: 1.4rem 1.2rem 1.2rem; } .hero-header { flex-direction: column; gap: 0.6rem; } }
@media (prefers-reduced-motion: reduce) { *, *::before, *::after { animation-duration: 0.01ms !important; transition-duration: 0.01ms !important; } }

/* ????? TOOLBAR ????? */
.toolbar {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0.7rem 1.4rem;
    border-bottom: 1px solid var(--hairline);
    flex-shrink: 0;
    z-index: 5;
}
.toolbar-title {
    font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif;
    font-size: 0.78rem;
    color: var(--navy-700);
    font-weight: 500;
}
.toolbar-title strong { color: var(--navy-900); font-weight: 600; }
.toolbar-actions { display: flex; align-items: center; gap: 0.45rem; }

/* ????? KPI STRIP ????? */
.kpis-head {
    font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif;
    font-size: 0.62rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    color: var(--text-muted);
    margin-bottom: 0.55rem;
}
.kpi-row {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 0.85rem;
    margin-bottom: 1.6rem;
}
.kpi {
    background: var(--surface);
    border-radius: 14px;
    padding: 1rem 1.15rem 1.05rem;
    box-shadow: var(--shadow-sm);
    transition: transform 0.25s var(--ease-out), box-shadow 0.25s var(--ease);
    position: relative;
    overflow: hidden;
}
.kpi::before {
    content: '';
    position: absolute;
    inset: 0;
    background: linear-gradient(135deg, rgba(255,255,255,0) 60%, rgba(28,79,192,0.04) 100%);
    opacity: 0;
    transition: opacity 0.25s var(--ease);
    pointer-events: none;
}
.kpi:hover { transform: translateY(-2px); box-shadow: var(--shadow-md); }
.kpi:hover::before { opacity: 1; }
.kpi .kpi-label {
    font-size: 0.74rem;
    color: var(--text-muted);
    font-weight: 500;
    margin-bottom: 0.3rem;
}
.kpi .kpi-value {
    font-family: 'Pfizer Diatype Office', Arial, Helvetica, sans-serif;
    font-size: 1.7rem;
    font-weight: 700;
    color: var(--navy-900);
    line-height: 1.05;
    letter-spacing: -0.025em;
    font-variant-numeric: tabular-nums;
    margin-bottom: 0.4rem;
}
.kpi .kpi-delta {
    display: inline-flex;
    align-items: center;
    gap: 0.3rem;
    font-size: 0.74rem;
    font-weight: 600;
    font-variant-numeric: tabular-nums;
}
.kpi .kpi-delta.up { color: var(--up); }
.kpi .kpi-delta.down { color: var(--down); }
.kpi .kpi-delta.flat { color: var(--flat); }
.kpi .kpi-delta .tri { font-size: 0.7rem; line-height: 1; }
.kpi .kpi-delta .vs { color: var(--text-muted); font-weight: 500; }

/* Card enhancements */
.card:active { transform: translateY(-1px) scale(0.995); transition-duration: 0.08s; }
.card:hover .icon-chip { transform: scale(1.05) rotate(-2deg); }
.card:hover .dest-pill { background: rgba(28,79,192,0.10); color: var(--navy-700); }
.card-updated { font-size: 0.68rem; color: var(--text-muted); font-weight: 500; margin-top: -0.65rem; margin-bottom: 0.7rem; opacity: 0.85; }
</style>
</head>
<body>

<div class="app">

<!-- ═══ SIDEBAR ═══ -->
<aside class="sidebar">
    <div class="sidebar-brand">
        <img src="https://cdn.pfizer.com/pfizercom/2022-10/Pfizer_Logo_Color_CMYK.png" alt="Pfizer">
        <div>
            <div class="title">Migraine<br>Intelligence Hub</div>
            <div class="subtitle">Nurtec&reg; ODT</div>
        </div>
    </div>
    <div class="sidebar-divider"></div>
    <div class="sidebar-section-label">Workspace</div>
    <nav class="nav" id="sidebarNav">
        <button class="nav-item active" data-target="dashboards">
            <span class="nav-icon"><svg viewBox="0 0 24 24"><rect x="3" y="3" width="7" height="9" rx="1.5"/><rect x="14" y="3" width="7" height="5" rx="1.5"/><rect x="14" y="12" width="7" height="9" rx="1.5"/><rect x="3" y="16" width="7" height="5" rx="1.5"/></svg></span>
            <span class="nav-label">Deep-Dive Dashboards</span>
            <span class="nav-count">6</span>
        </button>
        <button class="nav-item" data-target="newsletter">
            <span class="nav-icon"><svg viewBox="0 0 24 24"><path d="M4 4h16v16H4z"/><path d="M4 9h16M9 4v16"/></svg></span>
            <span class="nav-label">Weekly Newsletter</span>
            <span class="nav-count">6</span>
        </button>
        <!-- Sub-nav items (indented under Weekly Newsletter) -->
        <div class="nav-sub" id="newsletterSub">
            <button class="nav-sub-item active" data-panel="nl-exec">Executive Summary</button>
            <button class="nav-sub-item" data-panel="nl-xponent">Xponent Trends</button>
            <button class="nav-sub-item" data-panel="nl-market">Market Insights</button>
            <button class="nav-sub-item" data-panel="nl-access">Access Changes (WIP)</button>
            <a class="nav-sub-item external" href="https://dss-amer-design.pfizer.com/webapps/MIGRAINEDOPPLR/zXimTdd/" target="_blank" rel="noopener">DOppLR Alerts <svg viewBox="0 0 24 24"><path d="M18 13v6a2 2 0 01-2 2H5a2 2 0 01-2-2V8a2 2 0 012-2h6M15 3h6v6M10 14L21 3"/></svg></a>
            <button class="nav-sub-item" data-panel="nl-financial">Financial Summary</button>
        </div>
        <button class="nav-item" data-target="agents">
            <span class="nav-icon"><svg viewBox="0 0 24 24"><rect x="6" y="6" width="12" height="10" rx="2"/><path d="M9 16v3M15 16v3M9 6V3M15 6V3M3 11h3M18 11h3"/></svg></span>
            <span class="nav-label">CoWork Agents</span>
            <span class="nav-count">3</span>
        </button>
        <button class="nav-item" data-target="deliverables">
            <span class="nav-icon"><svg viewBox="0 0 24 24"><path d="M14 3v5h5M5 21h14a1 1 0 001-1V8l-5-5H5a1 1 0 00-1 1v16a1 1 0 001 1z"/></svg></span>
            <span class="nav-label">Analytics Deliverables</span>
            <span class="nav-count">3</span>
        </button>
        <button class="nav-item" data-target="docs">
            <span class="nav-icon"><svg viewBox="0 0 24 24"><path d="M5 3h11l3 3v15a1 1 0 01-1 1H5a1 1 0 01-1-1V4a1 1 0 011-1z"/><path d="M9 9h6M9 13h6M9 17h4"/></svg></span>
            <span class="nav-label">Business Rule Docs</span>
            <span class="nav-count">5</span>
        </button>
    </nav>
    <div class="sidebar-spacer"></div>
    <div class="sidebar-meta">
        <div class="meta-row"><strong>IIS Migraine Analytics</strong></div>
        <div class="meta-row"><a href="mailto:Team_ZS_US_Migraine_Analytics@zs.com">Team_ZS_US_Migraine_Analytics@zs.com</a></div>
        <div class="meta-row">Updated APP_LAST_UPDATED_DATE</div>
    </div>
</aside>

<!-- ═══ MAIN ═══ -->
<div class="main">
    
    <!-- TOOLBAR -->
    <div class="toolbar">
        <div class="toolbar-title">Workspace &middot; <strong id="toolbarSection">Deep-Dive Dashboards</strong></div>
        <div class="toolbar-actions">
            <button class="icon-btn" title="Help">
                <svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="9"/><path d="M9.5 9a2.5 2.5 0 015 0c0 1.5-2.5 2-2.5 4M12 17v.01"/></svg>
                Help
            </button>
        </div>
    </div>

    <main class="content">
        


        <!-- HERO BANNER -->
        <div class="hero">
            <div class="hero-header">
                <div class="hero-text">
                    <h1 class="hero-title">Nurtec Performance YTD</h1>
                    <div class="hero-subtitle"><span>Nurtec&reg; ODT</span><span class="dot"></span><span>IIS Analytics</span></div>
                </div>
                <span class="hero-badge"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="3"/><path d="M12 5V3M12 21v-2M16.95 7.05l1.41-1.41M5.64 18.36l1.41-1.41M19 12h2M3 12h2M16.95 16.95l1.41 1.41M5.64 5.64l1.41 1.41"/></svg>Executive KPIs</span>
            </div>
            <div class="hero-kpis">
                <div class="hero-kpi"><div class="kpi-label">Nurtec TRx</div><div class="kpi-value">HERO_TRX_YTD_VAL</div><div class="kpi-delta up"><span class="tri">&#9650;</span>HERO_TRX_GROWTH_VAL <span class="vs">vs STLY</span></div></div>
                <div class="hero-kpi"><div class="kpi-label">Nurtec NBRx</div><div class="kpi-value">HERO_NBRX_YTD_VAL</div><div class="kpi-delta up"><span class="tri">&#9650;</span>HERO_NBRX_GROWTH_VAL <span class="vs">vs STLY</span></div></div>
                <div class="hero-kpi"><div class="kpi-label">Nurtec oCGRP Mkt Share</div><div class="kpi-value">HERO_OVERALL_MS_VAL</div><div class="kpi-delta up"><span class="tri">&#9650;</span>HERO_OVERALL_MS_DELTA <span class="vs">vs STLY</span></div></div>
                <div class="hero-kpi"><div class="kpi-label">Acute Nurtec oCGRP Share</div><div class="kpi-value">HERO_ACUTE_MS_VAL</div><div class="kpi-delta up"><span class="tri">&#9650;</span>HERO_ACUTE_MS_DELTA <span class="vs">vs STLY</span></div></div>
                <div class="hero-kpi"><div class="kpi-label">Preventive Nurtec oCGRP Share</div><div class="kpi-value">HERO_PREV_MS_VAL</div><div class="kpi-delta down"><span class="tri">&#9660;</span>HERO_PREV_MS_DELTA <span class="vs">vs STLY</span></div></div>
            </div>
        </div>

        <div class="workspace-divider"></div>

        <!-- 1. DASHBOARDS -->
        <section class="section is-active is-visible" id="dashboards" data-label="Deep-Dive Dashboards">
            <div class="section-head">
                <div class="section-head-row">
                    <h2>Deep-Dive Dashboards</h2>
                    <div class="dropdown-wrap">
                        <button class="icon-btn" onclick="toggleDropdown('dataDropdown', event)"><svg viewBox="0 0 24 24"><ellipse cx="12" cy="6" rx="8" ry="3"/><path d="M4 6v6c0 1.7 3.6 3 8 3s8-1.3 8-3V6"/><path d="M4 12v6c0 1.7 3.6 3 8 3s8-1.3 8-3v-6"/></svg>Data Availability<svg viewBox="0 0 24 24" style="width:11px;height:11px;"><path d="M6 9l6 6 6-6"/></svg></button>
                        <div class="dropdown" id="dataDropdown">
                            <div class="dropdown-header">Last refresh by source</div>
                            DATA_REFRESH_ITEMS_PLACEHOLDER
                        </div>
                    </div>
                </div>
                <p>Interactive analytics across patient funnel, volume, HCP, payer, and financial performance.</p>
            </div>
            <div class="grid">
                <a class="card" href="https://us-east-1.online.tableau.com/#/site/amerdev/views/USMigraineDashboard/PatientFunnel?:iid=1" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s1"><svg viewBox="0 0 24 24"><path d="M5.5 5h13a1 1 0 011 1v10a1 1 0 01-1 1h-13a1 1 0 01-1-1V6a1 1 0 011-1zM4 20h16M10 5V3M14 5V3"/></svg></span></div><div class="card-title">Patient Funnel</div><div class="card-desc">Eligibility through treatment at claims and patient-level views.</div><span class="dest-pill dest-tableau"><span class="swatch">&#9632;</span>Tableau</span></a>
                <a class="card" href="https://us-east-1.online.tableau.com/#/site/amerdev/views/USMigraineConcurrentUsageTab/ConcurrentUsageTab" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s1"><svg viewBox="0 0 24 24"><circle cx="12" cy="7" r="4"/><path d="M6 21v-2a4 4 0 014-4h4a4 4 0 014 4v2"/></svg></span></div><div class="card-title">Patient Deep Dive</div><div class="card-desc">Concurrent usage, switching patterns, and patient segmentation.</div><span class="dest-pill dest-tableau"><span class="swatch">&#9632;</span>Tableau</span></a>
                <a class="card" href="https://us-east-1.online.tableau.com/#/site/amerdev/views/USMigraineDashboard/VolumeDeepDive" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s1"><svg viewBox="0 0 24 24"><path d="M3 12h4l3-9 4 18 3-9h4"/></svg></span></div><div class="card-title">Volume Deep Dive</div><div class="card-desc">TRx and NBRx volume trends across brands and channels.</div><span class="dest-pill dest-tableau"><span class="swatch">&#9632;</span>Tableau</span></a>
                <a class="card" href="https://us-east-1.online.tableau.com/#/site/amerdev/views/USMigraineDashboard/PrescriptionInsights?:iid=1" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s1"><svg viewBox="0 0 24 24"><path d="M12 4a5 5 0 015 5v3a1 1 0 01-1 1H8a1 1 0 01-1-1V9a5 5 0 015-5z"/><path d="M12 13v7M9 20h6"/></svg></span></div><div class="card-title">HCP Deep Dive</div><div class="card-desc">Prescriber behavior, decile analysis, and targeting insights.</div><span class="dest-pill dest-tableau"><span class="swatch">&#9632;</span>Tableau</span></a>
                <a class="card" href="https://us-east-1.online.tableau.com/#/site/amerdev/views/USMigraineDashboardPayerDeepDive/PayerSummary?:iid=1" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s1"><svg viewBox="0 0 24 24"><path d="M3 21h18M5 21V7l8-4v18M19 21V11l-6-4"/></svg></span></div><div class="card-title">Payer Deep Dive</div><div class="card-desc">Formulary status, payer mix, and access rates.</div><span class="dest-pill dest-tableau"><span class="swatch">&#9632;</span>Tableau</span></a>
                <a class="card" href="https://us-east-1.online.tableau.com/#/site/amerdev/views/USMigraineDashboard/FinancialSummary?:iid=1" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s1"><svg viewBox="0 0 24 24"><path d="M12 2v20M5 9h11a3 3 0 010 6H7a3 3 0 000 6h12"/></svg></span></div><div class="card-title">Financial Summary</div><div class="card-desc">Revenue, cost, and financial KPIs across the portfolio.</div><span class="dest-pill dest-tableau"><span class="swatch">&#9632;</span>Tableau</span></a>
            </div>
        </section>

        <!-- 2. NEWSLETTER -->
        <section class="section" id="newsletter" data-label="Weekly Newsletter">
            <!-- Card grid view (initial state when clicking Weekly Newsletter) -->
            <div id="nl-grid-view">
                <div class="section-head">
                    <h2>Weekly Newsletter Deep-Dive</h2>
                    <p>In-depth weekly analyses covering market momentum, competitive share shifts, and performance trends.</p>
                </div>
                <div class="grid">
                    <div class="card" data-nl-panel="nl-exec"><div class="card-top"><span class="icon-chip chip-s2"><svg viewBox="0 0 24 24"><path d="M4 19h16M4 15l4-6 4 2 4-5 4 4"/></svg></span><span class="badge weekly">Weekly</span></div><div class="card-title">Executive Summary</div><div class="card-desc">High-level performance overview and key takeaways.</div><span class="dest-pill dest-ppt"><span class="swatch">&#9632;</span>PowerPoint</span></div>
                    <div class="card" data-nl-panel="nl-xponent"><div class="card-top"><span class="icon-chip chip-s2"><svg viewBox="0 0 24 24"><path d="M3 17l6-6 4 4 8-8M14 7h7v7"/></svg></span><span class="badge weekly">Weekly</span></div><div class="card-title">Xponent Trends</div><div class="card-desc">Weekly Xponent data trends and brand performance.</div><span class="dest-pill dest-ppt"><span class="swatch">&#9632;</span>PowerPoint</span></div>
                    <div class="card" data-nl-panel="nl-market"><div class="card-top"><span class="icon-chip chip-s2"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="10"/><path d="M2 12h20M12 2a15.3 15.3 0 014 10 15.3 15.3 0 01-4 10 15.3 15.3 0 01-4-10 15.3 15.3 0 014-10z"/></svg></span><span class="badge weekly">Weekly</span></div><div class="card-title">Market Insights</div><div class="card-desc">Competitive landscape and market share dynamics.</div><span class="dest-pill dest-ppt"><span class="swatch">&#9632;</span>PowerPoint</span></div>
                    <div class="card" data-nl-panel="nl-access"><div class="card-top"><span class="icon-chip chip-s2"><svg viewBox="0 0 24 24"><path d="M9 12l2 2 4-4m5-2a9 9 0 11-4.219-7.619"/></svg></span><span class="badge weekly">Weekly</span></div><div class="card-title">Access Changes</div><div class="card-desc">Payer access updates and formulary change tracking.</div><span class="dest-pill dest-ppt"><span class="swatch">&#9632;</span>PowerPoint</span></div>
                    <a class="card" href="https://dss-amer-design.pfizer.com/webapps/MIGRAINEDOPPLR/zXimTdd/" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s2"><svg viewBox="0 0 24 24"><path d="M10 5a2 2 0 114 0v1a7 7 0 014 6v4l2 2H4l2-2v-4a7 7 0 014-6V5z"/><path d="M9 17v1a3 3 0 006 0v-1"/></svg></span><span class="badge weekly">Weekly</span></div><div class="card-title">DOppLR Alerts</div><div class="card-desc">Predictive alerts for market momentum shifts.</div><span class="dest-pill dest-ppt"><span class="swatch">&#9632;</span>PowerPoint</span></a>
                    <div class="card" data-nl-panel="nl-financial"><div class="card-top"><span class="icon-chip chip-s2"><svg viewBox="0 0 24 24"><path d="M5 21V5a2 2 0 012-2h10a2 2 0 012 2v16l-7-3-7 3z"/></svg></span><span class="badge weekly">Weekly</span></div><div class="card-title">Financial Summary</div><div class="card-desc">Weekly financial metrics and revenue tracking.</div><span class="dest-pill dest-xlsx"><span class="swatch">&#9632;</span>Excel</span></div>
                </div>
            </div>
            <!-- Deep-dive sub-panels (shown when a card is clicked) -->
            <div id="nl-deepdive-view" style="display:none;">
<style>
.nl-content .card { background: white; border: 1px solid #e5e7eb; border-radius: 10px; padding: 20px; margin-bottom: 16px; min-height: auto; box-shadow: none; cursor: default; }
.nl-content .card:hover { transform: none; box-shadow: none; }
.nl-content .card::after { display: none; }
.nl-content .card-title { font-size: 17px; font-weight: 700; margin-bottom: 6px; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; color: #1a2332; }
.nl-content .card-subtitle { font-size: 12px; color: #4b5563; margin-bottom: 16px; }
.nl-content .row { display: flex; gap: 16px; margin-bottom: 16px; }
.nl-content .row > * { flex: 1; }
.nl-content .stat-tile { background: #f9fafb; border: 1px solid #e5e7eb; border-radius: 8px; padding: 16px; }
.nl-content .stat-tile .label { font-size: 13px; color: #374151; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 6px; font-weight: 700; }
.nl-content .label { font-size: 13px; color: #374151; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 6px; font-weight: 700; }
.nl-content .stat-tile .value { font-size: 28px; font-weight: 700; color: #1a2332; }
.nl-content .stat-tile .sub { font-size: 10px; color: #9ca3af; margin-top: 4px; }
.nl-content .sub { font-size: 10px; color: #9ca3af; margin-top: 4px; }
.nl-content .value { font-size: 28px; font-weight: 700; color: #1a2332; }
.nl-content table { width: 100%; border-collapse: collapse; font-size: 12px; table-layout: fixed; }
.nl-content th { text-align: left; padding: 10px 12px; border-bottom: 2px solid #e5e7eb; font-size: 10px; text-transform: uppercase; color: #6b7280; letter-spacing: 0.5px; font-weight: 600; }
.nl-content td { padding: 10px 12px; border-bottom: 1px solid #f3f4f6; }
.nl-content tr:last-child td { border-bottom: none; }
.nl-content table th:first-child, .nl-content table td:first-child { width: 18%; }
.nl-content table th:nth-child(2), .nl-content table td:nth-child(2) { width: 10%; }
.nl-content table th:nth-child(3), .nl-content table td:nth-child(3) { width: 10%; }
.nl-content table th:nth-child(4), .nl-content table td:nth-child(4) { width: 11%; }
.nl-content table th:nth-child(5), .nl-content table td:nth-child(5) { width: 9%; }
.nl-content table th:nth-child(6), .nl-content table td:nth-child(6) { width: 9%; }
.nl-content table th:nth-child(7), .nl-content table td:nth-child(7) { width: 9%; }
.nl-content table th:nth-child(8), .nl-content table td:nth-child(8) { width: 8%; }
.nl-content table th:nth-child(9), .nl-content table td:nth-child(9) { width: 8%; }
.nl-content table th:nth-child(10), .nl-content table td:nth-child(10) { width: 8%; }
.nl-content .pill-group { display: inline-flex; gap: 0; margin-bottom: 16px; border: 2px solid #0000C9; border-radius: 50px; overflow: hidden; }
.nl-content .pill { padding: 8px 20px; border: none; cursor: pointer; font-size: 12px; font-weight: 600; background: white; color: #0000C9; }
.nl-content .pill:first-child { border-radius: 50px 0 0 50px; }
.nl-content .pill:last-child { border-radius: 0 50px 50px 0; }
.nl-content .pill.active { background: #0000C9; color: white; }
.nl-content .pill-sm { padding: 6px 14px; font-size: 11px; }
.nl-content .chart-container { position: relative; width: 100%; height: 280px; margin: 16px 0; margin-left: -20px; margin-right: -20px; width: calc(100% + 40px); }
.nl-content .chart-container-sm { height: 220px; margin-left: -20px; margin-right: -20px; width: calc(100% + 40px); }
.nl-content svg.chart { width: 100%; height: 100%; }
.nl-content svg.chart text { fill: #4b5563 !important; font-weight: 500 !important; font-size: 10px !important; }
.nl-content .legend { display: flex; gap: 16px; justify-content: center; margin-top: 8px; font-size: 11px; flex-wrap: wrap; }
.nl-content .axis-info { display: flex; justify-content: space-between; font-size: 9px; font-weight: 500; color: #6b7280; margin-top: 4px; margin-bottom: 6px; padding: 0 8px; }
.nl-content .axis-info span { font-style: italic; }
.nl-content .axis-label { font-size: 10px; font-weight: 600; color: #4b5563; text-align: center; margin-top: -8px; margin-bottom: 8px; }
.nl-content .legend-item { display: flex; align-items: center; gap: 6px; }
.nl-content .legend-dot { width: 8px; height: 8px; border-radius: 50%; }
.nl-content .brand-header { padding: 10px 16px; border-radius: 6px; margin-bottom: 12px; color: white; font-size: 12px; font-weight: 600; display: flex; gap: 16px; }
.nl-content .brand-header.brand-a { background: #7C6CFC; }
.nl-content .brand-header.brand-b { background: #16a34a; }
.nl-content .brand-header.brand-c { background: #ea580c; }
.nl-content .insight-bullet { display: flex; gap: 10px; margin-bottom: 12px; align-items: flex-start; }
.nl-content .insight-dot { width: 8px; height: 8px; border-radius: 50%; background: #f59e0b; margin-top: 4px; flex-shrink: 0; }
.nl-content .alert-item { display: flex; gap: 10px; margin-bottom: 14px; align-items: flex-start; }
.nl-content .alert-icon { width: 20px; height: 20px; flex-shrink: 0; }
.nl-content .alert-signal { font-weight: 600; font-size: 12px; }
.nl-content .alert-anchor { font-size: 11px; color: #6b7280; }
.nl-content .badge { padding: 2px 8px; border-radius: 4px; font-size: 10px; font-weight: 600; }
.nl-content .badge-gain { background: #dcfce7; color: #166534; }
.nl-content .badge-loss { background: #fee2e2; color: #991b1b; }
.nl-content .badge-restriction { background: #fef3c7; color: #92400e; }
.nl-content .delta-pos { color: #16a34a; }
.nl-content .delta-neg { color: #dc2626; }
.nl-content .card-header-row { display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 16px; }
.nl-content .footnote { font-size: 10px; color: #9ca3af; margin-top: 12px; padding-top: 8px; border-top: 1px solid #f3f4f6; }
.nl-content .split-section { display: flex; gap: 24px; margin-bottom: 16px; }
.nl-content .split-section > .split-col { flex: 1; min-width: 0; }
.nl-content .split-section.chart-split { position: relative; }
.nl-content .split-section.chart-split::after { content: ''; position: absolute; top: 5%; bottom: 5%; left: 50%; transform: translateX(-50%); width: 0; border-left: 1px dotted #c7d2fe; }
.nl-content .split-col .section-label { font-size: 13px; font-weight: 700; margin-bottom: 12px; color: #1a2332; text-align: center; }
.nl-content .filter-select { padding: 6px 12px; border: 1px solid #d1d5db; border-radius: 6px; font-size: 11px; background: white; }
.nl-content .sub-signal-table td { padding: 6px 12px; font-size: 11px; }
.nl-content .nl-tab { display: none; }
.nl-content .nl-tab.active { display: block; }
</style>
<div class="nl-content">
<div class="nl-tab active" id="nl-exec-tab">
<div style="padding:8px 32px 24px;">
<div class="section-head"><h2>Executive Summary</h2><p>High-level performance overview and key takeaways for the current reporting week.</p></div>

    <!-- KPI CARDS ROW -->
    <div style="display:grid;grid-template-columns:repeat(6,1fr);gap:16px;margin-bottom:32px;">
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);cursor:default;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div style="font-size:28px;font-weight:800;color:#0000C9;margin-bottom:6px;">EXEC_NBRX_WK_VAL</div>
        <div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:2px;">Nurtec NBRx</div>
        <div style="font-size:11px;color:#16a34a;font-weight:600;">EXEC_NBRX_WK_GROWTH WoW</div>
        <div style="font-size:10px;color:#9ca3af;margin-top:4px;">Latest Week</div>
      </div>
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);cursor:default;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div style="font-size:28px;font-weight:800;color:#0000C9;margin-bottom:6px;">EXEC_TRX_WK_VAL</div>
        <div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:2px;">Nurtec TRx</div>
        <div style="font-size:11px;color:#16a34a;font-weight:600;">EXEC_TRX_WK_GROWTH WoW</div>
        <div style="font-size:10px;color:#9ca3af;margin-top:4px;">Latest Week</div>
      </div>
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);cursor:default;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div style="font-size:28px;font-weight:800;color:#0000C9;margin-bottom:6px;">EXEC_MS_WK_VAL</div>
        <div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:2px;">oCGRP TRx Share</div>
        <div style="font-size:11px;color:#16a34a;font-weight:600;">EXEC_MS_WK_GROWTH WoW</div>
        <div style="font-size:10px;color:#9ca3af;margin-top:4px;">National &middot; All Segments</div>
      </div>
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);cursor:default;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div style="font-size:28px;font-weight:800;color:#0000C9;margin-bottom:6px;">EXEC_GROSS_WK_VAL</div>
        <div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:2px;">Gross Sales W/E</div>
        <div style="font-size:11px;color:#16a34a;font-weight:600;">EXEC_GROSS_VS_PY vs PY</div>
        <div style="font-size:10px;color:#9ca3af;margin-top:4px;">Latest Week</div>
      </div>
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);cursor:default;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div style="font-size:28px;font-weight:800;color:#0000C9;margin-bottom:6px;">EXEC_GOAL_ATT_VAL</div>
        <div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:2px;">YTD Goal Attainment</div>
        <div style="font-size:11px;color:#d97706;font-weight:600;">TRx vs OP26</div>
        <div style="font-size:10px;color:#9ca3af;margin-top:4px;">Nurtec TRx</div>
      </div>
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);cursor:default;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div style="font-size:28px;font-weight:800;color:#0000C9;margin-bottom:6px;">128M</div>
        <div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:2px;">Covered Lives w/ Access</div>
        <div style="font-size:11px;color:#16a34a;font-weight:600;">+1.2M vs prior month</div>
        <div style="font-size:10px;color:#9ca3af;margin-top:4px;">As of Jun 2026</div>
      </div>
    </div>

    <!-- PERFORMANCE INSIGHTS -->
    <div style="background:#ffffff;border:1px solid #e8ecf0;border-left:4px solid #0000C9;border-radius:8px;padding:24px 28px;margin-bottom:24px;">
      <div style="font-size:18px;font-weight:700;color:#1a2332;margin-bottom:6px;">Performance Insights</div>
      <div style="font-size:12px;color:#6b7280;margin-bottom:20px;">Key observations for week ending EXEC_PERF_WEEK_DATE &middot; Numbers in bold update weekly</div>

      <div style="font-size:14px;font-weight:700;color:#1a2332;margin-bottom:4px;text-decoration:underline;">Weekly Insights</div>
      <div style="font-size:11px;color:#6b7280;margin-bottom:12px;">Week-over-week changes and weekly-level observations</div>
      <ul style="list-style-type:disc;padding-left:20px;margin-bottom:16px;color:#0000C9;">
EXEC_WEEKLY_INSIGHTS_BULLETS
      </ul>

      <hr style="border:none;border-top:1px solid #e8ecf0;margin:24px 0;">

      <div style="font-size:14px;font-weight:700;color:#1a2332;margin-bottom:4px;text-decoration:underline;">Year to Date Insights</div>
      <div style="font-size:11px;color:#6b7280;margin-bottom:12px;">Cumulative performance statements as of week ending EXEC_PERF_WEEK_DATE</div>
      <ul style="list-style-type:disc;padding-left:20px;margin-bottom:16px;color:#0000C9;">
EXEC_YTD_INSIGHTS_BULLETS
      </ul>
    </div>

    <!-- PERFORMANCE SNAPSHOT + ACCESS CHANGES -->
    <div style="display:flex;gap:20px;margin-bottom:24px;">
      <div style="flex:2;background:#ffffff;border:1px solid #e8ecf0;border-left:4px solid #0000C9;border-radius:8px;padding:24px 28px;">
        <div style="font-size:16px;font-weight:700;margin-bottom:4px;">Performance Snapshot</div>
        <div style="font-size:11px;color:#6b7280;margin-bottom:16px;">Week ending EXEC_PERF_WEEK_DATE &middot; National</div>
        <table>
          <thead>
            <tr><th>METRIC</th><th style="text-align:right">LATEST WK</th><th style="text-align:right">R4W AVG</th><th style="text-align:right">YTD</th><th style="text-align:right">VS. GOAL</th><th style="text-align:right">VS. STLY</th></tr>
          </thead>
          <tbody>
EXEC_PERF_SNAPSHOT_ROWS
          </tbody>
        </table>
      </div>

      <div style="flex:1;background:#ffffff;border:1px solid #e8ecf0;border-left:4px solid #0000C9;border-radius:8px;padding:20px 24px;">
        <div style="font-size:16px;font-weight:700;margin-bottom:4px;">Access Change Lines</div>
        <div style="font-size:11px;color:#6b7280;margin-bottom:16px;">Only appears with material weekly access changes</div>
        <div style="margin-bottom:14px;">
          <div><strong>Louisiana Medicaid</strong></div>
          <div style="color:#6b7280;font-size:12px;">Nurtec removed from PDL &middot; est. <strong style="color:#0000C9;">-2,200 NRx</strong></div>
        </div>
        <div>
          <div><strong>UnitedHealth Commercial (IL)</strong></div>
          <div style="color:#6b7280;font-size:12px;">PA requirement added &middot; <strong style="color:#0000C9;">820K</strong> covered lives</div>
        </div>
      </div>
    </div>

  </div>
</div>
<div class="nl-tab" id="nl-xponent-tab">
<div style="padding:8px 32px 24px;">
<div class="section-head"><h2>Xponent Trends</h2><p>Weekly Xponent data trends, segment and channel performance analysis.</p></div>
    <div class="card">
      <div class="card-title">Claims MTD & YTD vs. STLY</div>
      <div class="card-subtitle">Nurtec NRx and TRx claim counts</div>
      
    

<div class="row">
        <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''"><div class="label">NRX MTD</div><div class="value" style="color:#0000C9;">XPT_NRX_MTD_VAL</div><div class="sub">XPT_NRX_MTD_STLY vs STLY</div></div>
        <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''"><div class="label">NRX YTD</div><div class="value" style="color:#0000C9;">XPT_NRX_YTD_VAL</div><div class="sub">XPT_NRX_YTD_STLY vs STLY</div></div>
        <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''"><div class="label">TRX MTD</div><div class="value" style="color:#0000C9;">XPT_TRX_MTD_VAL</div><div class="sub">XPT_TRX_MTD_STLY vs STLY</div></div>
        <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''"><div class="label">TRX YTD</div><div class="value" style="color:#0000C9;">XPT_TRX_YTD_VAL</div><div class="sub">XPT_TRX_YTD_STLY vs STLY</div></div>
      </div>
    </div>

    <div class="pill-group" id="xponent-metric-toggle" style="margin-bottom:16px;">
        <div class="pill pill-sm active" id="xp-trx-pill" onclick="switchXponentMetric('trx')">TRx</div>
        <div class="pill pill-sm" id="xp-nrx-pill" onclick="switchXponentMetric('nrx')">NRx</div>
      </div>
      <div style="display:flex;justify-content:flex-end;margin-bottom:6px;margin-top:-10px;align-items:center;gap:10px;">
        <div style="position:relative;display:inline-block;">
          <div style="width:28px;height:28px;border-radius:50%;background:#ffffff;color:#0000C9;font-size:16px;font-weight:800;display:flex;align-items:center;justify-content:center;cursor:pointer;border:1px solid #c7d2fe;" onmouseover="this.nextElementSibling.style.display='block'" onmouseout="this.nextElementSibling.style.display='none'">i</div>
          <div style="display:none;position:absolute;right:0;top:32px;background:#1a2332;color:#fff;padding:10px 14px;border-radius:6px;font-size:11px;line-height:1.5;width:240px;z-index:10;box-shadow:0 4px 12px rgba(0,0,0,0.15);">The Shares in this view are the Nurtec oCGRP shares across Payer's or Channel's</div>
        </div>
        <div class="dropdown-wrap" style="position:relative;">
          <button class="icon-btn" style="font-size:11px;padding:5px 12px;border-radius:6px;background:#0000C9;color:#fff;border:none;cursor:pointer;display:flex;align-items:center;gap:5px;font-weight:700;" onclick="toggleDropdown('xptDownloadDD', event)">
            <svg viewBox="0 0 24 24" style="width:13px;height:13px;fill:none;stroke:currentColor;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
            <svg viewBox="0 0 24 24" style="width:10px;height:10px;fill:none;stroke:currentColor;stroke-width:2;"><path d="M6 9l6 6 6-6"/></svg>
          </button>
          <div class="dropdown" id="xptDownloadDD" style="min-width:200px;">
            <a class="dropdown-item" id="xpt-dl-payer" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,XPT_DL_PAYER_TRX_B64" download="Xponent_TRx_Share_by_Payer.xlsx" style="text-decoration:none;color:inherit;display:flex;justify-content:space-between;" data-xpt-dl-label="{XM} Share by Payer" data-trx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,XPT_DL_PAYER_TRX_B64" data-nrx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,XPT_DL_PAYER_NRX_B64" data-trx-fname="Xponent_TRx_Share_by_Payer.xlsx" data-nrx-fname="Xponent_NRx_Share_by_Payer.xlsx"><span>TRx Share by Payer</span><svg viewBox="0 0 24 24" style="width:12px;height:12px;fill:none;stroke:#6b7280;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg></a>
            <a class="dropdown-item" id="xpt-dl-channel" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,XPT_DL_CH_TRX_B64" download="Xponent_TRx_Share_by_Channel.xlsx" style="text-decoration:none;color:inherit;display:flex;justify-content:space-between;" data-xpt-dl-label="{XM} Share by Channel" data-trx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,XPT_DL_CH_TRX_B64" data-nrx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,XPT_DL_CH_NRX_B64" data-trx-fname="Xponent_TRx_Share_by_Channel.xlsx" data-nrx-fname="Xponent_NRx_Share_by_Channel.xlsx"><span>TRx Share by Channel</span><svg viewBox="0 0 24 24" style="width:12px;height:12px;fill:none;stroke:#6b7280;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg></a>
          </div>
        </div>
      </div>
      <div class="row" style="margin-bottom:0;">
      
      <div class="card">
        <div class="card-title" data-xp-metric="{XM} Share Trend by Payer">TRx Share Trend by Payer</div>
        <div class="card-subtitle">Commercial / Medicare / Medicaid / Other · weekly</div>
        <div id="xpt-payer-trx" style="width:100%;overflow:hidden;">XPT_PAYER_TRX_PLACEHOLDER</div>
        <div id="xpt-payer-nrx" style="width:100%;overflow:hidden;height:0;visibility:hidden;">XPT_PAYER_NRX_PLACEHOLDER</div>
        <div class="legend" style="font-size:10px;margin-top:8px;">
          <div class="legend-item"><div class="legend-dot" style="background:#3b82f6"></div>Commercial</div>
          <div class="legend-item"><div class="legend-dot" style="background:#16a34a"></div>Medicare</div>
          <div class="legend-item"><div class="legend-dot" style="background:#f59e0b"></div>Medicaid</div>
          <div class="legend-item"><div class="legend-dot" style="background:#9ca3af"></div>Other</div>
        </div>
      </div>

      <div class="card">
        <div class="card-title" data-xp-metric="{XM} Share by Channel">TRx Share by Channel</div>
        <div class="card-subtitle">Retail / Mail-Order / LTC · weekly</div>
        <div id="xpt-ch-trx" style="width:100%;overflow:hidden;">XPT_CH_TRX_PLACEHOLDER</div>
        <div id="xpt-ch-nrx" style="width:100%;overflow:hidden;height:0;visibility:hidden;">XPT_CH_NRX_PLACEHOLDER</div>
        <div class="legend" style="font-size:10px;margin-top:8px;">
          <div class="legend-item"><div class="legend-dot" style="background:#0891b2"></div>Retail</div>
          <div class="legend-item"><div class="legend-dot" style="background:#7c3aed"></div>Mail-Order</div>
          <div class="legend-item"><div class="legend-dot" style="background:#9ca3af"></div>LTC</div>
        </div>
      </div>
      </div>

    <div class="row">
      <div class="card">
      <div class="card-title">Payer Performance</div>
      <div class="card-subtitle">Share % · Latest Wk / WoW % / MTD / vs. STLY</div>
      <table>
        <thead>
          <tr><th>CUT</th><th style="text-align:right" data-xp-metric="WoW Mkt Share ({XM})">WoW Mkt Share (TRx)</th><th style="text-align:right" data-xp-metric="WoW Growth ({XM})">WoW Growth (TRx)</th><th style="text-align:right" data-xp-metric="MTD Mkt Share ({XM})">MTD Mkt Share (TRx)</th><th style="text-align:right" data-xp-metric="MTD Growth vs STLY ({XM})">MTD Growth vs STLY (TRx)</th></tr>
        </thead>
        <tbody class="xpt-trx-body">
XPT_PAYER_TRX_TABLE_ROWS
        </tbody>
        <tbody class="xpt-nrx-body" style="display:none;">
XPT_PAYER_NRX_TABLE_ROWS
        </tbody>
      </table>
    </div>
      <div class="card">
      <div class="card-title">Channel Performance</div>
      <div class="card-subtitle">Share % · Latest Wk / WoW % / MTD / vs. STLY</div>
      <table>
        <thead>
          <tr><th>CUT</th><th style="text-align:right" data-xp-metric="WoW Mkt Share ({XM})">WoW Mkt Share (TRx)</th><th style="text-align:right" data-xp-metric="WoW Growth ({XM})">WoW Growth (TRx)</th><th style="text-align:right" data-xp-metric="MTD Mkt Share ({XM})">MTD Mkt Share (TRx)</th><th style="text-align:right" data-xp-metric="MTD Growth vs STLY ({XM})">MTD Growth vs STLY (TRx)</th></tr>
        </thead>
        <tbody class="xpt-trx-body">
XPT_CHANNEL_TRX_TABLE_ROWS
        </tbody>
        <tbody class="xpt-nrx-body" style="display:none;">
XPT_CHANNEL_NRX_TABLE_ROWS
        </tbody>
      </table>
    </div>
    </div>
  </div>
</div>
<div class="nl-tab" id="nl-market-tab">
<div style="padding:8px 32px 24px;">
<div class="section-head"><h2>NPA Performance</h2><p>National prescription audit competitive view across Oral CGRP brands.</p></div>
    <div class="pill-group" id="npa-subtabs">
      <div class="pill active" onclick="switchNpaSubtab('retail')">NPA Overall</div>
      <div class="pill" onclick="switchNpaSubtab('acute-prev')">NPA Acute / Preventive Split</div>
    </div>

    <!-- NPA Retail Sub-tab -->
    <div id="npa-retail">
      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;">
        <div class="pill-group" style="margin-bottom:0;">
          <div class="pill pill-sm active" id="npa-trx-pill" onclick="toggleNpaMetric('trx')">TRx</div>
          <div class="pill pill-sm" id="npa-nbrx-pill" onclick="toggleNpaMetric('nbrx')">NBRx</div>
        </div>
        <div class="dropdown-wrap" style="position:relative;">
          <button class="icon-btn" style="font-size:11px;padding:5px 12px;border-radius:6px;background:#0000C9;color:#fff;border:none;cursor:pointer;display:flex;align-items:center;gap:5px;font-weight:700;" onclick="toggleDropdown('npaDownloadDD', event)">
            <svg viewBox="0 0 24 24" style="width:13px;height:13px;fill:none;stroke:currentColor;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
            <svg viewBox="0 0 24 24" style="width:10px;height:10px;fill:none;stroke:currentColor;stroke-width:2;"><path d="M6 9l6 6 6-6"/></svg>
          </button>
          <div class="dropdown" id="npaDownloadDD" style="min-width:220px;">
            <a class="dropdown-item npa-dl-item" id="npa-dl-brand" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_BRAND_TRX_B64" download="NPA_Brand_Competitive_TRx.xlsx" style="text-decoration:none;color:inherit;display:flex;justify-content:space-between;" data-npa-dl-label="Brand Competitive ({M})" data-trx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_BRAND_TRX_B64" data-nbrx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_BRAND_NBRX_B64" data-trx-fname="NPA_Brand_Competitive_TRx.xlsx" data-nbrx-fname="NPA_Brand_Competitive_NBRx.xlsx"><span>Brand Competitive (TRx)</span><svg viewBox="0 0 24 24" style="width:12px;height:12px;fill:none;stroke:#6b7280;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg></a>
            <a class="dropdown-item npa-dl-item" id="npa-dl-channel" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_CH_NURTEC_TRX_B64" download="NPA_Channel_Nurtec_TRx.xlsx" style="text-decoration:none;color:inherit;display:flex;justify-content:space-between;" data-npa-dl-label="Channel - {B} ({M})" data-nurtec-trx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_CH_NURTEC_TRX_B64" data-nurtec-nbrx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_CH_NURTEC_NBRX_B64" data-ubrelvy-trx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_CH_UBRELVY_TRX_B64" data-ubrelvy-nbrx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_CH_UBRELVY_NBRX_B64" data-qulipta-trx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_CH_QULIPTA_TRX_B64" data-qulipta-nbrx-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,NPA_CH_QULIPTA_NBRX_B64"><span>Channel - Nurtec (TRx)</span><svg viewBox="0 0 24 24" style="width:12px;height:12px;fill:none;stroke:#6b7280;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg></a>
          </div>
        </div>
      </div>
      <div class="card">
        <div class="card-header-row">
          <div>
            <div class="card-title" data-metric-toggle="Nurtec NPA — Brand Competitive View ({M})">Nurtec NPA — Brand Competitive View (TRx)</div>
            <div class="card-subtitle">National · IQVIA NPA · Actuals 2026 vs Actuals 2025 (Same Time Last Year)</div>
          </div>
          
        </div>
        <div class="chart-container">
          <svg class="chart" preserveAspectRatio="none" viewBox="0 0 800 250">
BRAND_CHART_DATA_PLACEHOLDER
          </svg>
        </div>
        
        
        
        <div class="footnote">Time period reference: Actuals 2026 w.e. 06/05/2026 · Same Time Last Year w.e. 06/06/2025</div>
      </div>

      <div class="card">
        <div class="card-header-row">
          <div>
            <div class="card-title" id="overall-ch-title" data-metric-toggle="Nurtec NPA — Channel Performance View ({M})">Nurtec NPA — Channel Performance View (TRx)</div>
            <div class="card-subtitle">National · IQVIA NPA · Retail / Mail-Order / LTC · Actuals 2026 vs Actuals 2025 (STLY)</div>
          </div>
          <div class="pill-group" style="margin-bottom:0;">
            <div class="pill pill-sm active" id="channel-nurtec" onclick="switchChannelBrand('nurtec')">Nurtec</div>
            <div class="pill pill-sm" id="channel-ubrelvy" onclick="switchChannelBrand('ubrelvy')">Ubrelvy</div>
            <div class="pill pill-sm" id="channel-qulipta" onclick="switchChannelBrand('qulipta')">Qulipta</div>
          </div>
        
        </div>
        CHANNEL_CHART_PLACEHOLDER
        
        
        <div class="footnote">Time period reference: Actuals 2026 w.e. 06/05/2026 · Same Time Last Year w.e. 06/06/2025</div>
      </div>

      <div class="brand-header brand-a"><span>NURTEC</span><span style="font-weight:400;margin-left:12px;" data-metric-toggle="{M} Volume · Growth (vs STLY) · oCGRP Market Share %">TRx Volume · Growth (vs STLY) · oCGRP Market Share %</span></div>
      <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
        <table>
          <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">QTD %</th><th style="text-align:right">YTD %</th><th style="text-align:right">WK MS</th><th style="text-align:right">QTD MS</th><th style="text-align:right">YTD MS</th></tr></thead>
          <tbody class="npa-table-trx">
NPA_OVERALL_ROWS_NURTEC_TRx
          </tbody>
          <tbody class="npa-table-nbrx" style="display:none;">
NPA_OVERALL_ROWS_NURTEC_NBRx
          </tbody>
        </table>
      </div>

      <div class="brand-header brand-b" style="margin-top:16px;"><span>UBRELVY</span><span style="font-weight:400;margin-left:12px;" data-metric-toggle="{M} Volume · Growth (vs STLY) · oCGRP Market Share %">TRx Volume · Growth (vs STLY) · oCGRP Market Share %</span></div>
      <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
        <table>
          <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">QTD %</th><th style="text-align:right">YTD %</th><th style="text-align:right">WK MS</th><th style="text-align:right">QTD MS</th><th style="text-align:right">YTD MS</th></tr></thead>
          <tbody class="npa-table-trx">
NPA_OVERALL_ROWS_UBRELVY_TRx
          </tbody>
          <tbody class="npa-table-nbrx" style="display:none;">
NPA_OVERALL_ROWS_UBRELVY_NBRx
          </tbody>
        </table>
      </div>

      <div class="brand-header brand-c" style="margin-top:16px;"><span>QULIPTA</span><span style="font-weight:400;margin-left:12px;" data-metric-toggle="{M} Volume · Growth (vs STLY) · oCGRP Market Share %">TRx Volume · Growth (vs STLY) · oCGRP Market Share %</span></div>
      <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
        <table>
          <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">QTD %</th><th style="text-align:right">YTD %</th><th style="text-align:right">WK MS</th><th style="text-align:right">QTD MS</th><th style="text-align:right">YTD MS</th></tr></thead>
          <tbody class="npa-table-trx">
NPA_OVERALL_ROWS_QULIPTA_TRx
          </tbody>
          <tbody class="npa-table-nbrx" style="display:none;">
NPA_OVERALL_ROWS_QULIPTA_NBRx
          </tbody>
        </table>
      </div>

      

      
    </div>

    <!-- NPA Acute / Preventive Split Sub-tab -->
    <div id="npa-acute-prev" style="display:none;">
      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;">
        <div class="pill-group" id="acute-prev-toggle" style="margin-bottom:0;">
          <div class="pill active" onclick="switchAcutePrev('acute')">Acute</div>
          <div class="pill" onclick="switchAcutePrev('preventive')">Preventive</div>
        </div>
        <div class="dropdown-wrap" style="position:relative;">
          <button class="icon-btn" style="font-size:11px;padding:5px 12px;border-radius:6px;background:#0000C9;color:#fff;border:none;cursor:pointer;display:flex;align-items:center;gap:5px;font-weight:700;" onclick="toggleDropdown('apDownloadDD', event)">
            <svg viewBox="0 0 24 24" style="width:13px;height:13px;fill:none;stroke:currentColor;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
            <svg viewBox="0 0 24 24" style="width:10px;height:10px;fill:none;stroke:currentColor;stroke-width:2;"><path d="M6 9l6 6 6-6"/></svg>
          </button>
          <div class="dropdown" id="apDownloadDD" style="min-width:220px;">
            <a class="dropdown-item" id="ap-dl-brand" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,AP_ACUTE_BRAND_B64" download="NPA_Acute_Brand.xlsx" style="text-decoration:none;color:inherit;display:flex;justify-content:space-between;" data-acute-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,AP_ACUTE_BRAND_B64" data-prev-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,AP_PREV_BRAND_B64" data-acute-fname="NPA_Acute_Brand.xlsx" data-prev-fname="NPA_Preventive_Brand.xlsx"><span>Acute - Brand Competitive</span><svg viewBox="0 0 24 24" style="width:12px;height:12px;fill:none;stroke:#6b7280;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg></a>
            <a class="dropdown-item" id="ap-dl-channel" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,AP_ACUTE_CH_NURTEC_B64" download="NPA_Acute_Channel_Nurtec.xlsx" style="text-decoration:none;color:inherit;display:flex;justify-content:space-between;" data-acute-nurtec-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,AP_ACUTE_CH_NURTEC_B64" data-acute-ubrelvy-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,AP_ACUTE_CH_UBRELVY_B64" data-prev-nurtec-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,AP_PREV_CH_NURTEC_B64" data-prev-qulipta-href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,AP_PREV_CH_QULIPTA_B64"><span>Acute - Channel Nurtec</span><svg viewBox="0 0 24 24" style="width:12px;height:12px;fill:none;stroke:#6b7280;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg></a>
          </div>
        </div>
      </div>

      <!-- ACUTE VIEW -->
      <div id="acute-view">
        <div class="card">
          <div class="card-title">Nurtec Acute — Brand Competitive View</div>
          <div class="card-subtitle">National · IQVIA NPA · Actuals 2026 vs Actuals 2025 (Same Time Last Year)</div>

          <!-- TRx (left) and NBRx (right) charts side by side -->
          ACUTE_BRAND_CHART_PLACEHOLDER
        </div>

        
        
        <div class="card">
          <div class="card-header-row">
            <div>
              <div class="card-title" id="acute-ch-title">Nurtec Acute — Channel Performance View</div>
              <div class="card-subtitle">National &middot; IQVIA NPA &middot; Retail / Mail-Order / LTC &middot; Actuals 2026 vs STLY</div>
            </div>
            <div class="pill-group" style="margin-bottom:0;">
              <div class="pill pill-sm active" id="acute-ch-nurtec" onclick="switchAcuteChannel('nurtec')">Nurtec</div>
              <div class="pill pill-sm" id="acute-ch-ubrelvy" onclick="switchAcuteChannel('ubrelvy')">Ubrelvy</div>
            </div>
          </div>
          ACUTE_CHANNEL_CHART_PLACEHOLDER
        </div>

        

        <!-- Brand Tables: TRx left, NBRx right — aligned with charts above -->
        <div class="split-section">
          <div class="split-col">
            <div class="brand-header brand-a"><span>NURTEC ACUTE</span><span style="font-weight:400;margin-left:8px;">TRx Volume · Growth · MS</span></div>
            <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
              <table>
                <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">WK MS</th></tr></thead>
                <tbody>
NPA_ACUTE_ROWS_NURTEC_TRx
                </tbody>
              </table>
            </div>

            <div class="brand-header brand-b" style="margin-top:12px;"><span>UBRELVY ACUTE</span><span style="font-weight:400;margin-left:8px;">TRx Volume · Growth · MS</span></div>
            <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
              <table>
                <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">WK MS</th></tr></thead>
                <tbody>
NPA_ACUTE_ROWS_UBRELVY_TRx
                </tbody>
              </table>
            </div>
          </div>

          <div class="split-col">
            <div class="brand-header brand-a"><span>NURTEC ACUTE</span><span style="font-weight:400;margin-left:8px;">NBRx Volume · Growth · MS</span></div>
            <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
              <table>
                <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">WK MS</th></tr></thead>
                <tbody>
NPA_ACUTE_ROWS_NURTEC_NBRx
                </tbody>
              </table>
            </div>

            <div class="brand-header brand-b" style="margin-top:12px;"><span>UBRELVY ACUTE</span><span style="font-weight:400;margin-left:8px;">NBRx Volume · Growth · MS</span></div>
            <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
              <table>
                <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">WK MS</th></tr></thead>
                <tbody>
NPA_ACUTE_ROWS_UBRELVY_NBRx
                </tbody>
              </table>
            </div>
          </div>
        </div>
      </div>

      <!-- PREVENTIVE VIEW -->
      <div id="preventive-view" style="display:none;">
        <div class="card">
          <div class="card-title">Nurtec Preventive — Brand Competitive View</div>
          <div class="card-subtitle">National · IQVIA NPA · Actuals 2026 vs Actuals 2025 (Same Time Last Year)</div>

          PREV_BRAND_CHART_PLACEHOLDER
        </div>

        
        <div class="card">
          <div class="card-header-row">
            <div>
              <div class="card-title" id="prev-ch-title">Nurtec Preventive — Channel Performance View</div>
              <div class="card-subtitle">National &middot; IQVIA NPA &middot; Retail / Mail-Order / LTC &middot; Actuals 2026 vs STLY</div>
            </div>
            <div class="pill-group" style="margin-bottom:0;">
              <div class="pill pill-sm active" id="prev-ch-nurtec" onclick="switchPrevChannel('nurtec')">Nurtec</div>
              <div class="pill pill-sm" id="prev-ch-qulipta" onclick="switchPrevChannel('qulipta')">Qulipta</div>
            </div>
          </div>
          PREV_CHANNEL_CHART_PLACEHOLDER
        </div>

        <!-- Preventive tables: TRx left, NBRx right -->
        <div class="split-section">
          <div class="split-col">
            <div class="brand-header brand-a"><span>NURTEC PREVENTIVE</span><span style="font-weight:400;margin-left:8px;">TRx Volume · Growth · MS</span></div>
            <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
              <table>
                <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">WK MS</th></tr></thead>
                <tbody>
NPA_PREV_ROWS_NURTEC_TRx
                </tbody>
              </table>
            </div>

            <div class="brand-header brand-c" style="margin-top:12px;"><span>QULIPTA PREVENTIVE</span><span style="font-weight:400;margin-left:8px;">TRx Volume · Growth · MS</span></div>
            <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
              <table>
                <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">WK MS</th></tr></thead>
                <tbody>
NPA_PREV_ROWS_QULIPTA_TRx
                </tbody>
              </table>
            </div>
          </div>

          <div class="split-col">
            <div class="brand-header brand-a"><span>NURTEC PREVENTIVE</span><span style="font-weight:400;margin-left:8px;">NBRx Volume · Growth · MS</span></div>
            <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
              <table>
                <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">WK MS</th></tr></thead>
                <tbody>
NPA_PREV_ROWS_NURTEC_NBRx
                </tbody>
              </table>
            </div>

            <div class="brand-header brand-c" style="margin-top:12px;"><span>QULIPTA PREVENTIVE</span><span style="font-weight:400;margin-left:8px;">NBRx Volume · Growth · MS</span></div>
            <div class="card" style="border-top-left-radius:0;border-top-right-radius:0;margin-top:-1px;">
              <table>
                <thead><tr><th></th><th style="text-align:right">LATEST WK</th><th style="text-align:right">QTD</th><th style="text-align:right">YTD</th><th style="text-align:right">WK %</th><th style="text-align:right">WK MS</th></tr></thead>
                <tbody>
NPA_PREV_ROWS_QULIPTA_NBRx
                </tbody>
              </table>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
</div>
<div class="nl-tab" id="nl-access-tab">
<div style="position:relative;">
<div style="position:absolute;top:0;left:0;right:0;bottom:0;background:rgba(255,255,255,0.58);z-index:50;display:flex;align-items:flex-start;justify-content:center;padding-top:60px;border-radius:12px;">
  <div style="background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:28px 40px;box-shadow:0 4px 20px rgba(0,0,0,0.08);text-align:center;max-width:420px;">
    <div style="font-size:14px;font-weight:600;color:#1a2332;margin-bottom:6px;">Coming Soon</div>
    <div style="font-size:12px;color:#64748b;line-height:1.6;">This section is not yet live and views are just for illustrative purpose.</div>
  </div>
</div>
<div style="padding:8px 32px 24px;">
<div class="section-head"><h2>Access Summary</h2><p>Payer access changes, formulary tracking, and covered lives analysis.</p></div>
    <div class="card">
      <div class="card-header-row">
        <div>
          <div class="card-title">Rolled-up NRx Share by Access-Change Cohort</div>
          <div class="card-subtitle">Calculated share across all plans grouped by access change since Jan 2025. Plans contribute to their cohort only after their change effective date — pre-change months roll into Control.</div>
        </div>
                <div style="position:relative;display:inline-block;">
          <div style="width:28px;height:28px;border-radius:50%;background:#ffffff;color:#0000C9;font-size:16px;font-weight:800;display:flex;align-items:center;justify-content:center;cursor:pointer;border:1px solid #c7d2fe;" onmouseover="this.nextElementSibling.style.display='block'" onmouseout="this.nextElementSibling.style.display='none'">i</div>
          <div style="display:none;position:absolute;right:0;top:32px;background:#1a2332;color:#fff;padding:8px 12px;border-radius:6px;font-size:11px;line-height:1.5;width:auto;white-space:nowrap;z-index:10;box-shadow:0 4px 12px rgba(0,0,0,0.15);">Data Considered from 2025 onwards only</div>
        </div>
      </div>

    <div class="row">
        <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
          <div class="label">ACCESS GAIN PLANS</div>
          <div class="value" style="color:#16a34a;">7</div>
          <div class="sub">gain events since Jan 2025</div>
        </div>
        <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
          <div class="label">ACCESS LOSS / RESTRICTION</div>
          <div class="value" style="color:#dc2626;">4</div>
          <div class="sub">loss / restriction events</div>
        </div>
        <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
          <div class="label">CONTROL GROUP</div>
          <div class="value" style="color:#0000C9;">38</div>
          <div class="sub">no change (incl. pre-change months)</div>
        </div>
      </div>

      <div class="chart-container">
        <svg class="chart" preserveAspectRatio="none" viewBox="0 0 800 250">
          <line x1="60" y1="20" x2="60" y2="220" stroke="#f3f4f6" stroke-width="1"/>
          <line x1="60" y1="220" x2="780" y2="220" stroke="#e5e7eb" stroke-width="1"/>
          <line x1="60" y1="170" x2="780" y2="170" stroke="#f3f4f6" stroke-width="1" stroke-dasharray="4"/>
          <line x1="60" y1="120" x2="780" y2="120" stroke="#f3f4f6" stroke-width="1" stroke-dasharray="4"/>
          <line x1="60" y1="70" x2="780" y2="70" stroke="#f3f4f6" stroke-width="1" stroke-dasharray="4"/>
          <text x="50" y="224" text-anchor="end" font-size="10" fill="#9ca3af">28%</text>
          <text x="50" y="174" text-anchor="end" font-size="10" fill="#9ca3af">31%</text>
          <text x="50" y="124" text-anchor="end" font-size="10" fill="#9ca3af">34%</text>
          <text x="50" y="74" text-anchor="end" font-size="10" fill="#9ca3af">37%</text>
          <text x="100" y="240" text-anchor="middle" font-size="9" fill="#9ca3af">Jan'25</text>
          <text x="180" y="240" text-anchor="middle" font-size="9" fill="#9ca3af">Mar'25</text>
          <text x="260" y="240" text-anchor="middle" font-size="9" fill="#9ca3af">May'25</text>
          <text x="340" y="240" text-anchor="middle" font-size="9" fill="#9ca3af">Jul'25</text>
          <text x="420" y="240" text-anchor="middle" font-size="9" fill="#9ca3af">Sep'25</text>
          <text x="500" y="240" text-anchor="middle" font-size="9" fill="#9ca3af">Nov'25</text>
          <text x="580" y="240" text-anchor="middle" font-size="9" fill="#9ca3af">Jan'26</text>
          <text x="660" y="240" text-anchor="middle" font-size="9" fill="#9ca3af">Mar'26</text>
          <text x="740" y="240" text-anchor="middle" font-size="9" fill="#9ca3af">May'26</text>
          <line x1="260" y1="20" x2="260" y2="220" stroke="#9ca3af" stroke-width="1" stroke-dasharray="4"/>
          <text x="265" y="35" font-size="9" fill="#6b7280">Gain cohort starts</text>
          <polyline fill="none" stroke="#16a34a" stroke-width="2" points="80,180 120,178 160,175 200,172 240,170 280,165 320,155 360,148 400,140 440,130 480,120 520,108 560,95 600,85 640,72 680,60 720,50 760,42"/>
          <polyline fill="none" stroke="#dc2626" stroke-width="2" points="80,145 120,148 160,150 200,152 240,155 280,160 320,168 360,172 400,178 440,182 480,188 520,192 560,195 600,198 640,200 680,202 720,205 760,208"/>
          <polyline fill="none" stroke="#9ca3af" stroke-width="2" stroke-dasharray="6" points="80,130 120,132 160,130 200,128 240,130 280,128 320,127 360,128 400,126 440,125 480,124 520,123 560,122 600,120 640,119 680,118 720,117 760,115"/>
        </svg>
      </div>
      <div class="axis-info"><span>Y-Axis: NRx Share %</span><span>X-Axis: Time Period (Month)</span></div>
        <div class="legend">
        <div class="legend-item"><div class="legend-dot" style="background:#16a34a"></div>Plans w/ access gain (post-change)</div>
        <div class="legend-item"><div class="legend-dot" style="background:#dc2626"></div>Plans w/ access loss (post-change)</div>
        <div class="legend-item"><div class="legend-dot" style="background:#9ca3af"></div>Control (no change / pre-change)</div>
      </div>
    </div>

    <div class="card">
      <div class="card-title">Access Change Details</div>
      <div class="card-subtitle">Includes current share vs. share at time of change (delta is the story)</div>
      <div style="max-height:220px;overflow-y:auto;">
      <table style="table-layout:fixed;">
        <colgroup><col style="width:28%"><col style="width:9%"><col style="width:9%"><col style="width:8%"><col style="width:7%"><col style="width:11%"><col style="width:11%"><col style="width:7%"><col style="width:6%"><col style="width:6%"></colgroup>
        <thead style="position:sticky;top:0;z-index:1;background:#fff;">
          <tr><th>PLAN</th><th>CHANGE</th><th>EFF. DATE</th><th>PAYER</th><th style="text-align:right">LIVES</th><th style="text-align:right">SHARE @ CHANGE</th><th style="text-align:right">CURRENT SHARE</th><th style="text-align:right">Δ SHARE</th><th style="text-align:right">EST. NRX %</th><th style="text-align:right">EST. TRX %</th></tr>
        </thead>
        <tbody>
          <tr><td>Connecticut (TOP$)</td><td><span class="badge badge-gain">Gain</span></td><td>06/18/26</td><td>Commercial</td><td style="text-align:right">1.2M</td><td style="text-align:right">29.3%</td><td style="text-align:right">33.1%</td><td style="text-align:right"><span class="delta-pos">+3.8%</span></td><td style="text-align:right"><span class="delta-pos">+0.5%</span></td><td style="text-align:right"><span class="delta-pos">+0.3%</span></td></tr>
          <tr><td>Delaware (SSDC)</td><td><span class="badge badge-loss">Loss</span></td><td>06/12/26</td><td>Medicaid</td><td style="text-align:right">680K</td><td style="text-align:right">31.2%</td><td style="text-align:right">28.7%</td><td style="text-align:right"><span class="delta-neg">-2.5%</span></td><td style="text-align:right"><span class="delta-neg">-0.4%</span></td><td style="text-align:right"><span class="delta-neg">-0.3%</span></td></tr>
          <tr><td>Idaho (TOP$)</td><td><span class="badge badge-gain">Gain</span></td><td>06/05/26</td><td>Commercial</td><td style="text-align:right">920K</td><td style="text-align:right">25.8%</td><td style="text-align:right">30.2%</td><td style="text-align:right"><span class="delta-pos">+4.4%</span></td><td style="text-align:right"><span class="delta-pos">+0.7%</span></td><td style="text-align:right"><span class="delta-pos">+0.5%</span></td></tr>
          <tr><td>Iowa (SSDC)</td><td><span class="badge badge-restriction">Restriction</span></td><td>05/29/26</td><td>Medicaid</td><td style="text-align:right">1.1M</td><td style="text-align:right">33.6%</td><td style="text-align:right">31.9%</td><td style="text-align:right"><span class="delta-neg">-1.7%</span></td><td style="text-align:right"><span class="delta-neg">-0.2%</span></td><td style="text-align:right"><span class="delta-neg">-0.1%</span></td></tr>
          <tr><td>Kentucky (SSDC)</td><td><span class="badge badge-gain">Gain</span></td><td>05/22/26</td><td>Medicaid</td><td style="text-align:right">1.5M</td><td style="text-align:right">27.4%</td><td style="text-align:right">31.8%</td><td style="text-align:right"><span class="delta-pos">+4.4%</span></td><td style="text-align:right"><span class="delta-pos">+0.8%</span></td><td style="text-align:right"><span class="delta-pos">+0.6%</span></td></tr>
          <tr><td>Louisiana (TOP$)</td><td><span class="badge badge-loss">Loss</span></td><td>05/15/26</td><td>Commercial</td><td style="text-align:right">1.8M</td><td style="text-align:right">35.1%</td><td style="text-align:right">32.4%</td><td style="text-align:right"><span class="delta-neg">-2.7%</span></td><td style="text-align:right"><span class="delta-neg">-0.5%</span></td><td style="text-align:right"><span class="delta-neg">-0.4%</span></td></tr>
          <tr><td>Mississippi (SSDC)</td><td><span class="badge badge-restriction">Restriction</span></td><td>05/08/26</td><td>Medicaid</td><td style="text-align:right">740K</td><td style="text-align:right">28.9%</td><td style="text-align:right">26.5%</td><td style="text-align:right"><span class="delta-neg">-2.4%</span></td><td style="text-align:right"><span class="delta-neg">-0.3%</span></td><td style="text-align:right"><span class="delta-neg">-0.2%</span></td></tr>
          <tr><td>Montana (NMPI)</td><td><span class="badge badge-gain">Gain</span></td><td>05/01/26</td><td>Commercial</td><td style="text-align:right">560K</td><td style="text-align:right">22.1%</td><td style="text-align:right">27.6%</td><td style="text-align:right"><span class="delta-pos">+5.5%</span></td><td style="text-align:right"><span class="delta-pos">+0.4%</span></td><td style="text-align:right"><span class="delta-pos">+0.3%</span></td></tr>
          <tr><td>New York (NMPI)</td><td><span class="badge badge-gain">Gain</span></td><td>04/24/26</td><td>Commercial</td><td style="text-align:right">4.2M</td><td style="text-align:right">30.5%</td><td style="text-align:right">34.2%</td><td style="text-align:right"><span class="delta-pos">+3.7%</span></td><td style="text-align:right"><span class="delta-pos">+1.2%</span></td><td style="text-align:right"><span class="delta-pos">+0.9%</span></td></tr>
          <tr><td>Ohio (SSDC)</td><td><span class="badge badge-loss">Loss</span></td><td>04/17/26</td><td>Medicaid</td><td style="text-align:right">2.3M</td><td style="text-align:right">32.8%</td><td style="text-align:right">29.6%</td><td style="text-align:right"><span class="delta-neg">-3.2%</span></td><td style="text-align:right"><span class="delta-neg">-0.6%</span></td><td style="text-align:right"><span class="delta-neg">-0.5%</span></td></tr>
          <tr><td>Pennsylvania (SSDC)</td><td><span class="badge badge-restriction">Restriction</span></td><td>04/10/26</td><td>Commercial</td><td style="text-align:right">3.1M</td><td style="text-align:right">34.7%</td><td style="text-align:right">33.1%</td><td style="text-align:right"><span class="delta-neg">-1.6%</span></td><td style="text-align:right"><span class="delta-neg">-0.3%</span></td><td style="text-align:right"><span class="delta-neg">-0.2%</span></td></tr>
          <tr><td>Texas (Stand-Alone OptumRx)</td><td><span class="badge badge-gain">Gain</span></td><td>04/03/26</td><td>Commercial</td><td style="text-align:right">5.4M</td><td style="text-align:right">26.3%</td><td style="text-align:right">31.5%</td><td style="text-align:right"><span class="delta-pos">+5.2%</span></td><td style="text-align:right"><span class="delta-pos">+1.5%</span></td><td style="text-align:right"><span class="delta-pos">+1.1%</span></td></tr>
          <tr><td>Wyoming (SSDC)</td><td><span class="badge badge-loss">Loss</span></td><td>03/27/26</td><td>Medicaid</td><td style="text-align:right">320K</td><td style="text-align:right">29.4%</td><td style="text-align:right">26.1%</td><td style="text-align:right"><span class="delta-neg">-3.3%</span></td><td style="text-align:right"><span class="delta-neg">-0.1%</span></td><td style="text-align:right"><span class="delta-neg">-0.1%</span></td></tr>
        </tbody>
      </table>
      </div>
    </div>

    <div class="card">
      <div class="card-header-row">
        <div>
          <div class="card-title">Access Change Details — Top Plans by Payer</div>
          <div class="card-subtitle">Top 5 plans by covered lives within the selected payer type</div>
        </div>
        <div class="filter-group" style="margin-bottom:0;">
          <span style="font-size:11px;font-weight:600;">Payer Type</span>
          <select class="filter-select" id="payer-filter-select" onchange="filterPayerTable(this.value)">
            <option value="commercial">Commercial</option>
            <option value="medicaid">Medicaid</option>
            <option value="medicare">Medicare</option>
            <option value="other">Other</option>
          </select>
        </div>
      </div>
      <table style="table-layout:fixed;">
        <colgroup><col style="width:36%"><col style="width:9%"><col style="width:9%"><col style="width:7%"><col style="width:11%"><col style="width:11%"><col style="width:7%"><col style="width:6%"><col style="width:6%"></colgroup>
        <thead>
          <tr><th>PLAN</th><th>CHANGE</th><th>EFF. DATE</th><th style="text-align:right">LIVES</th><th style="text-align:right">SHARE @ CHANGE</th><th style="text-align:right">CURRENT SHARE</th><th style="text-align:right">Δ SHARE</th><th style="text-align:right">EST. NRX %</th><th style="text-align:right">EST. TRX %</th></tr>
        </thead>
        <tbody id="payer-table-body">
          <tr data-payer="commercial"><td>Aetna National</td><td><span class="badge badge-gain">Gain</span></td><td>06/10/26</td><td style="text-align:right">5.8M</td><td style="text-align:right">28.4%</td><td style="text-align:right">33.2%</td><td style="text-align:right"><span class="delta-pos">+4.8%</span></td><td style="text-align:right"><span class="delta-pos">+1.4%</span></td><td style="text-align:right"><span class="delta-pos">+1.1%</span></td></tr>
          <tr data-payer="commercial"><td>Cigna Commercial</td><td><span class="badge badge-gain">Gain</span></td><td>05/28/26</td><td style="text-align:right">4.3M</td><td style="text-align:right">30.1%</td><td style="text-align:right">34.6%</td><td style="text-align:right"><span class="delta-pos">+4.5%</span></td><td style="text-align:right"><span class="delta-pos">+1.1%</span></td><td style="text-align:right"><span class="delta-pos">+0.8%</span></td></tr>
          <tr data-payer="commercial"><td>UnitedHealth Comm (National)</td><td><span class="badge badge-restriction">Restriction</span></td><td>05/15/26</td><td style="text-align:right">8.2M</td><td style="text-align:right">36.7%</td><td style="text-align:right">35.1%</td><td style="text-align:right"><span class="delta-neg">-1.6%</span></td><td style="text-align:right"><span class="delta-neg">-0.3%</span></td><td style="text-align:right"><span class="delta-neg">-0.2%</span></td></tr>
          <tr data-payer="commercial"><td>Anthem BCBS</td><td><span class="badge badge-gain">Gain</span></td><td>04/30/26</td><td style="text-align:right">6.1M</td><td style="text-align:right">27.9%</td><td style="text-align:right">32.4%</td><td style="text-align:right"><span class="delta-pos">+4.5%</span></td><td style="text-align:right"><span class="delta-pos">+1.3%</span></td><td style="text-align:right"><span class="delta-pos">+1.0%</span></td></tr>
          <tr data-payer="commercial"><td>Humana Commercial</td><td><span class="badge badge-loss">Loss</span></td><td>04/18/26</td><td style="text-align:right">3.7M</td><td style="text-align:right">33.5%</td><td style="text-align:right">30.8%</td><td style="text-align:right"><span class="delta-neg">-2.7%</span></td><td style="text-align:right"><span class="delta-neg">-0.5%</span></td><td style="text-align:right"><span class="delta-neg">-0.4%</span></td></tr>
          <tr data-payer="medicaid"><td>California Medicaid</td><td><span class="badge badge-gain">Gain</span></td><td>06/05/26</td><td style="text-align:right">4.9M</td><td style="text-align:right">22.3%</td><td style="text-align:right">27.8%</td><td style="text-align:right"><span class="delta-pos">+5.5%</span></td><td style="text-align:right"><span class="delta-pos">+0.9%</span></td><td style="text-align:right"><span class="delta-pos">+0.7%</span></td></tr>
          <tr data-payer="medicaid"><td>Texas Medicaid</td><td><span class="badge badge-loss">Loss</span></td><td>05/22/26</td><td style="text-align:right">3.6M</td><td style="text-align:right">29.1%</td><td style="text-align:right">26.4%</td><td style="text-align:right"><span class="delta-neg">-2.7%</span></td><td style="text-align:right"><span class="delta-neg">-0.6%</span></td><td style="text-align:right"><span class="delta-neg">-0.4%</span></td></tr>
          <tr data-payer="medicaid"><td>New York Medicaid</td><td><span class="badge badge-restriction">Restriction</span></td><td>05/10/26</td><td style="text-align:right">3.2M</td><td style="text-align:right">31.6%</td><td style="text-align:right">29.8%</td><td style="text-align:right"><span class="delta-neg">-1.8%</span></td><td style="text-align:right"><span class="delta-neg">-0.3%</span></td><td style="text-align:right"><span class="delta-neg">-0.2%</span></td></tr>
          <tr data-payer="medicaid"><td>Florida Medicaid</td><td><span class="badge badge-gain">Gain</span></td><td>04/28/26</td><td style="text-align:right">2.8M</td><td style="text-align:right">24.7%</td><td style="text-align:right">29.3%</td><td style="text-align:right"><span class="delta-pos">+4.6%</span></td><td style="text-align:right"><span class="delta-pos">+0.7%</span></td><td style="text-align:right"><span class="delta-pos">+0.5%</span></td></tr>
          <tr data-payer="medicaid"><td>Illinois Medicaid</td><td><span class="badge badge-loss">Loss</span></td><td>04/15/26</td><td style="text-align:right">2.1M</td><td style="text-align:right">30.4%</td><td style="text-align:right">27.2%</td><td style="text-align:right"><span class="delta-neg">-3.2%</span></td><td style="text-align:right"><span class="delta-neg">-0.4%</span></td><td style="text-align:right"><span class="delta-neg">-0.3%</span></td></tr>
          <tr data-payer="medicare"><td>UnitedHealth Medicare</td><td><span class="badge badge-gain">Gain</span></td><td>06/01/26</td><td style="text-align:right">7.4M</td><td style="text-align:right">25.6%</td><td style="text-align:right">30.1%</td><td style="text-align:right"><span class="delta-pos">+4.5%</span></td><td style="text-align:right"><span class="delta-pos">+1.6%</span></td><td style="text-align:right"><span class="delta-pos">+1.2%</span></td></tr>
          <tr data-payer="medicare"><td>Humana Medicare</td><td><span class="badge badge-restriction">Restriction</span></td><td>05/18/26</td><td style="text-align:right">5.1M</td><td style="text-align:right">32.3%</td><td style="text-align:right">30.7%</td><td style="text-align:right"><span class="delta-neg">-1.6%</span></td><td style="text-align:right"><span class="delta-neg">-0.4%</span></td><td style="text-align:right"><span class="delta-neg">-0.3%</span></td></tr>
          <tr data-payer="medicare"><td>Aetna Medicare Advantage</td><td><span class="badge badge-gain">Gain</span></td><td>05/05/26</td><td style="text-align:right">3.9M</td><td style="text-align:right">26.8%</td><td style="text-align:right">31.4%</td><td style="text-align:right"><span class="delta-pos">+4.6%</span></td><td style="text-align:right"><span class="delta-pos">+0.9%</span></td><td style="text-align:right"><span class="delta-pos">+0.7%</span></td></tr>
          <tr data-payer="medicare"><td>BCBS Medicare</td><td><span class="badge badge-loss">Loss</span></td><td>04/22/26</td><td style="text-align:right">2.6M</td><td style="text-align:right">34.2%</td><td style="text-align:right">31.5%</td><td style="text-align:right"><span class="delta-neg">-2.7%</span></td><td style="text-align:right"><span class="delta-neg">-0.4%</span></td><td style="text-align:right"><span class="delta-neg">-0.3%</span></td></tr>
          <tr data-payer="medicare"><td>Cigna Medicare Part D</td><td><span class="badge badge-gain">Gain</span></td><td>04/08/26</td><td style="text-align:right">2.2M</td><td style="text-align:right">23.9%</td><td style="text-align:right">28.7%</td><td style="text-align:right"><span class="delta-pos">+4.8%</span></td><td style="text-align:right"><span class="delta-pos">+0.5%</span></td><td style="text-align:right"><span class="delta-pos">+0.4%</span></td></tr>
          <tr data-payer="other"><td>Kaiser Permanente</td><td><span class="badge badge-gain">Gain</span></td><td>06/12/26</td><td style="text-align:right">4.1M</td><td style="text-align:right">24.5%</td><td style="text-align:right">29.8%</td><td style="text-align:right"><span class="delta-pos">+5.3%</span></td><td style="text-align:right"><span class="delta-pos">+1.0%</span></td><td style="text-align:right"><span class="delta-pos">+0.8%</span></td></tr>
          <tr data-payer="other"><td>Tricare (DoD)</td><td><span class="badge badge-restriction">Restriction</span></td><td>05/20/26</td><td style="text-align:right">2.9M</td><td style="text-align:right">31.8%</td><td style="text-align:right">29.4%</td><td style="text-align:right"><span class="delta-neg">-2.4%</span></td><td style="text-align:right"><span class="delta-neg">-0.3%</span></td><td style="text-align:right"><span class="delta-neg">-0.2%</span></td></tr>
          <tr data-payer="other"><td>VA Health System</td><td><span class="badge badge-gain">Gain</span></td><td>05/06/26</td><td style="text-align:right">3.5M</td><td style="text-align:right">19.7%</td><td style="text-align:right">25.3%</td><td style="text-align:right"><span class="delta-pos">+5.6%</span></td><td style="text-align:right"><span class="delta-pos">+0.8%</span></td><td style="text-align:right"><span class="delta-pos">+0.6%</span></td></tr>
          <tr data-payer="other"><td>Workers Comp National</td><td><span class="badge badge-loss">Loss</span></td><td>04/25/26</td><td style="text-align:right">1.7M</td><td style="text-align:right">28.3%</td><td style="text-align:right">25.1%</td><td style="text-align:right"><span class="delta-neg">-3.2%</span></td><td style="text-align:right"><span class="delta-neg">-0.2%</span></td><td style="text-align:right"><span class="delta-neg">-0.2%</span></td></tr>
          <tr data-payer="other"><td>Health Exchange (ACA)</td><td><span class="badge badge-gain">Gain</span></td><td>04/11/26</td><td style="text-align:right">2.4M</td><td style="text-align:right">22.9%</td><td style="text-align:right">27.4%</td><td style="text-align:right"><span class="delta-pos">+4.5%</span></td><td style="text-align:right"><span class="delta-pos">+0.5%</span></td><td style="text-align:right"><span class="delta-pos">+0.4%</span></td></tr>
        </tbody>
      </table>
    </div>

    <div class="card">
      <div class="card-header-row">
        <div>
          <div class="card-title">Lives Covered by Access Status</div>
          <div class="card-subtitle">Longitudinal · cumulative covered lives (millions) by status, filterable by payer type</div>
        </div>
        <div class="filter-group">
          <span style="font-size:11px;">Payer</span><select class="filter-select"><option>All</option></select>
        </div>
      </div>
      <div class="chart-container">
        <svg class="chart" preserveAspectRatio="none" viewBox="0 0 800 250">
          <line x1="60" y1="220" x2="780" y2="220" stroke="#1a2332" stroke-width="1"/>
          <text x="50" y="224" text-anchor="end" font-size="10" fill="#9ca3af">0M</text>
          <text x="50" y="174" text-anchor="end" font-size="10" fill="#9ca3af">35M</text>
          <text x="50" y="124" text-anchor="end" font-size="10" fill="#9ca3af">70M</text>
          <text x="50" y="74" text-anchor="end" font-size="10" fill="#9ca3af">105M</text>
          <text x="50" y="24" text-anchor="end" font-size="10" fill="#9ca3af">140M</text>
          <polyline fill="none" stroke="#16a34a" stroke-width="2.5" points="80,205 120,195 200,185 300,175 400,165 500,155 600,148 700,142 760,140"/>
          <polyline fill="none" stroke="#dc2626" stroke-width="2.5" points="80,215 120,213 200,212 300,213 400,214 500,215 600,216 700,217 760,217"/>
          <polyline fill="none" stroke="#9ca3af" stroke-width="2.5" points="80,180 120,140 200,100 300,70 400,55 500,45 600,40 700,36 760,34"/>
          <text x="80" y="238" text-anchor="middle" font-size="9" fill="#9ca3af">Jan'25</text>
          <text x="165" y="238" text-anchor="middle" font-size="9" fill="#9ca3af">Mar'25</text>
          <text x="250" y="238" text-anchor="middle" font-size="9" fill="#9ca3af">May'25</text>
          <text x="335" y="238" text-anchor="middle" font-size="9" fill="#9ca3af">Jul'25</text>
          <text x="420" y="238" text-anchor="middle" font-size="9" fill="#9ca3af">Sep'25</text>
          <text x="505" y="238" text-anchor="middle" font-size="9" fill="#9ca3af">Nov'25</text>
          <text x="590" y="238" text-anchor="middle" font-size="9" fill="#9ca3af">Jan'26</text>
          <text x="675" y="238" text-anchor="middle" font-size="9" fill="#9ca3af">Mar'26</text>
          <text x="760" y="238" text-anchor="middle" font-size="9" fill="#9ca3af">May'26</text>
        </svg>
      </div>
      <div class="axis-info"><span>Y-Axis: Covered Lives (M)</span><span>X-Axis: Time Period (Month)</span></div>
        <div class="legend">
        <div class="legend-item"><div class="legend-dot" style="background:#16a34a"></div>Gain</div>
        <div class="legend-item"><div class="legend-dot" style="background:#dc2626"></div>Loss</div>
        <div class="legend-item"><div class="legend-dot" style="background:#9ca3af"></div>Control</div>
      </div>
    </div>
  </div>
</div>
</div>
<div class="nl-tab" id="nl-financial-tab">
<div style="padding:8px 32px 24px;">
<div class="section-head"><h2>Financial Tracker</h2><p>Weekly gross and net sales performance vs. budget and plan attainment.</p></div>
    <div class="pill-group" id="financial-toggle" style="margin-bottom:16px;">
      <div class="pill active" onclick="switchFinancial('gross')">Gross Sales</div>
      <div class="pill" onclick="switchFinancial('net')">Net Sales</div>
    </div>
    <div style="display:flex;justify-content:flex-end;margin-bottom:6px;margin-top:-10px;">
      <div style="position:relative;display:inline-block;">
        <div style="width:28px;height:28px;border-radius:50%;background:#ffffff;color:#0000C9;font-size:16px;font-weight:800;display:flex;align-items:center;justify-content:center;cursor:pointer;border:1px solid #c7d2fe;" onmouseover="this.nextElementSibling.style.display='block'" onmouseout="this.nextElementSibling.style.display='none'">i</div>
        <div style="display:none;position:absolute;right:0;top:32px;background:#1a2332;color:#fff;padding:8px 12px;border-radius:6px;font-size:11px;line-height:1.5;width:auto;white-space:nowrap;z-index:10;box-shadow:0 4px 12px rgba(0,0,0,0.15);"><strong>Reporting Cadence:</strong><br>Gross Sales = Weekly &nbsp;|&nbsp; Net Sales = Monthly</div>
      </div>
    </div>

    <div id="gross-view">
    <div class="row">
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div class="label">FIN_GROSS_WK_LABEL</div>
        <div class="value" style="color:#0000C9;">FIN_GROSS_WK_VAL</div>
        <div class="sub">FIN_GROSS_WK_PY&nbsp;&nbsp;vs Prior Year</div>
      </div>
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div class="label">MTD ATTAINMENT VS OP'26</div>
        <div class="value" style="color:#0000C9;">FIN_GROSS_MTD_ATT</div>
        <div class="sub">FIN_GROSS_MTD_DELTA</div>
      </div>
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div class="label">YTD ATTAINMENT VS OP'26</div>
        <div class="value" style="color:#0000C9;">FIN_GROSS_YTD_ATT</div>
        <div class="sub">FIN_GROSS_YTD_DELTA</div>
      </div>
    </div>

    <div style="display:flex;justify-content:flex-end;margin-bottom:8px;margin-top:8px;">
      <div class="dropdown-wrap" style="position:relative;">
        <button class="icon-btn" style="font-size:11px;padding:5px 12px;border-radius:6px;background:#0000C9;color:#fff;border:none;cursor:pointer;display:flex;align-items:center;gap:5px;font-weight:700;" onclick="toggleDropdown('finDownloadDD', event)">
          <svg viewBox="0 0 24 24" style="width:13px;height:13px;fill:none;stroke:currentColor;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
          <svg viewBox="0 0 24 24" style="width:10px;height:10px;fill:none;stroke:currentColor;stroke-width:2;"><path d="M6 9l6 6 6-6"/></svg>
        </button>
        <div class="dropdown" id="finDownloadDD" style="min-width:180px;">
          <a class="dropdown-item" id="fin-dl-gross" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,FIN_GROSS_DL_B64" download="Finance_Gross_Sales.xlsx" style="text-decoration:none;color:inherit;display:flex;justify-content:space-between;"><span>Gross Sales</span><svg viewBox="0 0 24 24" style="width:12px;height:12px;fill:none;stroke:#6b7280;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg></a>
          <a class="dropdown-item" id="fin-dl-net" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,FIN_NET_DL_B64" download="Finance_Net_Sales.xlsx" style="text-decoration:none;color:inherit;display:flex;justify-content:space-between;"><span>Net Sales</span><svg viewBox="0 0 24 24" style="width:12px;height:12px;fill:none;stroke:#6b7280;stroke-width:2;"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg></a>
        </div>
      </div>
    </div>

        <div class="card">
      <div class="card-title">Weekly Gross Revenue vs Budget (2026)</div>
      <div class="card-subtitle">Actuals · STLY · Budget OP'26 · $M</div>
      FIN_GROSS_CHART_PLACEHOLDER
      <div class="legend" style="font-size:10px;margin-top:8px;">
        <div class="legend-item"><div class="legend-dot" style="background:#7C6CFC;border-radius:2px;width:12px;"></div>Actual</div>
        <div class="legend-item"><div class="legend-dot" style="background:#0000C9"></div>Actual (STLY)</div>
        <div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#9ca3af" stroke-width="2" stroke-dasharray="3"/></svg>Budget</div>
      </div>
    </div>

    </div>

    <div id="net-view" style="display:none;">
    <div class="row">
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div class="label">FIN_NET_MONTH_LABEL</div>
        <div class="value" style="color:#0000C9;">FIN_NET_MONTH_VAL</div>
        <div class="sub">FIN_NET_MONTH_PY&nbsp;&nbsp;vs Prior Year</div>
      </div>
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div class="label">LATEST MONTH ATTAINMENT VS OP'26</div>
        <div class="value" style="color:#0000C9;">FIN_NET_MTD_ATT</div>
        <div class="sub">FIN_NET_MTD_DELTA</div>
      </div>
      <div style="background:#EEF5FB;border-radius:12px;padding:14px 16px;text-align:left;border:1px solid rgba(15,23,42,0.05);transition:transform 0.25s cubic-bezier(0.16,1,0.3,1),box-shadow 0.25s cubic-bezier(0.4,0,0.2,1);flex:1;" onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 6px 16px rgba(15,23,42,0.07),0 2px 4px rgba(15,23,42,0.04)'" onmouseout="this.style.transform='';this.style.boxShadow=''">
        <div class="label">YTD ATTAINMENT VS OP'26</div>
        <div class="value" style="color:#0000C9;">FIN_NET_YTD_ATT</div>
        <div class="sub">FIN_NET_YTD_DELTA</div>
      </div>
    </div>

        <div class="card">
      <div class="card-title">Monthly Net Revenue vs Budget (2026)</div>
      <div class="card-subtitle">Actuals · STLY · Budget OP'26 · $M</div>
      FIN_NET_CHART_PLACEHOLDER
      <div class="legend" style="font-size:10px;margin-top:8px;">
        <div class="legend-item"><div class="legend-dot" style="background:#7C6CFC;border-radius:2px;width:12px;"></div>Actual</div>
        <div class="legend-item"><div class="legend-dot" style="background:#0000C9"></div>Actual (STLY)</div>
        <div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#9ca3af" stroke-width="2" stroke-dasharray="3"/></svg>Budget</div>
      </div>
    </div>

    </div>
</div>
</div>

        </section>

        <!-- 3. AGENTS -->
        <section class="section" id="agents" data-label="CoWork Agents">
            <div class="section-head"><h2>Snowflake CoWork Agents</h2><p>AI-powered analytical agents for automated insights and conversational data exploration.</p></div>
            <div class="grid">
                <a class="card" href="https://app.us-east-1.privatelink.snowflakecomputing.com/pfe/amerprod01/#/ai" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s3"><svg viewBox="0 0 24 24"><rect x="6" y="4" width="12" height="7" rx="2"/><path d="M9 11v4M15 11v4M8 18h8M12 15v3M7 4V2M17 4V2"/></svg></span></div><div class="card-title">Migraine NPA Agent</div><div class="card-desc">Conversational NPA data querying and automated insight generation.</div><span class="dest-pill dest-agent"><span class="swatch">&#9632;</span>Cortex Agent</span></a>
                <a class="card" href="https://app.us-east-1.privatelink.snowflakecomputing.com/pfe/amerprod01/#/ai" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s3"><svg viewBox="0 0 24 24"><rect x="6" y="4" width="12" height="7" rx="2"/><path d="M9 11v4M15 11v4M8 18h8M12 15v3M7 4V2M17 4V2"/></svg></span><span class="badge weekly">Weekly</span></div><div class="card-title">LAAD Weekly Agent</div><div class="card-desc">Weekly LAAD data processing and trend surfacing.</div><span class="dest-pill dest-agent"><span class="swatch">&#9632;</span>Cortex Agent</span></a>
                <a class="card" href="https://app.us-east-1.privatelink.snowflakecomputing.com/pfe/amerprod01/#/ai" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s3"><svg viewBox="0 0 24 24"><rect x="6" y="4" width="12" height="7" rx="2"/><path d="M9 11v4M15 11v4M8 18h8M12 15v3M7 4V2M17 4V2"/></svg></span><span class="badge monthly">Monthly</span></div><div class="card-title">eLAAD Monthly Agent</div><div class="card-desc">Monthly eLAAD aggregation and reporting automation.</div><span class="dest-pill dest-agent"><span class="swatch">&#9632;</span>Cortex Agent</span></a>
            </div>
        </section>

        <!-- 4. DELIVERABLES -->
        <section class="section" id="deliverables" data-label="Analytics Deliverables">
            <div class="section-head"><h2>Analytics Deliverables</h2><p>PowerPoint and Excel deliverables for leadership and cross-functional stakeholders.</p></div>
            <div class="grid">
                <a class="card" href="https://pfizer.sharepoint.com/:p:/s/MigraineAnalytics/IQA7CzbiZxm9SYbErQAXLG5kASpgzKHmNLmQz8yBenX6kNc?e=KgQ7yz" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s4"><svg viewBox="0 0 24 24"><path d="M4 4h16a1 1 0 011 1v14a1 1 0 01-1 1H4a1 1 0 01-1-1V5a1 1 0 011-1z"/><path d="M8 9h8M8 13h6M8 17h4"/></svg></span></div><div class="card-title">Nurtec Offsite Deck</div><div class="card-desc">Strategic offsite presentation with market overview.</div><div class="card-updated">Updated Feb 24, 2026</div><span class="dest-pill dest-ppt"><span class="swatch">&#9632;</span>PowerPoint</span></a>
                <a class="card" href="https://pfizer.sharepoint.com/:p:/s/MigraineAnalytics/IQBvAxCrDDmeTbpe9RH14ha6AeG6J09GqT4ft1vRBHtmTk8?e=qFF3D9" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s4"><svg viewBox="0 0 24 24"><circle cx="6" cy="18" r="2"/><circle cx="18" cy="18" r="2"/><circle cx="12" cy="6" r="2"/><path d="M6 16V8l6-2M18 16V8l-6-2"/></svg></span></div><div class="card-title">oCGRP Market Dynamics</div><div class="card-desc">Competitive dynamics and market share analysis.</div><div class="card-updated">Updated Feb 4, 2026</div><span class="dest-pill dest-ppt"><span class="swatch">&#9632;</span>PPT &middot; Excel</span></a>
                <a class="card" href="#" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s4"><svg viewBox="0 0 24 24"><path d="M14 3v4a1 1 0 001 1h4"/><path d="M17 21H7a2 2 0 01-2-2V5a2 2 0 012-2h7l5 5v11a2 2 0 01-2 2z"/><path d="M9 17v-6M12 17v-1M15 17v-4"/></svg></span></div><div class="card-title">Nurtec Deep-Dives</div><div class="card-desc">Detailed deep-dive analyses across key metrics.</div><span class="dest-pill dest-ppt"><span class="swatch">&#9632;</span>PowerPoint</span></a>
            </div>
        </section>

        <!-- 5. DOCS -->
        <section class="section" id="docs" data-label="Business Rule Docs">
            <div class="section-head"><h2>Business Rule Documentation</h2><p>Data definitions, metric calculations, business logic, and reporting standards.</p></div>
            <div class="grid">
                <a class="card" href="https://pfizer.sharepoint.com/:p:/s/MigraineAnalytics/IQBdBFg1vxdFR6HZvdxtKVPZAUdmK-vTwF4Trvm4JdEBru4?e=TG0guS" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s5"><svg viewBox="0 0 24 24"><path d="M3 19a9 9 0 019-9 9 9 0 019 9"/><path d="M3 19h18M12 10V3"/><path d="M7.8 4.8L12 3l4.2 1.8"/></svg></span></div><div class="card-title">Migraine Market Overview</div><div class="card-desc">Market landscape and key therapeutic area definitions.</div><div class="card-updated">Updated Dec 23, 2025</div><span class="dest-pill dest-doc"><span class="swatch">&#9632;</span>Doc</span></a>
                <a class="card" href="#" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s5"><svg viewBox="0 0 24 24"><path d="M9 5H7a2 2 0 00-2 2v12a2 2 0 002 2h10a2 2 0 002-2V7a2 2 0 00-2-2h-2"/><rect x="9" y="3" width="6" height="4" rx="1"/><path d="M9 12l2 2 4-4"/></svg></span></div><div class="card-title">BR Alignment</div><div class="card-desc">Cross-functional business rule alignment documentation.</div><span class="dest-pill dest-doc"><span class="swatch">&#9632;</span>Doc</span></a>
                <a class="card" href="https://pfizer.sharepoint.com/:p:/s/MigraineAnalytics/IQA9ft3L-N9VT4YQzW36UFsXAQU2ZNBXbO3UjFvDwa6NV5E?e=8ytpZI" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s5"><svg viewBox="0 0 24 24"><path d="M12 4L4 8l8 4 8-4-8-4zM4 12l8 4 8-4M4 16l8 4 8-4"/></svg></span></div><div class="card-title">IIS BR Master Deck</div><div class="card-desc">Master business rules deck for IIS analytics.</div><div class="card-updated">Updated Jan 1, 2026</div><span class="dest-pill dest-ppt"><span class="swatch">&#9632;</span>PowerPoint</span></a>
                <a class="card" href="https://pfizer.sharepoint.com/:p:/s/MigraineAnalytics/IQBOAt2MOv_OTqhwLardlsHPAZdTa2hsvBq5pUuZfNbVWTM?e=tDWmaa" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s5"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="3"/><circle cx="12" cy="4" r="1.5"/><circle cx="12" cy="20" r="1.5"/><circle cx="4" cy="12" r="1.5"/><circle cx="20" cy="12" r="1.5"/><path d="M12 7v2M12 15v2M7 12h2M15 12h2"/></svg></span></div><div class="card-title">IIS Migraine Data Ecosystem</div><div class="card-desc">Data sources, flows, and ecosystem mapping.</div><div class="card-updated">Updated Jan 1, 2026</div><span class="dest-pill dest-doc"><span class="swatch">&#9632;</span>Doc</span></a>
                <a class="card" href="https://pfizer.sharepoint.com/:p:/s/MigraineAnalytics/IQADCiim_iD7QbxZE7Fm3F89Ab8O10p4xzPmog3z4X7KQ2g?e=jA6LCi" target="_blank" rel="noopener"><div class="card-top"><span class="icon-chip chip-s5"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="3"/><path d="M12 5V3M12 21v-2M16.95 7.05l1.41-1.41M5.64 18.36l1.41-1.41M19 12h2M3 12h2M16.95 16.95l1.41 1.41M5.64 5.64l1.41 1.41"/></svg></span></div><div class="card-title">Migraine Dashboard BR</div><div class="card-desc">Dashboard-specific metric logic and standards.</div><div class="card-updated">Updated Feb 1, 2026</div><span class="dest-pill dest-doc"><span class="swatch">&#9632;</span>Doc</span></a>
            </div>
        </section>

    </main>
</div>
</div>

<script>
(function() {
    'use strict';

    // Dropdown
    window.toggleDropdown = function(id, ev) {
        ev.stopPropagation();
        document.querySelectorAll('.dropdown.show').forEach(function(d) { if (d.id !== id) d.classList.remove('show'); });
        document.getElementById(id).classList.toggle('show');
    };
    document.addEventListener('click', function(e) {
        if (!e.target.closest('.dropdown-wrap')) {
            document.querySelectorAll('.dropdown.show').forEach(function(d) { d.classList.remove('show'); });
        }
    });

    // Sidebar main navigation
    var nav = document.getElementById('sidebarNav');
    var items = nav.querySelectorAll('.nav-item');
    var sections = {};
    items.forEach(function(it) { sections[it.dataset.target] = document.getElementById(it.dataset.target); });
    var toolbarSection = document.getElementById('toolbarSection');
    var newsletterSub = document.getElementById('newsletterSub');
    var nlGridView = document.getElementById('nl-grid-view');
    var nlDeepdiveView = document.getElementById('nl-deepdive-view');
    var heroEl = document.querySelector('.hero');
    var dividerEl = document.querySelector('.workspace-divider');
    var inDeepdive = false;

    function setActive(item) {
        items.forEach(function(i) { i.classList.remove('active'); });
        item.classList.add('active');
    }

    function exitDeepdive() {
        inDeepdive = false;
        nlGridView.style.display = '';
        nlDeepdiveView.style.display = 'none';
        newsletterSub.classList.remove('is-open');
        heroEl.style.display = '';
        dividerEl.style.display = '';
        
        newsletterSub.querySelectorAll('.nav-sub-item').forEach(function(s) { s.classList.remove('active'); });
    }

    function enterDeepdive(panelId) {
        inDeepdive = true;
        nlGridView.style.display = 'none';
        nlDeepdiveView.style.display = '';
        newsletterSub.classList.add('is-open');
        heroEl.style.display = 'none';
        dividerEl.style.display = 'none';
        
        // Activate the correct nl-tab
        nlDeepdiveView.querySelectorAll('.nl-tab').forEach(function(p) { p.classList.remove('active'); });
        var target = document.getElementById(panelId + '-tab');
        if (target) target.classList.add('active');
        // Activate the correct sub-nav item
        newsletterSub.querySelectorAll('.nav-sub-item').forEach(function(s) { s.classList.remove('active'); });
        var matchingSub = newsletterSub.querySelector('[data-panel="' + panelId + '"]');
        if (matchingSub) matchingSub.classList.add('active');
        // Trigger resize for Plotly charts
        setTimeout(function() {
            document.querySelectorAll('.plotly-graph-div').forEach(function(gd) {
                if (gd && gd.data && typeof Plotly !== 'undefined') {
                    Plotly.relayout(gd, {autosize: true});
                }
            });
        }, 150);
    }

    var switching = false;
    function showSection(id, label) {
        if (switching) return;
        var current = document.querySelector('.section.is-active');
        var next = sections[id];
        if (!next || next === current) return;
        switching = true;
        if (current) current.classList.remove('is-visible');
        setTimeout(function() {
            if (current) current.classList.remove('is-active');
            var content = document.querySelector('.content');
            if (content) content.scrollTop = 0;
            next.classList.add('is-active');
            void next.offsetWidth;
            next.classList.add('is-visible');
            // Trigger resize for Plotly charts - staggered for reliability
            setTimeout(function() {
                document.querySelectorAll('.plotly-graph-div').forEach(function(gd) {
                    if (gd && gd.data && typeof Plotly !== 'undefined') {
                        Plotly.relayout(gd, {autosize: true});
                    }
                });
            }, 400);
            if (toolbarSection && label) toolbarSection.textContent = label;
            switching = false;
        }, 220);
    }

    items.forEach(function(item) {
        item.addEventListener('click', function(e) {
            e.preventDefault();
            setActive(item);
            var label = sections[item.dataset.target] ? sections[item.dataset.target].dataset.label : '';
            showSection(item.dataset.target, label);

            // When clicking Newsletter from nav, show card grid (reset deepdive)
            if (item.dataset.target === 'newsletter') {
                exitDeepdive();
            } else {
                newsletterSub.classList.remove('is-open');
                if (inDeepdive) exitDeepdive();
            }

            // Restore hero for non-deepdive views
            if (item.dataset.target !== 'newsletter' || !inDeepdive) {
                heroEl.style.display = '';
                dividerEl.style.display = '';
            }
        });
    });

    // Newsletter card clicks → enter deepdive
    var nlCards = document.querySelectorAll('#nl-grid-view .card[data-nl-panel]');
    nlCards.forEach(function(card) {
        card.addEventListener('click', function(e) {
            e.preventDefault();
            var panelId = card.dataset.nlPanel;
            if (panelId) enterDeepdive(panelId);
        });
    });

    // Newsletter sub-nav items (sidebar) → switch panels
    var subItems = newsletterSub.querySelectorAll('.nav-sub-item[data-panel]');
    subItems.forEach(function(subItem) {
        subItem.addEventListener('click', function(e) {
            e.preventDefault();
            newsletterSub.querySelectorAll('.nav-sub-item').forEach(function(s) { s.classList.remove('active'); });
            subItem.classList.add('active');
            nlDeepdiveView.querySelectorAll('.nl-tab').forEach(function(p) { p.classList.remove('active'); });
            var target = document.getElementById(subItem.dataset.panel + '-tab');
            if (target) target.classList.add('active');
        });
    });

    // NPA Download dropdown helpers
    function updateNpaDownload(metricLabel) {
        var m = metricLabel.toLowerCase() === 'trx' ? 'trx' : 'nbrx';
        var brandEl = document.getElementById('npa-dl-brand');
        if (brandEl) {
            var tmpl = brandEl.getAttribute('data-npa-dl-label');
            brandEl.querySelector('span').textContent = tmpl.replace('{M}', metricLabel);
            brandEl.href = brandEl.getAttribute('data-' + m + '-href');
            brandEl.download = brandEl.getAttribute('data-' + m + '-fname');
        }
        // Also update channel link metric
        var chEl = document.getElementById('npa-dl-channel');
        if (chEl) {
            var activeBrand = 'nurtec';
            var chNurtec = document.getElementById('channel-nurtec');
            var chUbrelvy = document.getElementById('channel-ubrelvy');
            var chQulipta = document.getElementById('channel-qulipta');
            if (chNurtec && chNurtec.classList.contains('active')) activeBrand = 'nurtec';
            if (chUbrelvy && chUbrelvy.classList.contains('active')) activeBrand = 'ubrelvy';
            if (chQulipta && chQulipta.classList.contains('active')) activeBrand = 'qulipta';
            var brandNames = {nurtec:'Nurtec',ubrelvy:'Ubrelvy',qulipta:'Qulipta'};
            var tmpl2 = chEl.getAttribute('data-npa-dl-label');
            chEl.querySelector('span').textContent = tmpl2.replace('{B}', brandNames[activeBrand]).replace('{M}', metricLabel);
            chEl.href = chEl.getAttribute('data-' + activeBrand + '-' + m + '-href');
            chEl.download = 'NPA_Channel_' + brandNames[activeBrand] + '_' + metricLabel + '.xlsx';
        }
    }
    function updateNpaChannelDownload(brand) {
        var chEl = document.getElementById('npa-dl-channel');
        if (!chEl) return;
        var m = document.getElementById('npa-trx-pill').classList.contains('active') ? 'trx' : 'nbrx';
        var metricLabel = m === 'trx' ? 'TRx' : 'NBRx';
        var brandNames = {nurtec:'Nurtec',ubrelvy:'Ubrelvy',qulipta:'Qulipta'};
        var tmpl = chEl.getAttribute('data-npa-dl-label');
        chEl.querySelector('span').textContent = tmpl.replace('{B}', brandNames[brand]).replace('{M}', metricLabel);
        chEl.href = chEl.getAttribute('data-' + brand + '-' + m + '-href');
        chEl.download = 'NPA_Channel_' + brandNames[brand] + '_' + metricLabel + '.xlsx';
    }
    function updateApDownload() {
        var isAcute = document.getElementById('acute-view').style.display !== 'none';
        var view = isAcute ? 'acute' : 'prev';
        var viewLabel = isAcute ? 'Acute' : 'Preventive';
        var brandEl = document.getElementById('ap-dl-brand');
        if (brandEl) {
            brandEl.href = brandEl.getAttribute('data-' + view + '-href');
            brandEl.download = brandEl.getAttribute('data-' + view + '-fname');
            brandEl.querySelector('span').textContent = viewLabel + ' - Brand Competitive';
        }
        var chEl = document.getElementById('ap-dl-channel');
        if (chEl) {
            var activeBrand = 'nurtec';
            if (isAcute) {
                var acUb = document.getElementById('acute-ch-ubrelvy');
                if (acUb && acUb.classList.contains('active')) activeBrand = 'ubrelvy';
            } else {
                var prQu = document.getElementById('prev-ch-qulipta');
                if (prQu && prQu.classList.contains('active')) activeBrand = 'qulipta';
            }
            var brandNames = {nurtec:'Nurtec',ubrelvy:'Ubrelvy',qulipta:'Qulipta'};
            chEl.href = chEl.getAttribute('data-' + view + '-' + activeBrand + '-href');
            chEl.download = 'NPA_' + viewLabel + '_Channel_' + brandNames[activeBrand] + '.xlsx';
            chEl.querySelector('span').textContent = viewLabel + ' - Channel ' + brandNames[activeBrand];
        }
    }

    // Newsletter sub-tab functions
    window.switchNpaSubtab = function(subtab) {
        var pills = document.querySelectorAll('#npa-subtabs .pill');
        pills.forEach(function(el) { el.classList.remove('active'); });
        if (subtab === 'retail') {
            pills[0].classList.add('active');
            document.getElementById('npa-retail').style.display = 'block';
            document.getElementById('npa-acute-prev').style.display = 'none';
        } else {
            pills[1].classList.add('active');
            document.getElementById('npa-retail').style.display = 'none';
            document.getElementById('npa-acute-prev').style.display = 'block';
            window.dispatchEvent(new Event('resize'));
            setTimeout(function(){ window.dispatchEvent(new Event('resize')); }, 300);
        }
    };
    window.switchAcutePrev = function(view) {
        var pills = document.querySelectorAll('#acute-prev-toggle .pill');
        pills.forEach(function(el) { el.classList.remove('active'); });
        if (view === 'acute') {
            pills[0].classList.add('active');
            document.getElementById('acute-view').style.display = 'block';
            document.getElementById('preventive-view').style.display = 'none';
        } else {
            pills[1].classList.add('active');
            document.getElementById('acute-view').style.display = 'none';
            document.getElementById('preventive-view').style.display = 'block';
        }
        window.dispatchEvent(new Event('resize'));
        setTimeout(function(){ window.dispatchEvent(new Event('resize')); }, 300);
        updateApDownload();
    };
    
    window.switchFinancial = function(view) {
        document.querySelectorAll('#financial-toggle .pill').forEach(function(el) { el.classList.remove('active'); });
        event.currentTarget.classList.add('active');
        if (view === 'gross') {
            document.getElementById('gross-view').style.display = 'block';
            document.getElementById('net-view').style.display = 'none';
        } else {
            document.getElementById('gross-view').style.display = 'none';
            document.getElementById('net-view').style.display = 'block';
        }
        setTimeout(function() {
            document.querySelectorAll('.plotly-graph-div').forEach(function(gd) {
                if (gd && gd.data && typeof Plotly !== 'undefined') { Plotly.relayout(gd, {autosize: true}); }
            });
        }, 100);
    };
    
    
    window.switchXponentMetric = function(metric) {
        document.getElementById('xp-trx-pill').classList.toggle('active', metric === 'trx');
        document.getElementById('xp-nrx-pill').classList.toggle('active', metric === 'nrx');
        var label = metric === 'trx' ? 'TRx' : 'NRx';
        document.querySelectorAll('[data-xp-metric]').forEach(function(el) {
            var template = el.getAttribute('data-xp-metric');
            el.textContent = template.replace('{XM}', label);
        });
        // Toggle Xponent charts
        var pTrx = document.getElementById('xpt-payer-trx');
        var pNrx = document.getElementById('xpt-payer-nrx');
        var cTrx = document.getElementById('xpt-ch-trx');
        var cNrx = document.getElementById('xpt-ch-nrx');
        if (metric === 'trx') {
            pTrx.style.height = ''; pTrx.style.visibility = '';
            pNrx.style.height = '0'; pNrx.style.visibility = 'hidden';
            cTrx.style.height = ''; cTrx.style.visibility = '';
            cNrx.style.height = '0'; cNrx.style.visibility = 'hidden';
        } else {
            pTrx.style.height = '0'; pTrx.style.visibility = 'hidden';
            pNrx.style.height = ''; pNrx.style.visibility = '';
            cTrx.style.height = '0'; cTrx.style.visibility = 'hidden';
            cNrx.style.height = ''; cNrx.style.visibility = '';
        }
        window.dispatchEvent(new Event('resize'));
        // Toggle Xponent table tbody (TRx/NRx)
        document.querySelectorAll('.xpt-trx-body').forEach(function(el) {
            el.style.display = metric === 'trx' ? '' : 'none';
        });
        document.querySelectorAll('.xpt-nrx-body').forEach(function(el) {
            el.style.display = metric === 'nrx' ? '' : 'none';
        });
        // Toggle download dropdown labels and hrefs
        document.querySelectorAll('[data-xpt-dl-label]').forEach(function(el) {
            var template = el.getAttribute('data-xpt-dl-label');
            el.querySelector('span').textContent = template.replace('{XM}', label);
            el.href = el.getAttribute('data-' + metric + '-href');
            el.download = el.getAttribute('data-' + metric + '-fname');
        });
    };
    
    window.filterPayerTable = function(payer) {
        var rows = document.querySelectorAll('#payer-table-body tr');
        var shown = 0;
        rows.forEach(function(row) {
            var rowPayer = row.getAttribute('data-payer');
            if (rowPayer === payer) {
                row.style.display = shown < 5 ? '' : 'none';
                shown++;
            } else {
                row.style.display = 'none';
            }
        });
    };
    filterPayerTable('commercial');

    window.switchAcuteChannel = function(brand) {
        document.getElementById('acute-ch-nurtec').classList.toggle('active', brand === 'nurtec');
        document.getElementById('acute-ch-ubrelvy').classList.toggle('active', brand === 'ubrelvy');
        var label = brand === 'nurtec' ? 'Nurtec' : 'Ubrelvy';
        document.getElementById('acute-ch-title').innerHTML = label + ' Acute \u2014 Channel Performance View';
        document.getElementById('ap-ch-nurtec-acute').style.display = brand === 'nurtec' ? 'block' : 'none';
        document.getElementById('ap-ch-ubrelvy-acute').style.display = brand === 'ubrelvy' ? 'block' : 'none';
        window.dispatchEvent(new Event('resize'));
        updateApDownload();
    };
    window.switchPrevChannel = function(brand) {
        document.getElementById('prev-ch-nurtec').classList.toggle('active', brand === 'nurtec');
        document.getElementById('prev-ch-qulipta').classList.toggle('active', brand === 'qulipta');
        var label = brand === 'nurtec' ? 'Nurtec' : 'Qulipta';
        document.getElementById('prev-ch-title').innerHTML = label + ' Preventive \u2014 Channel Performance View';
        document.getElementById('ap-ch-nurtec-prev').style.display = brand === 'nurtec' ? 'block' : 'none';
        document.getElementById('ap-ch-qulipta-prev').style.display = brand === 'qulipta' ? 'block' : 'none';
        window.dispatchEvent(new Event('resize'));
        updateApDownload();
    };
    window.switchChannelBrand = function(brand) {
        document.getElementById('channel-nurtec').classList.toggle('active', brand === 'nurtec');
        document.getElementById('channel-ubrelvy').classList.toggle('active', brand === 'ubrelvy');
        document.getElementById('channel-qulipta').classList.toggle('active', brand === 'qulipta');
        var labels = {'nurtec':'Nurtec','ubrelvy':'Ubrelvy','qulipta':'Qulipta'};
        var el = document.getElementById('overall-ch-title');
        var metric = document.getElementById('npa-trx-pill').classList.contains('active') ? 'TRx' : 'NBRx';
        if (el) {
            el.innerHTML = labels[brand] + ' NPA \u2014 Channel Performance View (' + metric + ')';
            el.setAttribute('data-metric-toggle', labels[brand] + ' NPA \u2014 Channel Performance View ({M})');
        }
        // Switch channel chart based on brand + current metric
        var metric = document.getElementById('npa-trx-pill').classList.contains('active') ? 'trx' : 'nbrx';
        var brands = ['nurtec','ubrelvy','qulipta'];
        brands.forEach(function(b) {
            var trxEl = document.getElementById('ch-' + b + '-trx');
            var nbrxEl = document.getElementById('ch-' + b + '-nbrx');
            if (trxEl) trxEl.style.display = (b === brand && metric === 'trx') ? 'block' : 'none';
            if (nbrxEl) nbrxEl.style.display = (b === brand && metric === 'nbrx') ? 'block' : 'none';
        });
        updateNpaChannelDownload(brand);
        window.dispatchEvent(new Event('resize'));
    };
    window.toggleNpaMetric = function(metric) {
        document.getElementById('npa-trx-pill').classList.toggle('active', metric === 'trx');
        document.getElementById('npa-nbrx-pill').classList.toggle('active', metric === 'nbrx');
        // Update all elements with data-metric-toggle attribute
        var label = metric === 'trx' ? 'TRx' : 'NBRx';
        document.querySelectorAll('[data-metric-toggle]').forEach(function(el) {
            var template = el.getAttribute('data-metric-toggle');
            el.textContent = template.replace('{M}', label);
        });
        // Toggle TRx/NBRx charts
        var trxChart = document.getElementById('trx-brand-chart');
        var nbrxChart = document.getElementById('nbrx-brand-chart');
        if (trxChart && nbrxChart) {
            trxChart.style.display = metric === 'trx' ? 'block' : 'none';
            nbrxChart.style.display = metric === 'nbrx' ? 'block' : 'none';
        }
        // Toggle channel charts based on metric + current brand
        var brands = ['nurtec','ubrelvy','qulipta'];
        var activeBrand = 'nurtec';
        var chNurtec = document.getElementById('channel-nurtec');
        var chUbrelvy = document.getElementById('channel-ubrelvy');
        var chQulipta = document.getElementById('channel-qulipta');
        if (chNurtec && chNurtec.classList.contains('active')) activeBrand = 'nurtec';
        if (chUbrelvy && chUbrelvy.classList.contains('active')) activeBrand = 'ubrelvy';
        if (chQulipta && chQulipta.classList.contains('active')) activeBrand = 'qulipta';
        brands.forEach(function(b) {
            var trxEl = document.getElementById('ch-' + b + '-trx');
            var nbrxEl = document.getElementById('ch-' + b + '-nbrx');
            if (trxEl) trxEl.style.display = (b === activeBrand && metric === 'trx') ? 'block' : 'none';
            if (nbrxEl) nbrxEl.style.display = (b === activeBrand && metric === 'nbrx') ? 'block' : 'none';
        });
        // Toggle NPA table tbody (TRx/NBRx)
        document.querySelectorAll('.npa-table-trx').forEach(function(el) {
            el.style.display = metric === 'trx' ? '' : 'none';
        });
        document.querySelectorAll('.npa-table-nbrx').forEach(function(el) {
            el.style.display = metric === 'nbrx' ? '' : 'none';
        });
        // Update NPA download dropdown
        updateNpaDownload(metric === 'trx' ? 'TRx' : 'NBRx');
        window.dispatchEvent(new Event('resize'));
    };
    // Definitive fix: MutationObserver watches for any element becoming visible,
    // then resizes all Plotly charts. Handles hidden tabs, section switches, toggles.
    function resizeVisiblePlotly() {
        document.querySelectorAll('.plotly-graph-div').forEach(function(gd) {
            if (gd && gd.data && typeof Plotly !== 'undefined' && gd.offsetWidth > 0) {
                Plotly.relayout(gd, {autosize: true});
            }
        });
    }
    var observer = new MutationObserver(function(mutations) {
        var needsResize = false;
        mutations.forEach(function(m) {
            if (m.type === 'attributes' && (m.attributeName === 'style' || m.attributeName === 'class')) {
                needsResize = true;
            }
        });
        if (needsResize) {
            setTimeout(resizeVisiblePlotly, 50);
            setTimeout(resizeVisiblePlotly, 300);
        }
    });
    observer.observe(document.body, {attributes: true, subtree: true, attributeFilter: ['style', 'class']});
    // Also fire on initial load with longer delays for iframe settling
    setTimeout(resizeVisiblePlotly, 500);
    setTimeout(resizeVisiblePlotly, 1000);
    setTimeout(resizeVisiblePlotly, 2000);
    setTimeout(resizeVisiblePlotly, 3500);
})();
</script>
</body>
</html>
"""


# Inject live Plotly chart - replace the entire chart container
trx_nbrx_html = '<div id="trx-brand-chart" style="width:100%;overflow:hidden;">' + brand_chart_svg + '</div><div id="nbrx-brand-chart" style="width:100%;overflow:hidden;display:none;">' + nbrx_chart_svg + '</div>'
html_content = html_content.replace('<div class="chart-container">\n          <svg class="chart" preserveAspectRatio="none" viewBox="0 0 800 250">\nBRAND_CHART_DATA_PLACEHOLDER\n          </svg>\n        </div>', trx_nbrx_html)

# Dynamic "Updated" date (current date when backend starts)
from datetime import datetime as _dt
html_content = html_content.replace('APP_LAST_UPDATED_DATE', _dt.now().strftime('%b %d, %Y'))

# Dynamic Data Availability refresh dates
_refresh_df = load_data_refresh()
_refresh_html = ''
for _, row in _refresh_df.iterrows():
    src = row['DATA_SOURCE']
    raw_date = str(row['REFRESH_DATE']).strip('"')
    try:
        dt_obj = _dt.strptime(raw_date, '%Y-%m-%d')
        formatted = dt_obj.strftime('%b %d, %Y')
    except Exception:
        formatted = raw_date
    _refresh_html += f'<div class="dropdown-item"><span class="src">{src}</span><span class="date">{formatted}</span></div>\n'
html_content = html_content.replace('DATA_REFRESH_ITEMS_PLACEHOLDER', _refresh_html)


channel_html = '<div id="ch-nurtec-trx" style="width:100%;overflow:hidden;">' + ch_nurtec_trx + '</div>'
channel_html += '<div id="ch-nurtec-nbrx" style="width:100%;overflow:hidden;display:none;">' + ch_nurtec_nbrx + '</div>'
channel_html += '<div id="ch-ubrelvy-trx" style="width:100%;overflow:hidden;display:none;">' + ch_ubrelvy_trx + '</div>'
channel_html += '<div id="ch-ubrelvy-nbrx" style="width:100%;overflow:hidden;display:none;">' + ch_ubrelvy_nbrx + '</div>'
channel_html += '<div id="ch-qulipta-trx" style="width:100%;overflow:hidden;display:none;">' + ch_qulipta_trx + '</div>'
channel_html += '<div id="ch-qulipta-nbrx" style="width:100%;overflow:hidden;display:none;">' + ch_qulipta_nbrx + '</div>'
# Acute Channel chart injection (brand toggle: nurtec/ubrelvy, side by side TRx|NBRx)
acute_ch_legend = '<div class="legend" style="margin-top:8px;justify-content:center;"><div class="legend-item"><div class="legend-dot" style="background:#0891b2"></div>Retail Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#0891b2" stroke-width="2" stroke-dasharray="3"/></svg>Retail STLY</div><div class="legend-item"><div class="legend-dot" style="background:#7c3aed"></div>Mail-Order Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#7c3aed" stroke-width="2" stroke-dasharray="3"/></svg>Mail-Order STLY</div><div class="legend-item"><div class="legend-dot" style="background:#9ca3af"></div>LTC Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#9ca3af" stroke-width="2" stroke-dasharray="3"/></svg>LTC STLY</div></div>'
acute_ch_html = '<div id="ap-ch-nurtec-acute" style="width:100%;"><div class="split-section chart-split"><div class="split-col"><div class="section-label">TRx</div>' + ap_ch_nurtec_acute_trx + '</div><div class="split-col"><div class="section-label">NBRx</div>' + ap_ch_nurtec_acute_nbrx + '</div></div></div>'
acute_ch_html += '<div id="ap-ch-ubrelvy-acute" style="width:100%;display:none;"><div class="split-section chart-split"><div class="split-col"><div class="section-label">TRx</div>' + ap_ch_ubrelvy_acute_trx + '</div><div class="split-col"><div class="section-label">NBRx</div>' + ap_ch_ubrelvy_acute_nbrx + '</div></div></div>'
acute_ch_html += acute_ch_legend
html_content = html_content.replace('ACUTE_CHANNEL_CHART_PLACEHOLDER', acute_ch_html)

# Preventive Channel chart injection (brand toggle: nurtec/qulipta, side by side TRx|NBRx)
prev_ch_legend = '<div class="legend" style="margin-top:8px;justify-content:center;"><div class="legend-item"><div class="legend-dot" style="background:#0891b2"></div>Retail Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#0891b2" stroke-width="2" stroke-dasharray="3"/></svg>Retail STLY</div><div class="legend-item"><div class="legend-dot" style="background:#7c3aed"></div>Mail-Order Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#7c3aed" stroke-width="2" stroke-dasharray="3"/></svg>Mail-Order STLY</div><div class="legend-item"><div class="legend-dot" style="background:#9ca3af"></div>LTC Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#9ca3af" stroke-width="2" stroke-dasharray="3"/></svg>LTC STLY</div></div>'
prev_ch_html = '<div id="ap-ch-nurtec-prev" style="width:100%;"><div class="split-section chart-split"><div class="split-col"><div class="section-label">TRx</div>' + ap_ch_nurtec_prev_trx + '</div><div class="split-col"><div class="section-label">NBRx</div>' + ap_ch_nurtec_prev_nbrx + '</div></div></div>'
prev_ch_html += '<div id="ap-ch-qulipta-prev" style="width:100%;display:none;"><div class="split-section chart-split"><div class="split-col"><div class="section-label">TRx</div>' + ap_ch_qulipta_prev_trx + '</div><div class="split-col"><div class="section-label">NBRx</div>' + ap_ch_qulipta_prev_nbrx + '</div></div></div>'
prev_ch_html += prev_ch_legend
html_content = html_content.replace('PREV_CHANNEL_CHART_PLACEHOLDER', prev_ch_html)

# NPA Overall channel chart injection (must come AFTER acute/prev channel replacements)
html_content = html_content.replace('CHANNEL_CHART_PLACEHOLDER', channel_html)

# Acute/Preventive chart injection - side by side TRx | NBRx
acute_legend = '<div class="legend" style="margin-top:8px;justify-content:center;"><div class="legend-item"><div class="legend-dot" style="background:#7C6CFC"></div>Nurtec Acute Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#7C6CFC" stroke-width="2" stroke-dasharray="3"/></svg>Nurtec Acute STLY</div><div class="legend-item"><div class="legend-dot" style="background:#4ADE80"></div>Ubrelvy Acute Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#4ADE80" stroke-width="2" stroke-dasharray="3"/></svg>Ubrelvy Acute STLY</div></div>'
acute_html = '<div class="split-section chart-split"><div class="split-col"><div class="section-label">TRx</div>' + acute_trx_chart + '</div><div class="split-col"><div class="section-label">NBRx</div>' + acute_nbrx_chart + '</div></div>' + acute_legend
html_content = html_content.replace('ACUTE_BRAND_CHART_PLACEHOLDER', acute_html)

prev_legend = '<div class="legend" style="margin-top:8px;justify-content:center;"><div class="legend-item"><div class="legend-dot" style="background:#7C6CFC"></div>Nurtec Prev Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#7C6CFC" stroke-width="2" stroke-dasharray="3"/></svg>Nurtec Prev STLY</div><div class="legend-item"><div class="legend-dot" style="background:#FB923C"></div>Qulipta Prev Actuals</div><div class="legend-item"><svg width="20" height="2" style="margin-right:4px"><line x1="0" y1="1" x2="20" y2="1" stroke="#FB923C" stroke-width="2" stroke-dasharray="3"/></svg>Qulipta Prev STLY</div></div>'
prev_html = '<div class="split-section chart-split"><div class="split-col"><div class="section-label">TRx</div>' + prev_trx_chart + '</div><div class="split-col"><div class="section-label">NBRx</div>' + prev_nbrx_chart + '</div></div>' + prev_legend
html_content = html_content.replace('PREV_BRAND_CHART_PLACEHOLDER', prev_html)

# Xponent Trends chart injection
html_content = html_content.replace('XPT_PAYER_TRX_PLACEHOLDER', xpt_payer_trx_html)
html_content = html_content.replace('XPT_PAYER_NRX_PLACEHOLDER', xpt_payer_nrx_html)
html_content = html_content.replace('XPT_CH_TRX_PLACEHOLDER', xpt_ch_trx_html)
html_content = html_content.replace('XPT_CH_NRX_PLACEHOLDER', xpt_ch_nrx_html)
# Xponent download Excel base64 injection
html_content = html_content.replace('XPT_DL_PAYER_TRX_B64', _xpt_dl_payer_trx_b64)
html_content = html_content.replace('XPT_DL_PAYER_NRX_B64', _xpt_dl_payer_nrx_b64)
html_content = html_content.replace('XPT_DL_CH_TRX_B64', _xpt_dl_ch_trx_b64)
html_content = html_content.replace('XPT_DL_CH_NRX_B64', _xpt_dl_ch_nrx_b64)
# NPA Market Insights download Excel base64 injection
html_content = html_content.replace('NPA_BRAND_TRX_B64', _npa_brand_trx_b64)
html_content = html_content.replace('NPA_BRAND_NBRX_B64', _npa_brand_nbrx_b64)
html_content = html_content.replace('NPA_CH_NURTEC_TRX_B64', _npa_ch_nurtec_trx_b64)
html_content = html_content.replace('NPA_CH_NURTEC_NBRX_B64', _npa_ch_nurtec_nbrx_b64)
html_content = html_content.replace('NPA_CH_UBRELVY_TRX_B64', _npa_ch_ubrelvy_trx_b64)
html_content = html_content.replace('NPA_CH_UBRELVY_NBRX_B64', _npa_ch_ubrelvy_nbrx_b64)
html_content = html_content.replace('NPA_CH_QULIPTA_TRX_B64', _npa_ch_qulipta_trx_b64)
html_content = html_content.replace('NPA_CH_QULIPTA_NBRX_B64', _npa_ch_qulipta_nbrx_b64)
# Acute/Preventive download Excel base64 injection
html_content = html_content.replace('AP_ACUTE_BRAND_B64', _acute_brand_b64)
html_content = html_content.replace('AP_PREV_BRAND_B64', _prev_brand_b64)
html_content = html_content.replace('AP_ACUTE_CH_NURTEC_B64', _acute_ch_nurtec_b64)
html_content = html_content.replace('AP_ACUTE_CH_UBRELVY_B64', _acute_ch_ubrelvy_b64)
html_content = html_content.replace('AP_PREV_CH_NURTEC_B64', _prev_ch_nurtec_b64)
html_content = html_content.replace('AP_PREV_CH_QULIPTA_B64', _prev_ch_qulipta_b64)
# Finance download Excel base64 injection
html_content = html_content.replace('FIN_GROSS_DL_B64', _fin_gross_b64)
html_content = html_content.replace('FIN_NET_DL_B64', _fin_net_b64)
# Xponent KPI grid injection (dynamic from _xpt_national)
if _xpt_national is not None:
    def _xpt_fmt_val(v):
        if v is None or (v != v): return '\u2014'
        v = float(v)
        if v >= 1e6: return f'{v/1e6:.2f}M'
        if v >= 1e3: return f'{v/1e3:.1f}K'
        return f'{int(v):,}'
    def _xpt_fmt_pct(v):
        if v is None or (v != v): return '\u2014'
        cls = 'delta-pos' if v >= 0 else 'delta-neg'
        return f'<span class="{cls}">{v:+.1f}%</span>'
    html_content = html_content.replace('XPT_NRX_MTD_VAL', _xpt_fmt_val(_xpt_national['NRX_MTD']))
    html_content = html_content.replace('XPT_NRX_MTD_STLY', _xpt_fmt_pct(_xpt_national['NRX_MTD_VS_STLY_PCT']))
    html_content = html_content.replace('XPT_NRX_YTD_VAL', _xpt_fmt_val(_xpt_national['NRX_YTD']))
    html_content = html_content.replace('XPT_NRX_YTD_STLY', _xpt_fmt_pct(_xpt_national['NRX_YTD_VS_STLY_PCT']))
    html_content = html_content.replace('XPT_TRX_MTD_VAL', _xpt_fmt_val(_xpt_national['TRX_MTD']))
    html_content = html_content.replace('XPT_TRX_MTD_STLY', _xpt_fmt_pct(_xpt_national['TRX_MTD_VS_STLY_PCT']))
    html_content = html_content.replace('XPT_TRX_YTD_VAL', _xpt_fmt_val(_xpt_national['TRX_YTD']))
    html_content = html_content.replace('XPT_TRX_YTD_STLY', _xpt_fmt_pct(_xpt_national['TRX_YTD_VS_STLY_PCT']))

# Xponent Payer & Channel table injection (dynamic from _xpt_kpi_df)
def _xpt_table_row(row, metric_prefix):
    """Build a single <tr> for the XPT table. metric_prefix is 'TRX' or 'NRX'."""
    label = row['CUT_VALUE']
    # Display name mapping
    display_names = {'MAIL_ORDER': 'Mail-Order', 'RETAIL': 'Retail', 'LTC': 'LTC',
                     'Commercial': 'Commercial', 'Medicare': 'Medicare', 'Medicaid': 'Medicaid',
                     'Other': 'Other', 'National': 'National'}
    display = display_names.get(label, label)
    wk_share = row[f'WK_{metric_prefix}_SHARE']
    wow_pp = row[f'WOW_{metric_prefix}_PP']
    mtd_share = row[f'MTD_{metric_prefix}_SHARE']
    mtd_stly = row[f'MTD_VS_STLY_{metric_prefix}_PP']
    def fmt_share(v):
        if v is None or (v != v): return '\u2014'
        return f'{v:.1f}%'
    def fmt_delta(v):
        if v is None or (v != v): return '\u2014'
        v = float(v)
        if v == 0: return '0.0%'
        cls = 'delta-pos' if v > 0 else 'delta-neg'
        return f'<span class="{cls}">{v:+.1f}%</span>'
    return f'<tr><td><strong>{display}</strong></td><td style="text-align:right">{fmt_share(wk_share)}</td><td style="text-align:right">{fmt_delta(wow_pp)}</td><td style="text-align:right">{fmt_share(mtd_share)}</td><td style="text-align:right">{fmt_delta(mtd_stly)}</td></tr>'

def _build_xpt_table(cut_type, metric_prefix):
    """Build table rows for a given cut_type (Payer/Channel) and metric (TRX/NRX)."""
    rows = []
    # National row first
    nat = _xpt_kpi_df[_xpt_kpi_df['CUT_TYPE'] == 'National']
    for _, r in nat.iterrows():
        rows.append(_xpt_table_row(r, metric_prefix))
    # Then specific cuts
    cut_df = _xpt_kpi_df[_xpt_kpi_df['CUT_TYPE'] == cut_type]
    order = ['Commercial', 'Medicare', 'Medicaid', 'Other'] if cut_type == 'Payer' else ['RETAIL', 'MAIL_ORDER', 'LTC']
    for cv in order:
        match = cut_df[cut_df['CUT_VALUE'] == cv]
        for _, r in match.iterrows():
            rows.append(_xpt_table_row(r, metric_prefix))
    return '\n'.join(rows)

html_content = html_content.replace('XPT_PAYER_TRX_TABLE_ROWS', _build_xpt_table('Payer', 'TRX'))
html_content = html_content.replace('XPT_PAYER_NRX_TABLE_ROWS', _build_xpt_table('Payer', 'NRX'))
html_content = html_content.replace('XPT_CHANNEL_TRX_TABLE_ROWS', _build_xpt_table('Channel', 'TRX'))
html_content = html_content.replace('XPT_CHANNEL_NRX_TABLE_ROWS', _build_xpt_table('Channel', 'NRX'))

# Finance KPI injection (dynamic from _fin_kpi_df)
def _inject_finance_kpis():
    """Build all finance KPI placeholder values from live stacked data."""
    replacements = {}
    # --- GROSS ---
    gross_wk = _fin_kpi_df[_fin_kpi_df['KPI_TITLE'].str.contains('LATEST_WEEK', na=False) & (_fin_kpi_df['SECTION_NAME'] == 'Gross')]
    gross_mtd = _fin_kpi_df[_fin_kpi_df['KPI_TITLE'].str.contains('MTD', na=False) & (_fin_kpi_df['SECTION_NAME'] == 'Gross')]
    gross_ytd = _fin_kpi_df[_fin_kpi_df['KPI_TITLE'].str.contains('YTD', na=False) & (_fin_kpi_df['SECTION_NAME'] == 'Gross')]

    if len(gross_wk) > 0:
        r = gross_wk.iloc[0]
        actual = r['ACTUAL_VALUE']
        py_pct = r['VARIANCE_TO_PY_PCT']
        date_str = str(r['DATA_AS_OF_DATE']).replace('"', '').strip()
        # Parse date for label
        try:
            from datetime import datetime
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            label = f"GROSS SALES W/E {dt.day} {dt.strftime('%b')}\u2019{dt.strftime('%y')}"
        except Exception:
            label = "GROSS SALES W/E"
        replacements['FIN_GROSS_WK_LABEL'] = label
        replacements['FIN_GROSS_WK_VAL'] = f'${actual/1e6:.1f}M' if actual else '\u2014'
        py_cls = 'delta-pos' if py_pct and py_pct > 0 else 'delta-neg'
        replacements['FIN_GROSS_WK_PY'] = f'<span class="{py_cls}">{py_pct:+.1f}%</span>' if py_pct else '\u2014'

    if len(gross_mtd) > 0:
        r = gross_mtd.iloc[0]
        att = r['BUDGET_ATT']
        actual = r['ACTUAL_VALUE'] or 0
        budget = r['BUDGET_VALUE'] or 0
        gap = abs(actual - budget) / 1e6
        replacements['FIN_GROSS_MTD_ATT'] = f'{att:.1f}%' if att else '\u2014'
        if actual < budget:
            replacements['FIN_GROSS_MTD_DELTA'] = f'<span class="delta-neg">${gap:.1f}M</span> behind MTD plan'
        else:
            replacements['FIN_GROSS_MTD_DELTA'] = f'<span class="delta-pos">+${gap:.1f}M</span> ahead of MTD plan'

    if len(gross_ytd) > 0:
        r = gross_ytd.iloc[0]
        att = r['BUDGET_ATT']
        actual = r['ACTUAL_VALUE'] or 0
        budget = r['BUDGET_VALUE'] or 0
        gap = abs(actual - budget) / 1e6
        replacements['FIN_GROSS_YTD_ATT'] = f'{att:.1f}%' if att else '\u2014'
        if actual < budget:
            replacements['FIN_GROSS_YTD_DELTA'] = f'<span class="delta-neg">${gap:.1f}M</span> behind YTD plan'
        else:
            replacements['FIN_GROSS_YTD_DELTA'] = f'<span class="delta-pos">+${gap:.1f}M</span> ahead of YTD plan'

    # --- NET ---
    net_month = _fin_kpi_df[_fin_kpi_df['KPI_TITLE'].str.contains('LATEST_MONTH', na=False) & (_fin_kpi_df['SECTION_NAME'] == 'Net')]
    net_ytd = _fin_kpi_df[_fin_kpi_df['KPI_TITLE'].str.contains('YTD', na=False) & (_fin_kpi_df['SECTION_NAME'] == 'Net')]

    if len(net_month) > 0:
        r = net_month.iloc[0]
        actual = r['ACTUAL_VALUE']
        py_pct = r['VARIANCE_TO_PY_PCT']
        att = r['BUDGET_ATT']
        budget = r['BUDGET_VALUE'] or 0
        date_str = str(r['DATA_AS_OF_DATE']).replace('"', '').strip()
        try:
            from datetime import datetime
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            month_num = dt.month
            label = f"NET SALES P{month_num:02d}\u2019{dt.strftime('%y')}"
        except Exception:
            label = "NET SALES LATEST MONTH"
        replacements['FIN_NET_MONTH_LABEL'] = label
        replacements['FIN_NET_MONTH_VAL'] = f'${actual/1e6:.1f}M' if actual else '\u2014'
        py_cls = 'delta-pos' if py_pct and py_pct > 0 else 'delta-neg'
        replacements['FIN_NET_MONTH_PY'] = f'<span class="{py_cls}">{py_pct:+.1f}%</span>' if py_pct else '\u2014'
        replacements['FIN_NET_MTD_ATT'] = f'{att:.1f}%' if att else '\u2014'
        gap = abs((actual or 0) - budget) / 1e6
        if (actual or 0) < budget:
            replacements['FIN_NET_MTD_DELTA'] = f'<span class="delta-neg">${gap:.1f}M</span> behind monthly plan'
        else:
            replacements['FIN_NET_MTD_DELTA'] = f'<span class="delta-pos">+${gap:.1f}M</span> ahead of monthly plan'

    if len(net_ytd) > 0:
        r = net_ytd.iloc[0]
        att = r['BUDGET_ATT']
        actual = r['ACTUAL_VALUE'] or 0
        budget = r['BUDGET_VALUE'] or 0
        gap = abs(actual - budget) / 1e6
        replacements['FIN_NET_YTD_ATT'] = f'{att:.1f}%' if att else '\u2014'
        if actual < budget:
            replacements['FIN_NET_YTD_DELTA'] = f'<span class="delta-neg">${gap:.1f}M</span> behind YTD plan'
        else:
            replacements['FIN_NET_YTD_DELTA'] = f'<span class="delta-pos">+${gap:.1f}M</span> ahead of YTD plan'

    return replacements

# Finance chart injection
_fin_replacements = _inject_finance_kpis()
if FINANCE_RESTRICTED:
    _fin_restricted_msg = '<div style="display:flex;align-items:center;justify-content:center;height:300px;color:#9ca3af;font-size:24px;font-weight:700;">\u2014</div>'
    html_content = html_content.replace('FIN_GROSS_CHART_PLACEHOLDER', _fin_restricted_msg)
    html_content = html_content.replace('FIN_NET_CHART_PLACEHOLDER', _fin_restricted_msg)
    # Blank out KPI values for restricted users
    for _k in _fin_replacements:
        html_content = html_content.replace(_k, '\u2014')
else:
    html_content = html_content.replace('FIN_GROSS_CHART_PLACEHOLDER', fin_gross_html)
    html_content = html_content.replace('FIN_NET_CHART_PLACEHOLDER', fin_net_html)
    # Inject live KPI values
    for _k, _v in _fin_replacements.items():
        html_content = html_content.replace(_k, _v)


# NPA Table dynamic injection
# Overall tables
for _brand in ['NURTEC', 'UBRELVY', 'QULIPTA']:
    for _rx, _cls_name in [('TRx', 'npa-table-trx'), ('NBRx', 'npa-table-nbrx')]:
        _key = f'{_brand}_{_rx}_OVERALL'
        _placeholder = f'NPA_OVERALL_ROWS_{_brand}_{_rx}'
        html_content = html_content.replace(_placeholder, _npa_tables.get(_key, ''))

# Acute tables
for _brand in ['NURTEC', 'UBRELVY']:
    for _rx in ['TRx', 'NBRx']:
        _key = f'{_brand}_{_rx}_ACUTE'
        _placeholder = f'NPA_ACUTE_ROWS_{_brand}_{_rx}'
        html_content = html_content.replace(_placeholder, _npa_tables.get(_key, ''))

# Preventive tables
for _brand in ['NURTEC', 'QULIPTA']:
    for _rx in ['TRx', 'NBRx']:
        _key = f'{_brand}_{_rx}_PREVENTIVE'
        _placeholder = f'NPA_PREV_ROWS_{_brand}_{_rx}'
        html_content = html_content.replace(_placeholder, _npa_tables.get(_key, ''))

# Hero KPI dynamic injection
_hero_trx_ytd = _get_npa_kpi('NURTEC', 'TRx', 'OVERALL', 'YTD', 'CURR_VALUE')
_hero_nbrx_ytd = _get_npa_kpi('NURTEC', 'NBRx', 'OVERALL', 'YTD', 'CURR_VALUE')
_hero_trx_growth = _get_npa_kpi('NURTEC', 'TRx', 'OVERALL', 'YTD', 'GROWTH_PCT')
_hero_nbrx_growth = _get_npa_kpi('NURTEC', 'NBRx', 'OVERALL', 'YTD', 'GROWTH_PCT')
_hero_overall_ms = _get_npa_kpi('NURTEC', 'TRx', 'OVERALL', 'YTD', 'MARKET_SHARE_PCT')
_hero_acute_ms = _get_npa_kpi('NURTEC', 'TRx', 'ACUTE', 'YTD', 'MARKET_SHARE_PCT')
_hero_prev_ms = _get_npa_kpi('NURTEC', 'TRx', 'PREVENTIVE', 'YTD', 'MARKET_SHARE_PCT')
# Previous year market shares for delta calculation
_hero_overall_ms_py = _npa_stacked_df[(_npa_stacked_df['BRAND']=='NURTEC')&(_npa_stacked_df['PRESCRIPTION']=='TRx')&(_npa_stacked_df['RX_CLASSIFICATION']=='OVERALL')&(_npa_stacked_df['ROW_LABEL']=="Actuals '25")&(_npa_stacked_df['TIME_PERIOD']=='YTD')]
_hero_overall_ms_py = _hero_overall_ms_py.iloc[0]['MARKET_SHARE_PCT'] if len(_hero_overall_ms_py) > 0 else 0
_hero_acute_ms_py = _npa_stacked_df[(_npa_stacked_df['BRAND']=='NURTEC')&(_npa_stacked_df['PRESCRIPTION']=='TRx')&(_npa_stacked_df['RX_CLASSIFICATION']=='ACUTE')&(_npa_stacked_df['ROW_LABEL']=="Actuals '25")&(_npa_stacked_df['TIME_PERIOD']=='YTD')]
_hero_acute_ms_py = _hero_acute_ms_py.iloc[0]['MARKET_SHARE_PCT'] if len(_hero_acute_ms_py) > 0 else 0
_hero_prev_ms_py = _npa_stacked_df[(_npa_stacked_df['BRAND']=='NURTEC')&(_npa_stacked_df['PRESCRIPTION']=='TRx')&(_npa_stacked_df['RX_CLASSIFICATION']=='PREVENTIVE')&(_npa_stacked_df['ROW_LABEL']=="Actuals '25")&(_npa_stacked_df['TIME_PERIOD']=='YTD')]
_hero_prev_ms_py = _hero_prev_ms_py.iloc[0]['MARKET_SHARE_PCT'] if len(_hero_prev_ms_py) > 0 else 0

html_content = html_content.replace('HERO_TRX_YTD_VAL', f'{_hero_trx_ytd/1e6:.2f}M' if _hero_trx_ytd else '—')
html_content = html_content.replace('HERO_NBRX_YTD_VAL', f'{int(_hero_nbrx_ytd/1000)}K' if _hero_nbrx_ytd else '—')
html_content = html_content.replace('HERO_TRX_GROWTH_VAL', f'+{_hero_trx_growth:.1f}%' if _hero_trx_growth else '—')
html_content = html_content.replace('HERO_NBRX_GROWTH_VAL', f'+{_hero_nbrx_growth:.1f}%' if _hero_nbrx_growth else '—')
html_content = html_content.replace('HERO_OVERALL_MS_VAL', f'{_hero_overall_ms:.1f}%' if _hero_overall_ms else '—')
html_content = html_content.replace('HERO_ACUTE_MS_VAL', f'{_hero_acute_ms:.1f}%' if _hero_acute_ms else '—')
html_content = html_content.replace('HERO_PREV_MS_VAL', f'{_hero_prev_ms:.1f}%' if _hero_prev_ms else '—')
html_content = html_content.replace('HERO_OVERALL_MS_DELTA', f'+{_hero_overall_ms - _hero_overall_ms_py:.1f}%' if _hero_overall_ms and _hero_overall_ms_py else '—')
html_content = html_content.replace('HERO_ACUTE_MS_DELTA', f'+{_hero_acute_ms - _hero_acute_ms_py:.1f}%' if _hero_acute_ms and _hero_acute_ms_py else '—')
_prev_delta = _hero_prev_ms - _hero_prev_ms_py if _hero_prev_ms and _hero_prev_ms_py else 0
html_content = html_content.replace('HERO_PREV_MS_DELTA', f'{_prev_delta:+.1f}%' if _hero_prev_ms else '—')

# Executive Summary KPI dynamic injection
_exec_nbrx_wk = _get_npa_kpi('NURTEC', 'NBRx', 'OVERALL', 'Latest Week', 'CURR_VALUE')
_exec_trx_wk = _get_npa_kpi('NURTEC', 'TRx', 'OVERALL', 'Latest Week', 'CURR_VALUE')
_exec_ms_wk = _get_npa_kpi('NURTEC', 'TRx', 'OVERALL', 'Latest Week', 'MARKET_SHARE_PCT')
_exec_nbrx_growth = _get_npa_kpi('NURTEC', 'NBRx', 'OVERALL', 'Latest Week', 'GROWTH_PCT')
_exec_trx_growth = _get_npa_kpi('NURTEC', 'TRx', 'OVERALL', 'Latest Week', 'GROWTH_PCT')
_exec_goal_att = _get_npa_kpi('NURTEC', 'TRx', 'OVERALL', 'YTD', 'GOAL_ATTAINMENT_PCT')
_exec_ms_py = _npa_stacked_df[(_npa_stacked_df['BRAND']=='NURTEC')&(_npa_stacked_df['PRESCRIPTION']=='TRx')&(_npa_stacked_df['RX_CLASSIFICATION']=='OVERALL')&(_npa_stacked_df['ROW_LABEL']=="Actuals '25")&(_npa_stacked_df['TIME_PERIOD']=='Latest Week')]
_exec_ms_py_val = _exec_ms_py.iloc[0]['MARKET_SHARE_PCT'] if len(_exec_ms_py) > 0 else 0
_exec_ms_delta = _exec_ms_wk - _exec_ms_py_val if _exec_ms_wk and _exec_ms_py_val else 0
_exec_fin = _fin_kpi_df[_fin_kpi_df['SECTION_NAME'] == 'Gross']
_exec_fin_wk = _exec_fin[_exec_fin['KPI_TITLE'].str.contains('LATEST_WEEK', na=False)]
_exec_gross_val = _exec_fin_wk.iloc[0]['ACTUAL_VALUE'] if len(_exec_fin_wk) > 0 else None
_exec_gross_py_pct = _exec_fin_wk.iloc[0]['VARIANCE_TO_PY_PCT'] if len(_exec_fin_wk) > 0 else None

html_content = html_content.replace('EXEC_NBRX_WK_VAL', f'{_exec_nbrx_wk/1000:.1f}K' if _exec_nbrx_wk else '—')
html_content = html_content.replace('EXEC_TRX_WK_VAL', f'{_exec_trx_wk/1000:.1f}K' if _exec_trx_wk else '—')
html_content = html_content.replace('EXEC_MS_WK_VAL', f'{_exec_ms_wk:.1f}%' if _exec_ms_wk else '—')
html_content = html_content.replace('EXEC_NBRX_WK_GROWTH', f'+{_exec_nbrx_growth:.1f}%' if _exec_nbrx_growth and _exec_nbrx_growth > 0 else f'{_exec_nbrx_growth:.1f}%' if _exec_nbrx_growth else '—')
html_content = html_content.replace('EXEC_TRX_WK_GROWTH', f'+{_exec_trx_growth:.1f}%' if _exec_trx_growth and _exec_trx_growth > 0 else f'{_exec_trx_growth:.1f}%' if _exec_trx_growth else '—')
html_content = html_content.replace('EXEC_MS_WK_GROWTH', f'{_exec_ms_delta:+.1f}%' if _exec_ms_wk else '—')
html_content = html_content.replace('EXEC_GROSS_WK_VAL', f'${_exec_gross_val/1e6:.1f}M' if _exec_gross_val else '—')
html_content = html_content.replace('EXEC_GROSS_VS_PY', f'+{_exec_gross_py_pct:.1f}%' if _exec_gross_py_pct and _exec_gross_py_pct > 0 else f'{_exec_gross_py_pct:.1f}%' if _exec_gross_py_pct else '—')
html_content = html_content.replace('EXEC_GOAL_ATT_VAL', f'{_exec_goal_att:.1f}%' if _exec_goal_att and not (_exec_goal_att != _exec_goal_att) else '—')

# Dynamic week ending date from latest NPA trends data
_latest_week_id = str(int(npa_brand_df['WEEK_ID'].max()))
_exec_week_date = f'{_latest_week_id[4:6]}/{_latest_week_id[6:8]}/{_latest_week_id[0:4]}' if len(_latest_week_id) == 8 else '\u2014'
html_content = html_content.replace('EXEC_PERF_WEEK_DATE', _exec_week_date)

# --- Performance Insights dynamic bullets ---
def _li(text):
    return f'<li style="margin-bottom:8px;font-size:13px;line-height:1.7;"><span style="color:#374151;">{text}</span></li>'

def _b(v):
    return f'<strong style="color:#0000C9;">{v}</strong>'

def _wow_dir(pct):
    return 'up' if pct >= 0 else 'down'

def _build_weekly_insights():
    bullets = []
    latest_wk = npa_brand_df['WEEK_ID'].max()
    all_weeks = sorted(npa_brand_df['WEEK_ID'].unique(), reverse=True)
    prev_wk = all_weeks[1] if len(all_weeks) > 1 else latest_wk

    # Bullet 1: Nurtec NBRx volume, WoW, vs STLY
    nbrx_curr = nbrx_brand_df[(nbrx_brand_df['BRAND']=='NURTEC')&(nbrx_brand_df['WEEK_ID']==latest_wk)]['ACTUALS'].sum()
    nbrx_prev = nbrx_brand_df[(nbrx_brand_df['BRAND']=='NURTEC')&(nbrx_brand_df['WEEK_ID']==prev_wk)]['ACTUALS'].sum()
    nbrx_stly = nbrx_brand_df[(nbrx_brand_df['BRAND']=='NURTEC')&(nbrx_brand_df['WEEK_ID']==latest_wk)]['STLY'].sum()
    nbrx_wow = (nbrx_curr - nbrx_prev) / nbrx_prev * 100 if nbrx_prev else 0
    nbrx_vs_stly = (nbrx_curr - nbrx_stly) / nbrx_stly * 100 if nbrx_stly else 0
    bullets.append(_li(
        f'Nurtec NBRx volume this week was {_b(f"{nbrx_curr/1000:.1f}K")}, '
        f'{_wow_dir(nbrx_wow)} {_b(f"{nbrx_wow:+.1f}%")} vs. prior week ({nbrx_prev/1000:.1f}K) '
        f'and {_b(f"{nbrx_vs_stly:+.1f}%")} vs. same week last year.'
    ))

    # Bullet 2: Nurtec TRx volume, WoW, retail share delta
    trx_curr = npa_brand_df[(npa_brand_df['BRAND']=='NURTEC')&(npa_brand_df['WEEK_ID']==latest_wk)]['ACTUALS'].sum()
    trx_prev = npa_brand_df[(npa_brand_df['BRAND']=='NURTEC')&(npa_brand_df['WEEK_ID']==prev_wk)]['ACTUALS'].sum()
    trx_wow = (trx_curr - trx_prev) / trx_prev * 100 if trx_prev else 0
    # Retail share delta vs STLY
    ret_nurtec_curr = load_channel_data("TRx", "NURTEC")
    ret_nurtec_wk = ret_nurtec_curr[(ret_nurtec_curr['CHANNEL_TYPE']=='Retail')&(ret_nurtec_curr['WEEK_ID']==latest_wk)]
    r_n_curr = ret_nurtec_wk['ACTUALS'].sum() if len(ret_nurtec_wk) > 0 else 0
    r_n_stly = ret_nurtec_wk['STLY'].sum() if len(ret_nurtec_wk) > 0 else 0
    ret_total_curr = r_n_curr
    ret_total_stly = r_n_stly
    for _br in ['UBRELVY', 'QULIPTA']:
        _ch = load_channel_data("TRx", _br)
        _wk = _ch[(_ch['CHANNEL_TYPE']=='Retail')&(_ch['WEEK_ID']==latest_wk)]
        ret_total_curr += _wk['ACTUALS'].sum() if len(_wk) > 0 else 0
        ret_total_stly += _wk['STLY'].sum() if len(_wk) > 0 else 0
    ret_share_curr = (r_n_curr / ret_total_curr * 100) if ret_total_curr else 0
    ret_share_stly = (r_n_stly / ret_total_stly * 100) if ret_total_stly else 0
    ret_share_delta = ret_share_curr - ret_share_stly
    delta_word = 'gain' if ret_share_delta >= 0 else 'loss'
    bullets.append(_li(
        f'Nurtec TRx volume came in at {_b(f"{trx_curr/1000:.1f}K")}, '
        f'{"an increase" if trx_wow >= 0 else "a decrease"} of {_b(f"{trx_wow:+.1f}%")} WoW, '
        f'with retail channel contributing {_b(f"{ret_share_delta:+.1f}%")} share {delta_word} vs. same week last year.'
    ))

    # Bullet 3: oCGRP TRx market shares all brands
    def _wk_share(brand):
        c = npa_brand_df[(npa_brand_df['BRAND']==brand)&(npa_brand_df['WEEK_ID']==latest_wk)]['ACTUALS'].sum()
        t = npa_brand_df[npa_brand_df['WEEK_ID']==latest_wk]['ACTUALS'].sum()
        return (c / t * 100) if t else 0
    def _prev_share(brand):
        c = npa_brand_df[(npa_brand_df['BRAND']==brand)&(npa_brand_df['WEEK_ID']==prev_wk)]['ACTUALS'].sum()
        t = npa_brand_df[npa_brand_df['WEEK_ID']==prev_wk]['ACTUALS'].sum()
        return (c / t * 100) if t else 0
    n_share = _wk_share('NURTEC'); n_wow_s = n_share - _prev_share('NURTEC')
    u_share = _wk_share('UBRELVY'); u_wow_s = u_share - _prev_share('UBRELVY')
    q_share = _wk_share('QULIPTA'); q_wow_s = q_share - _prev_share('QULIPTA')
    u_desc = f'flat at {_b(f"{u_share:.1f}%")}' if abs(u_wow_s) < 0.1 else f'{"gained" if u_wow_s > 0 else "lost"} {_b(f"{u_wow_s:+.1f}%")} to {_b(f"{u_share:.1f}%")}'
    q_desc = f'flat at {_b(f"{q_share:.1f}%")}' if abs(q_wow_s) < 0.1 else f'{"gained" if q_wow_s > 0 else "lost"} {_b(f"{q_wow_s:+.1f}%")} to {_b(f"{q_share:.1f}%")}'
    bullets.append(_li(
        f'Nurtec oCGRP TRx market share moved to {_b(f"{n_share:.1f}%")} this week, '
        f'{_wow_dir(n_wow_s)} {_b(f"{n_wow_s:+.1f}%")} WoW. Ubrelvy {u_desc}; Qulipta {q_desc}.'
    ))

    # Bullet 4: Acute & Preventive share
    acute_curr = _get_npa_kpi('NURTEC', 'TRx', 'ACUTE', 'Latest Week', 'MARKET_SHARE_PCT') or 0
    acute_py = _npa_stacked_df[(_npa_stacked_df['BRAND']=='NURTEC')&(_npa_stacked_df['PRESCRIPTION']=='TRx')&(_npa_stacked_df['RX_CLASSIFICATION']=='ACUTE')&(_npa_stacked_df['ROW_LABEL']=="Actuals '25")&(_npa_stacked_df['TIME_PERIOD']=='Latest Week')]
    acute_delta = acute_curr - (acute_py.iloc[0]['MARKET_SHARE_PCT'] if len(acute_py) > 0 else 0)
    prev_curr = _get_npa_kpi('NURTEC', 'TRx', 'PREVENTIVE', 'Latest Week', 'MARKET_SHARE_PCT') or 0
    prev_py = _npa_stacked_df[(_npa_stacked_df['BRAND']=='NURTEC')&(_npa_stacked_df['PRESCRIPTION']=='TRx')&(_npa_stacked_df['RX_CLASSIFICATION']=='PREVENTIVE')&(_npa_stacked_df['ROW_LABEL']=="Actuals '25")&(_npa_stacked_df['TIME_PERIOD']=='Latest Week')]
    prev_delta = prev_curr - (prev_py.iloc[0]['MARKET_SHARE_PCT'] if len(prev_py) > 0 else 0)
    bullets.append(_li(
        f'Acute oCGRP share for Nurtec at {_b(f"{acute_curr:.1f}%")} ({_b(f"{acute_delta:+.1f}%")} WoW); '
        f'Preventive share at {_b(f"{prev_curr:.1f}%")} ({_b(f"{prev_delta:+.1f}%")} WoW).'
    ))

    # Bullet 5: Gross sales
    gross_wk_row = _fin_kpi_df[_fin_kpi_df['KPI_TITLE'].str.contains('LATEST_WEEK', na=False) & (_fin_kpi_df['SECTION_NAME'] == 'Gross')]
    gross_mtd_row = _fin_kpi_df[_fin_kpi_df['KPI_TITLE'].str.contains('MTD', na=False) & (_fin_kpi_df['SECTION_NAME'] == 'Gross')]
    if len(gross_wk_row) > 0:
        gw = gross_wk_row.iloc[0]
        g_actual = gw['ACTUAL_VALUE'] / 1e6
        g_py_pct = gw['VARIANCE_TO_PY_PCT']
        mtd_att = gross_mtd_row.iloc[0]['BUDGET_ATT'] if len(gross_mtd_row) > 0 else 0
        bullets.append(_li(
            f'Gross sales for the week were {_b(f"${g_actual:.1f}M")}, '
            f'{"up" if g_py_pct >= 0 else "down"} {_b(f"{g_py_pct:+.1f}%")} vs. prior year. '
            f'MTD attainment at {_b(f"{mtd_att:.1f}%")} vs. OP26.'
        ))

    # Bullet 6: Preventive segment gap
    prev_brand_df = load_acute_prev_brand_data("TRx", "Preventive")
    q_p_curr = prev_brand_df[(prev_brand_df['BRAND']=='QULIPTA')&(prev_brand_df['WEEK_ID']==latest_wk)]['ACTUALS'].sum()
    q_p_prev = prev_brand_df[(prev_brand_df['BRAND']=='QULIPTA')&(prev_brand_df['WEEK_ID']==prev_wk)]['ACTUALS'].sum()
    n_p_curr = prev_brand_df[(prev_brand_df['BRAND']=='NURTEC')&(prev_brand_df['WEEK_ID']==latest_wk)]['ACTUALS'].sum()
    n_p_prev = prev_brand_df[(prev_brand_df['BRAND']=='NURTEC')&(prev_brand_df['WEEK_ID']==prev_wk)]['ACTUALS'].sum()
    q_p_wow = (q_p_curr - q_p_prev) / q_p_prev * 100 if q_p_prev else 0
    n_p_wow = (n_p_curr - n_p_prev) / n_p_prev * 100 if n_p_prev else 0
    gap = abs(q_p_wow - n_p_wow)
    gap_word = 'widening' if q_p_wow > n_p_wow else 'narrowing'
    bullets.append(_li(
        f'Preventive segment: Qulipta weekly volume {"grew" if q_p_wow >= 0 else "fell"} {_b(f"{q_p_wow:+.1f}%")} WoW '
        f'vs. Nurtec Preventive at {_b(f"{n_p_wow:+.1f}%")} WoW '
        f'&#8212; gap {gap_word} by {_b(f"{gap:.1f}%")} this week.'
    ))

    return '\n'.join(bullets)

def _build_ytd_insights():
    bullets = []
    def _ytd(brand, presc, field):
        return _get_npa_kpi(brand, presc, 'OVERALL', 'YTD', field)
    def _ytd_py_share(brand, presc):
        r = _npa_stacked_df[(_npa_stacked_df['BRAND']==brand)&(_npa_stacked_df['PRESCRIPTION']==presc)&(_npa_stacked_df['RX_CLASSIFICATION']=='OVERALL')&(_npa_stacked_df['ROW_LABEL']=="Actuals '25")&(_npa_stacked_df['TIME_PERIOD']=='YTD')]
        return r.iloc[0]['MARKET_SHARE_PCT'] if len(r) > 0 else 0
    def _ytd_py_growth(brand, presc):
        r = _npa_stacked_df[(_npa_stacked_df['BRAND']==brand)&(_npa_stacked_df['PRESCRIPTION']==presc)&(_npa_stacked_df['RX_CLASSIFICATION']=='OVERALL')&(_npa_stacked_df['ROW_LABEL']=="Actuals '25")&(_npa_stacked_df['TIME_PERIOD']=='YTD')]
        return r.iloc[0]['GROWTH_PCT'] if len(r) > 0 else 0

    # Bullet 1: oCGRP market growth
    n_growth = _ytd('NURTEC', 'TRx', 'GROWTH_PCT') or 0
    total_curr = npa_brand_df['ACTUALS'].sum()
    total_stly = npa_brand_df['STLY'].sum()
    mkt_growth = (total_curr - total_stly) / total_stly * 100 if total_stly else 0
    n_py_g = _ytd_py_growth('NURTEC', 'TRx')
    u_py_g = _ytd_py_growth('UBRELVY', 'TRx')
    q_py_g = _ytd_py_growth('QULIPTA', 'TRx')
    mkt_py_growth = (n_py_g * 0.43 + u_py_g * 0.33 + q_py_g * 0.24) if (n_py_g and u_py_g and q_py_g) else 0
    mkt_diff = mkt_growth - mkt_py_growth
    from datetime import datetime as _dtf
    try:
        dt = _dtf.strptime(_latest_week_id, '%Y%m%d')
        date_fmt = dt.strftime('%B %d, %Y')
    except:
        date_fmt = _exec_week_date
    bullets.append(_li(
        f'As of {_b(date_fmt)}, the oral CGRP TRx market is growing at {_b(f"{mkt_growth:.1f}% YTD")} vs. same time last year, '
        f'which is {_b(f"{abs(mkt_diff):.1f}%")} {"lower" if mkt_diff < 0 else "higher"} growth than the same time period in 2025 vs. 2024 ({mkt_py_growth:.1f}%).'
    ))

    # Bullet 2: Nurtec TRx YTD growth
    n_diff = n_growth - n_py_g if n_py_g else 0
    bullets.append(_li(
        f'Nurtec TRx is growing at {_b(f"{n_growth:.1f}% YTD")}, which is {_b(f"{abs(n_diff):.1f}%")} '
        f'{"higher" if n_diff >= 0 else "lower"} than Nurtec\'s growth in 2025 by this point in the year ({n_py_g:.1f}%).'
    ))

    # Bullet 3: Ubrelvy & Qulipta YTD growth
    u_growth = _ytd('UBRELVY', 'TRx', 'GROWTH_PCT') or 0
    q_growth = _ytd('QULIPTA', 'TRx', 'GROWTH_PCT') or 0
    u_diff = u_growth - u_py_g if u_py_g else 0
    q_diff = q_growth - q_py_g if q_py_g else 0
    bullets.append(_li(
        f"Comparatively, Ubrelvy's YTD growth is {_b(f'{u_growth:.1f}%')} ({abs(u_diff):.1f}% {'lower' if u_diff < 0 else 'higher'} than in 2025), "
        f"and Qulipta's is {_b(f'{q_growth:.1f}%')} ({abs(q_diff):.1f}% {'lower' if q_diff < 0 else 'higher'} than in 2025)."
    ))

    # Bullet 4: Nurtec share leader
    n_trx_s = _ytd('NURTEC', 'TRx', 'MARKET_SHARE_PCT') or 0
    n_nbrx_s = _ytd('NURTEC', 'NBRx', 'MARKET_SHARE_PCT') or 0
    n_trx_d = n_trx_s - _ytd_py_share('NURTEC', 'TRx')
    n_nbrx_d = n_nbrx_s - _ytd_py_share('NURTEC', 'NBRx')
    bullets.append(_li(
        f'Nurtec remains the market leader in Oral CGRP with {_b(f"{n_trx_s:.1f}% TRx share YTD")} '
        f'and {_b(f"{n_nbrx_s:.1f}% NBRx share YTD")} '
        f'({n_trx_d:+.1f}% TRx share and {n_nbrx_d:+.1f}% NBRx share vs. same period last year).'
    ))

    # Bullet 5: Ubrelvy shares
    u_trx_s = _ytd('UBRELVY', 'TRx', 'MARKET_SHARE_PCT') or 0
    u_nbrx_s = _ytd('UBRELVY', 'NBRx', 'MARKET_SHARE_PCT') or 0
    u_trx_d = u_trx_s - _ytd_py_share('UBRELVY', 'TRx')
    u_nbrx_d = u_nbrx_s - _ytd_py_share('UBRELVY', 'NBRx')
    bullets.append(_li(
        f'Ubrelvy has {_b(f"{u_trx_s:.1f}%")} TRx share YTD and {_b(f"{u_nbrx_s:.1f}%")} NBRx share YTD '
        f'({u_trx_d:+.1f}% and {u_nbrx_d:+.1f}% respectively vs. same time last year).'
    ))

    # Bullet 6: Qulipta shares
    q_trx_s = _ytd('QULIPTA', 'TRx', 'MARKET_SHARE_PCT') or 0
    q_nbrx_s = _ytd('QULIPTA', 'NBRx', 'MARKET_SHARE_PCT') or 0
    q_trx_d = q_trx_s - _ytd_py_share('QULIPTA', 'TRx')
    q_nbrx_d = q_nbrx_s - _ytd_py_share('QULIPTA', 'NBRx')
    bullets.append(_li(
        f'Qulipta has {_b(f"{q_trx_s:.1f}%")} TRx share YTD and {_b(f"{q_nbrx_s:.1f}%")} NBRx share YTD '
        f'({q_trx_d:+.1f}% and {q_nbrx_d:+.1f}% respectively vs. same time last year).'
    ))

    # Bullet 7: Budget attainment
    n_goal = _ytd('NURTEC', 'TRx', 'GOAL_ATTAINMENT_PCT') or 0
    # oCGRP attainment: only use brands that have non-null goal attainment
    ocgrp_rows = _npa_stacked_df[(_npa_stacked_df['PRESCRIPTION']=='TRx')&(_npa_stacked_df['RX_CLASSIFICATION']=='OVERALL')&(_npa_stacked_df['ROW_LABEL']=="Actuals '26")&(_npa_stacked_df['TIME_PERIOD']=='YTD')&(_npa_stacked_df['GOAL_ATTAINMENT_PCT'].notna())]
    # If only Nurtec has goal data, don't show oCGRP (it would be same number)
    ocgrp_brands_with_goal = ocgrp_rows['BRAND'].nunique() if len(ocgrp_rows) > 0 else 0
    if ocgrp_brands_with_goal > 1:
        ocgrp_att = ocgrp_rows['GOAL_ATTAINMENT_PCT'].mean()
        bullets.append(_li(
            f'vs Budget: Nurtec TRx attainment is {_b(f"{n_goal:.1f}% YTD")}; '
            f'oCGRP TRx attainment is {_b(f"{ocgrp_att:.1f}% YTD")}.'
        ))
    else:
        bullets.append(_li(
            f'vs Budget: Nurtec TRx attainment is {_b(f"{n_goal:.1f}% YTD")}.'
        ))

    return '\n'.join(bullets)

html_content = html_content.replace('EXEC_WEEKLY_INSIGHTS_BULLETS', _build_weekly_insights())
html_content = html_content.replace('EXEC_YTD_INSIGHTS_BULLETS', _build_ytd_insights())

def _calc_r4w_avg(prescription):
    """Calculate Rolling 4-Week Average from NPA trends weekly data."""
    src_df = nbrx_brand_df if prescription == 'NBRx' else npa_brand_df
    nurtec_wk = src_df[src_df['BRAND'] == 'NURTEC'].sort_values('WEEK_ID', ascending=False)
    last4 = nurtec_wk.head(4)['ACTUALS'].dropna()
    return last4.mean() if len(last4) > 0 else None

def _calc_r4w_share(prescription):
    """Calculate R4W average share: avg of (Nurtec / oCGRP) for last 4 weeks."""
    src_df = nbrx_brand_df if prescription == 'NBRx' else npa_brand_df
    last4_weeks = sorted(src_df['WEEK_ID'].unique(), reverse=True)[:4]
    shares = []
    for wk in last4_weeks:
        wk_df = src_df[src_df['WEEK_ID'] == wk]
        nurtec_vol = wk_df[wk_df['BRAND'] == 'NURTEC']['ACTUALS'].sum()
        total = wk_df['ACTUALS'].sum()
        if total > 0:
            shares.append(nurtec_vol / total * 100)
    return sum(shares) / len(shares) if shares else None

def _calc_nrx_share_weekly():
    """Calculate NRx share as Nurtec NBRx / (Nurtec + Ubrelvy + Qulipta NBRx) for latest week."""
    latest_wk = nbrx_brand_df['WEEK_ID'].max()
    wk_df = nbrx_brand_df[nbrx_brand_df['WEEK_ID'] == latest_wk]
    nurtec_vol = wk_df[wk_df['BRAND'] == 'NURTEC']['ACTUALS'].sum()
    total = wk_df['ACTUALS'].sum()
    return (nurtec_vol / total * 100) if total > 0 else None

def _calc_nrx_share_ytd():
    """Calculate NRx share YTD as Nurtec NBRx / oCGRP NBRx total across all weeks."""
    nurtec_ytd = nbrx_brand_df[nbrx_brand_df['BRAND'] == 'NURTEC']['ACTUALS'].sum()
    total = nbrx_brand_df['ACTUALS'].sum()
    return (nurtec_ytd / total * 100) if total > 0 else None

def _calc_nrx_share_stly():
    """Calculate NRx share VS STLY: current YTD share - prior year YTD share (using STLY column)."""
    nurtec_ytd = nbrx_brand_df[nbrx_brand_df['BRAND'] == 'NURTEC']['ACTUALS'].sum()
    total_ytd = nbrx_brand_df['ACTUALS'].sum()
    nurtec_py = nbrx_brand_df[nbrx_brand_df['BRAND'] == 'NURTEC']['STLY'].sum()
    total_py = nbrx_brand_df['STLY'].sum()
    curr_share = (nurtec_ytd / total_ytd * 100) if total_ytd > 0 else None
    py_share = (nurtec_py / total_py * 100) if total_py > 0 else None
    if curr_share is not None and py_share is not None:
        return curr_share - py_share
    return None

def _build_perf_snapshot():
    rows = []
    for label, prescription in [('Nurtec NBRx', 'NBRx'), ('Nurtec TRx', 'TRx')]:
        wk_val = _get_npa_kpi('NURTEC', prescription, 'OVERALL', 'Latest Week', 'CURR_VALUE')
        r4w_val = _calc_r4w_avg(prescription)
        ytd_val = _get_npa_kpi('NURTEC', prescription, 'OVERALL', 'YTD', 'CURR_VALUE')
        goal_att = _get_npa_kpi('NURTEC', prescription, 'OVERALL', 'YTD', 'GOAL_ATTAINMENT_PCT')
        growth = _get_npa_kpi('NURTEC', prescription, 'OVERALL', 'YTD', 'GROWTH_PCT')
        wk_fmt = f'{wk_val/1000:.1f}K' if wk_val else '\u2014'
        r4w_fmt = f'{r4w_val/1000:.1f}K' if r4w_val else '\u2014'
        ytd_fmt = f'{ytd_val/1e6:.2f}M' if ytd_val and ytd_val >= 1e6 else f'{int(ytd_val):,}' if ytd_val else '\u2014'
        goal_fmt = f'{goal_att:.1f}%' if goal_att and not (goal_att != goal_att) else '\u2014'
        growth_fmt = _fmt_pct(growth) if growth and not (growth != growth) else '\u2014'
        rows.append(f'<tr><td><strong>{label}</strong></td><td style="text-align:right">{wk_fmt}</td><td style="text-align:right">{r4w_fmt}</td><td style="text-align:right">{ytd_fmt}</td><td style="text-align:right">{goal_fmt}</td><td style="text-align:right">{growth_fmt}</td></tr>')
    for label, prescription in [('Nurtec NBRx Share', 'NBRx'), ('Nurtec TRx Share', 'TRx')]:
        wk_ms = _get_npa_kpi('NURTEC', prescription, 'OVERALL', 'Latest Week', 'MARKET_SHARE_PCT')
        r4w_ms = _calc_r4w_share(prescription)
        ytd_ms = _get_npa_kpi('NURTEC', prescription, 'OVERALL', 'YTD', 'MARKET_SHARE_PCT')
        # Delta vs STLY
        py_ms = _npa_stacked_df[(_npa_stacked_df['BRAND']=='NURTEC')&(_npa_stacked_df['PRESCRIPTION']==prescription)&(_npa_stacked_df['RX_CLASSIFICATION']=='OVERALL')&(_npa_stacked_df['ROW_LABEL']=="Actuals '25")&(_npa_stacked_df['TIME_PERIOD']=='YTD')]
        py_ms_val = py_ms.iloc[0]['MARKET_SHARE_PCT'] if len(py_ms) > 0 else 0
        delta = ytd_ms - py_ms_val if ytd_ms and py_ms_val else None
        wk_fmt = f'{wk_ms:.1f}%' if wk_ms else '\u2014'
        r4w_fmt = f'{r4w_ms:.1f}%' if r4w_ms else '\u2014'
        ytd_fmt = f'{ytd_ms:.1f}%' if ytd_ms else '\u2014'
        delta_fmt = _fmt_pct(delta) if delta is not None else '\u2014'
        rows.append(f'<tr><td><strong>{label}</strong></td><td style="text-align:right">{wk_fmt}</td><td style="text-align:right">{r4w_fmt}</td><td style="text-align:right">{ytd_fmt}</td><td style="text-align:right">\u2014</td><td style="text-align:right">{delta_fmt}</td></tr>')
    # NRx Share row: Nurtec NBRx / (Nurtec + Ubrelvy + Qulipta) NBRx
    nrx_wk = _calc_nrx_share_weekly()
    nrx_r4w = _calc_r4w_share('NBRx')
    nrx_ytd = _calc_nrx_share_ytd()
    nrx_stly = _calc_nrx_share_stly()
    nrx_wk_fmt = f'{nrx_wk:.1f}%' if nrx_wk else '\u2014'
    nrx_r4w_fmt = f'{nrx_r4w:.1f}%' if nrx_r4w else '\u2014'
    nrx_ytd_fmt = f'{nrx_ytd:.1f}%' if nrx_ytd else '\u2014'
    nrx_stly_fmt = _fmt_pct(nrx_stly) if nrx_stly is not None else '\u2014'
    rows.append(f'<tr><td><strong>Nurtec NRx Share</strong></td><td style="text-align:right">{nrx_wk_fmt}</td><td style="text-align:right">{nrx_r4w_fmt}</td><td style="text-align:right">{nrx_ytd_fmt}</td><td style="text-align:right">\u2014</td><td style="text-align:right">{nrx_stly_fmt}</td></tr>')
    return '\n'.join(rows)

html_content = html_content.replace('EXEC_PERF_SNAPSHOT_ROWS', _build_perf_snapshot())

components.html(html_content, height=960, scrolling=False)
